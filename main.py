import sys
import re
import json
import os
import shutil
import calendar
import zipfile
from datetime import date
from pathlib import Path

DATA_DIR = "data"
import pandas as pd

from core import energy
from core.utils import fmt_money
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout,
    QListWidget, QListWidgetItem, QStackedWidget, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QLineEdit, QComboBox,
    QDateEdit, QFrame, QFileDialog, QMessageBox, QMenu, QInputDialog, QDialog,
    QCheckBox, QSpinBox,
    QGraphicsView, QGraphicsScene, QGraphicsEllipseItem, QGraphicsTextItem,
    QFormLayout, QDialogButtonBox, QScrollArea, QSizePolicy,
    QStyleOption, QStyle,
)
from PyQt6.QtCore import Qt, QDate, QPoint, QRectF, pyqtSignal
from PyQt6.QtGui import QFont, QFontMetrics, QColor, QAction, QPainter, QPixmap, QPen, QFontDatabase, QPalette, QBitmap, QPainterPath


from ui.categorization import CATEGORY_COLORS, ALL_CATEGORIES, categorize_row, apply_categorization


from ui.plot_detection import get_plot, apply_plot_column, _PLOTS_FILE


def _ensure_fonts():
    """Download Material Icons and Roboto Slab from GitHub if not present."""
    import urllib.request
    fonts_dir = Path(__file__).parent / "resources" / "fonts"
    fonts_dir.mkdir(parents=True, exist_ok=True)
    fonts = {
        "MaterialIcons-Regular.ttf": (
            "https://github.com/google/material-design-icons"
            "/raw/master/font/MaterialIcons-Regular.ttf"
        ),
        "RobotoSlab-Regular.ttf": (
            "https://github.com/googlefonts/robotoslab"
            "/raw/main/fonts/ttf/RobotoSlab-Regular.ttf"
        ),
        "RobotoSlab-Bold.ttf": (
            "https://github.com/googlefonts/robotoslab"
            "/raw/main/fonts/ttf/RobotoSlab-Bold.ttf"
        ),
    }
    for filename, url in fonts.items():
        dest = fonts_dir / filename
        if not dest.exists():
            try:
                urllib.request.urlretrieve(url, str(dest))
            except Exception:
                pass


# ======================================================================= #
#  ВКЛАДКА ДЕТАЛИЗАЦИЯ
# ======================================================================= #

class _SortItem(QTableWidgetItem):
    """QTableWidgetItem с корректной числовой и датовой сортировкой."""
    _DATE_FMT = "%d.%m.%Y"

    def __lt__(self, other: "QTableWidgetItem") -> bool:
        a, b = self.text().strip(), other.text().strip()
        if not a:
            return False  # пустые значения — в конец
        if not b:
            return True
        try:
            def _num(s: str) -> float:
                return float(
                    s.replace(" ", "").replace(" ", "")
                     .replace("₽", "").replace(",", ".")
                )
            return _num(a) < _num(b)
        except ValueError:
            pass
        try:
            from datetime import datetime
            return datetime.strptime(a, self._DATE_FMT) < datetime.strptime(b, self._DATE_FMT)
        except ValueError:
            pass
        return a < b


def _merge_to_summa(df: "pd.DataFrame") -> "pd.DataFrame":
    """Если в DataFrame есть Поступление/Списание — объединяет их в Сумма.
    Если Сумма уже есть — ничего не делает."""
    if "Сумма" in df.columns:
        return df
    if "Поступление" not in df.columns and "Списание" not in df.columns:
        return df
    inc = pd.to_numeric(df["Поступление"], errors="coerce") if "Поступление" in df.columns \
        else pd.Series(0.0, index=df.index)
    exp = pd.to_numeric(df["Списание"],    errors="coerce") if "Списание"    in df.columns \
        else pd.Series(0.0, index=df.index)
    summa = inc.fillna(0) - exp.fillna(0)
    summa = summa.where(summa != 0, other=float("nan"))
    pos = list(df.columns).index("Поступление") if "Поступление" in df.columns else len(df.columns)
    df = df.drop(columns=[c for c in ("Поступление", "Списание") if c in df.columns])
    df.insert(min(pos, len(df.columns)), "Сумма", summa)
    return df


class LoadSettingsDialog(QDialog):
    """Диалог настроек перед загрузкой файла выписки."""

    _STYLE = """
        QDialog { background: #FFFFFF; color: #374151; }
        QLabel  { background: transparent; }
        QLabel#sectionLabel {
            color: #9CA3AF; font-size: 11px; font-weight: 600;
            letter-spacing: 0.5px; text-transform: uppercase;
        }
        QPushButton#fmtActive {
            background: #4F46E5; color: #ffffff; border: none;
            border-radius: 6px; padding: 8px 20px; font-size: 13px; font-weight: 600;
        }
        QPushButton#fmtInactive {
            background: #E5E7EB; color: #6B7280; border: 1px solid #D1D5DB;
            border-radius: 6px; padding: 8px 20px; font-size: 13px;
        }
        QPushButton#fmtInactive:hover { background: #E5E7EB; color: #374151; }
        QPushButton#btnPrimary {
            background: #4F46E5; color: white; border: none;
            border-radius: 6px; padding: 8px 20px; font-size: 13px; font-weight: 600;
        }
        QPushButton#btnPrimary:hover  { background: #6366F1; }
        QPushButton#btnPrimary:pressed { background: #4338CA; }
        QPushButton#btnSecondary {
            background: #E5E7EB; color: #6B7280; border: 1px solid #D1D5DB;
            border-radius: 6px; padding: 7px 16px; font-size: 13px;
        }
        QPushButton#btnSecondary:hover { background: #E5E7EB; color: #374151; }
        QCheckBox {
            color: #374151; background: transparent; font-size: 13px; spacing: 8px;
        }
        QCheckBox::indicator {
            width: 16px; height: 16px; border-radius: 4px;
            border: 1px solid #D1D5DB; background: #F8F9FA;
        }
        QCheckBox::indicator:checked {
            background: #4F46E5; border-color: #4F46E5;
            image: url(none);
        }
        QCheckBox::indicator:hover { border-color: #818CF8; }
        QFrame#divider { background: #E5E7EB; max-height: 1px; }
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Загрузка детализации")
        self.setModal(True)
        self.setFixedWidth(400)
        self._fmt = "sber"
        self._setup_ui()
        self.setStyleSheet(self._STYLE)

    def _setup_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(24, 24, 24, 20)
        lay.setSpacing(14)

        # Заголовок
        title = QLabel("Загрузка детализации")
        title.setStyleSheet("font-size:15px; font-weight:700; color:#111827;")
        lay.addWidget(title)

        div0 = QFrame(objectName="divider")
        div0.setFixedHeight(1)
        lay.addWidget(div0)

        # ── Формат файла ──────────────────────────────────────────────────
        lay.addWidget(QLabel("ФОРМАТ ФАЙЛА", objectName="sectionLabel"))

        fmt_row = QHBoxLayout()
        fmt_row.setSpacing(8)
        self._btn_sber = QPushButton("СберБизнес (операции)", objectName="fmtActive")
        self._btn_snt  = QPushButton("СНТ Учёт",              objectName="fmtInactive")
        self._btn_sber.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._btn_snt .setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._btn_sber.clicked.connect(lambda: self._set_fmt("sber"))
        self._btn_snt .clicked.connect(lambda: self._set_fmt("snt"))
        fmt_row.addWidget(self._btn_sber)
        fmt_row.addWidget(self._btn_snt)
        fmt_row.addStretch()
        lay.addLayout(fmt_row)

        # Подсказка под кнопками формата
        self._fmt_hint = QLabel()
        self._fmt_hint.setWordWrap(True)
        self._fmt_hint.setStyleSheet("color:#9CA3AF; font-size:11px; background:transparent;")
        lay.addWidget(self._fmt_hint)
        self._update_hint()

        div1 = QFrame(objectName="divider")
        div1.setFixedHeight(1)
        lay.addWidget(div1)

        # ── Автораспределение ─────────────────────────────────────────────
        lay.addWidget(QLabel("АВТОМАТИЧЕСКОЕ РАСПРЕДЕЛЕНИЕ", objectName="sectionLabel"))

        self.chk_cat  = QCheckBox("Категория")
        self.chk_plot = QCheckBox("Участок")
        self.chk_cat .setChecked(True)
        self.chk_plot.setChecked(True)
        lay.addWidget(self.chk_cat)
        lay.addWidget(self.chk_plot)

        div2 = QFrame(objectName="divider")
        div2.setFixedHeight(1)
        lay.addWidget(div2)

        # ── Кнопки ────────────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)
        btn_cancel = QPushButton("Отмена",           objectName="btnSecondary")
        btn_ok     = QPushButton("Выбрать файл", objectName="btnPrimary")
        btn_cancel.clicked.connect(self.reject)
        btn_ok    .clicked.connect(self.accept)
        btn_row.addStretch()
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_ok)
        lay.addLayout(btn_row)

    def _set_fmt(self, fmt: str):
        self._fmt = fmt
        self._btn_sber.setObjectName("fmtActive"   if fmt == "sber" else "fmtInactive")
        self._btn_snt .setObjectName("fmtInactive" if fmt == "sber" else "fmtActive")
        # Перерисовываем стиль после смены objectName
        for btn in (self._btn_sber, self._btn_snt):
            btn.style().unpolish(btn)
            btn.style().polish(btn)
        self._update_hint()

    def _update_hint(self):
        if self._fmt == "sber":
            self._fmt_hint.setText(
                "Стандартная выгрузка операций из СберБизнес (.xlsx)")
        else:
            self._fmt_hint.setText(
                "Файл в формате программы СНТ Учёт — столбцы уже приведены к нужному виду")

    @property
    def fmt(self) -> str:
        return self._fmt

    @property
    def auto_cat(self) -> bool:
        return self.chk_cat.isChecked()

    @property
    def auto_plot(self) -> bool:
        return self.chk_plot.isChecked()


class DetailWidget(QWidget):
    dataLoaded = pyqtSignal(object)   # эмитится после успешной загрузки выписки

    def __init__(self):
        super().__init__()
        self.setAutoFillBackground(True)
        self.df_full = None
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(14)

        # Заголовок + кнопка
        top_bar = QHBoxLayout()
        title = QLabel("Детализация операций")
        title.setObjectName("pageTitle")
        top_bar.addWidget(title)
        top_bar.addStretch()
        self.btn_load = QPushButton("Загрузить файл")
        self.btn_load.setObjectName("btnPrimary")
        self.btn_load.clicked.connect(self.load_file)
        top_bar.addWidget(self.btn_load)

        btn_excel = QPushButton("Экспорт в Excel", objectName="btnSecondary")
        btn_excel.clicked.connect(self._export_excel)
        top_bar.addWidget(btn_excel)

        layout.addLayout(top_bar)

        # Панель фильтров
        filter_frame = QFrame()
        filter_frame.setObjectName("filterFrame")
        fl = QHBoxLayout(filter_frame)
        fl.setContentsMargins(16, 12, 16, 12)
        fl.setSpacing(10)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Поиск по контрагенту или назначению...")
        self.search_input.setObjectName("searchInput")
        self.search_input.textChanged.connect(self.apply_filters)
        fl.addWidget(self.search_input, stretch=3)

        self.combo_type = QComboBox()
        self.combo_type.setObjectName("filterCombo")
        self.combo_type.addItems(["Все операции", "Поступления", "Списания"])
        self.combo_type.currentIndexChanged.connect(self.apply_filters)
        fl.addWidget(self.combo_type, stretch=1)

        self.combo_cat = QComboBox()
        self.combo_cat.setObjectName("filterCombo")
        self.combo_cat.addItem("Все категории")
        for cat in ALL_CATEGORIES:
            self.combo_cat.addItem(cat)
        self.combo_cat.currentIndexChanged.connect(self.apply_filters)
        fl.addWidget(self.combo_cat, stretch=2)

        fl.addWidget(QLabel("с", objectName="filterLabel"))
        self.date_from = QDateEdit(calendarPopup=True, objectName="datePicker",
                                   displayFormat="dd.MM.yyyy")
        self.date_from.setDate(QDate(2021, 1, 1))
        self.date_from.dateChanged.connect(self.apply_filters)
        fl.addWidget(self.date_from)

        fl.addWidget(QLabel("по", objectName="filterLabel"))
        self.date_to = QDateEdit(calendarPopup=True, objectName="datePicker",
                                 displayFormat="dd.MM.yyyy")
        self.date_to.setDate(QDate.currentDate())
        self.date_to.dateChanged.connect(self.apply_filters)
        fl.addWidget(self.date_to)

        btn_reset = QPushButton("✕  Сбросить", objectName="btnSecondary")
        btn_reset.clicked.connect(self.reset_filters)
        fl.addWidget(btn_reset)

        layout.addWidget(filter_frame)

        self.status_label = QLabel("Файл не загружен", objectName="statusLabel")
        layout.addWidget(self.status_label)

        self.table = QTableWidget(objectName="mainTable")
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(True)
        self.table.setShowGrid(True)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_context_menu)
        self.table.itemChanged.connect(self._on_item_changed)
        layout.addWidget(self.table)

        summary_layout = QHBoxLayout()
        self.lbl_income  = QLabel("Поступления: —", objectName="summaryIncome")
        self.lbl_expense = QLabel("Списания: —",    objectName="summaryExpense")
        summary_layout.addWidget(self.lbl_income)
        summary_layout.addStretch()
        summary_layout.addWidget(self.lbl_expense)
        layout.addLayout(summary_layout)

    # ------------------------------------------------------------------ #
    def load_file(self):
        settings_dlg = LoadSettingsDialog(self)
        if settings_dlg.exec() != QDialog.DialogCode.Accepted:
            return

        fmt       = settings_dlg.fmt
        auto_cat  = settings_dlg.auto_cat
        auto_plot = settings_dlg.auto_plot

        path, _ = QFileDialog.getOpenFileName(
            self, "Открыть файл выписки", "", "Excel файлы (*.xlsx *.xls)")
        if not path:
            return
        try:
            df = pd.read_excel(path, engine="openpyxl")
            # Убираем пустые и валютные столбцы (актуально для СберБизнес)
            cols = [c for c in df.columns
                    if not str(c).strip().startswith("Валюта") and str(c).strip() != ""]
            df = df[cols]

            if fmt == "sber":
                # СберБизнес: удаляем технические столбцы банка
                drop_cols = {"Номер", "Номер счёта", "Контрагент счёт", "Контрагент cчёт"}
                df.drop(columns=[c for c in drop_cols if c in df.columns], inplace=True)

            # Приводим к единому формату: Поступление/Списание → Сумма
            df = _merge_to_summa(df)

            df["Дата"] = pd.to_datetime(df["Дата"], dayfirst=True, errors="coerce")
            df = df[df["Дата"].notna()].copy()

            if auto_cat:
                df = apply_categorization(df)
            if auto_plot:
                df = apply_plot_column(df)

            # Баг 1: гарантируем наличие столбцов Категория и Участок
            if "Категория" not in df.columns:
                df["Категория"] = ""
            if "Участок" not in df.columns:
                df["Участок"] = ""

            # Баг 2: столбец Участок может прийти как float (43.0) — приводим к строке
            df["Участок"] = df["Участок"].apply(
                lambda v: "" if pd.isna(v) else
                str(int(v)) if isinstance(v, float) and v == int(v) else str(v)
            )

            self.df_full = df
            min_d, max_d = df["Дата"].min(), df["Дата"].max()
            if pd.notna(min_d):
                self.date_from.setDate(QDate(min_d.year, min_d.month, min_d.day))
            if pd.notna(max_d):
                self.date_to.setDate(QDate(max_d.year, max_d.month, max_d.day))
            self.apply_filters()
            self.dataLoaded.emit(self.df_full)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка загрузки", f"Не удалось загрузить файл:\n{e}")

    # ------------------------------------------------------------------ #
    def load_dataframe(self, df: "pd.DataFrame"):
        """Восстанавливает DataFrame из сохранённого проекта без диалога выбора файла."""
        drop_cols = {"Номер", "Номер счёта", "Контрагент счёт", "Контрагент cчёт"}
        df = df.drop(columns=[c for c in drop_cols if c in df.columns])
        df = _merge_to_summa(df)   # совместимость со старыми проектами
        self.df_full = df
        min_d, max_d = df["Дата"].min(), df["Дата"].max()
        if pd.notna(min_d):
            self.date_from.setDate(QDate(min_d.year, min_d.month, min_d.day))
        if pd.notna(max_d):
            self.date_to.setDate(QDate(max_d.year, max_d.month, max_d.day))
        self.apply_filters()
        self.dataLoaded.emit(self.df_full)

    # ------------------------------------------------------------------ #
    def refresh_plot_column(self):
        """Пересчитывает столбец «Участок» по актуальным данным из snt_plots.json."""
        if self.df_full is None:
            return
        self.df_full = apply_plot_column(self.df_full)
        self.apply_filters()

    # ------------------------------------------------------------------ #
    def apply_filters(self):
        if self.df_full is None:
            return
        df = self.df_full.copy()

        d_from = self.date_from.date().toPyDate()
        d_to   = self.date_to.date().toPyDate()
        df = df[(df["Дата"].dt.date >= d_from) & (df["Дата"].dt.date <= d_to)]

        op_type = self.combo_type.currentText()
        if op_type == "Поступления" and "Сумма" in df.columns:
            df = df[pd.to_numeric(df["Сумма"], errors="coerce") > 0]
        elif op_type == "Списания" and "Сумма" in df.columns:
            df = df[pd.to_numeric(df["Сумма"], errors="coerce") < 0]

        cat_filter = self.combo_cat.currentText()
        if cat_filter != "Все категории":
            df = df[df["Категория"] == cat_filter]

        search = self.search_input.text().strip().lower()
        if search:
            mask = (
                df["Контрагент"].astype(str).str.lower().str.contains(search, na=False) |
                df["Назначение"].astype(str).str.lower().str.contains(search, na=False)
            )
            df = df[mask]

        self._fill_table(df)

    def reset_filters(self):
        self.search_input.clear()
        self.combo_type.setCurrentIndex(0)
        self.combo_cat.setCurrentIndex(0)
        if self.df_full is not None:
            min_d, max_d = self.df_full["Дата"].min(), self.df_full["Дата"].max()
            if pd.notna(min_d):
                self.date_from.setDate(QDate(min_d.year, min_d.month, min_d.day))
            if pd.notna(max_d):
                self.date_to.setDate(QDate(max_d.year, max_d.month, max_d.day))
        self.apply_filters()

    # ------------------------------------------------------------------ #
    def _export_excel(self):
        if self.df_full is None:
            QMessageBox.warning(self, "Нет данных", "Сначала загрузите файл выписки.")
            return

        # Экспортируем текущий отфильтрованный вид (с оригинальными столбцами)
        df = self.df_full.copy()
        d_from = self.date_from.date().toPyDate()
        d_to   = self.date_to.date().toPyDate()
        df = df[(df["Дата"].dt.date >= d_from) & (df["Дата"].dt.date <= d_to)]

        op_type = self.combo_type.currentText()
        if op_type == "Поступления" and "Сумма" in df.columns:
            df = df[pd.to_numeric(df["Сумма"], errors="coerce") > 0]
        elif op_type == "Списания" and "Сумма" in df.columns:
            df = df[pd.to_numeric(df["Сумма"], errors="coerce") < 0]

        cat_filter = self.combo_cat.currentText()
        if cat_filter != "Все категории":
            df = df[df["Категория"] == cat_filter]

        search = self.search_input.text().strip().lower()
        if search:
            mask = (
                df["Контрагент"].astype(str).str.lower().str.contains(search, na=False) |
                df["Назначение"].astype(str).str.lower().str.contains(search, na=False)
            )
            df = df[mask]

        if df.empty:
            QMessageBox.information(self, "Нет данных", "После применения фильтров нет строк для экспорта.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Экспорт детализации", "детализация.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        if not path.endswith(".xlsx"):
            path += ".xlsx"

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Детализация"

            headers = list(df.columns)
            summa_col_idx = headers.index("Сумма") + 1 if "Сумма" in headers else None

            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for _, row in df.iterrows():
                out_row = []
                for col in headers:
                    val = row[col]
                    if col == "Сумма":
                        num = pd.to_numeric(val, errors="coerce")
                        out_row.append(float(num) if pd.notna(num) else "")
                    elif col == "Дата":
                        out_row.append(val.strftime("%d.%m.%Y") if pd.notna(val) else "")
                    else:
                        out_row.append("" if pd.isna(val) else val)
                ws.append(out_row)

            # Цвет ячеек Сумма: зелёный для поступлений, красный для списаний
            if summa_col_idx is not None:
                green_fill = PatternFill("solid", fgColor="C8E6C9")
                red_fill   = PatternFill("solid", fgColor="FFCDD2")
                summa_letter = get_column_letter(summa_col_idx)
                for r in range(2, ws.max_row + 1):
                    cell = ws[f"{summa_letter}{r}"]
                    if isinstance(cell.value, (int, float)):
                        cell.fill  = green_fill if cell.value >= 0 else red_fill
                        cell.number_format = '#,##0.00 ₽'
                        cell.alignment = Alignment(horizontal="right")

            # Авто-ширина столбцов
            for col_cells in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

            wb.save(path)
            QMessageBox.information(self, "Экспорт завершён", f"Файл сохранён:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка экспорта", str(e))

    # ------------------------------------------------------------------ #
    def _fill_table(self, df: pd.DataFrame):
        self.table.blockSignals(True)
        self.table.setSortingEnabled(False)
        self.table.clearContents()

        columns = list(df.columns)
        self.table.setColumnCount(len(columns))
        self.table.setRowCount(len(df))
        self.table.setHorizontalHeaderLabels(columns)

        col_widths = {
            "Дата": 95, "Контрагент": 260,
            "Сумма": 140,
            "Назначение": 340, "Категория": 210, "Участок": 80,
        }

        for row_idx, (df_idx, row) in enumerate(df.iterrows()):
            cat       = str(row.get("Категория", "Прочее"))
            row_color = CATEGORY_COLORS.get(cat, QColor(55, 55, 60))

            for col_idx, col in enumerate(columns):
                val = row[col]
                if col == "Сумма":
                    num = pd.to_numeric(val, errors="coerce")
                    if pd.notna(num) and num > 0:
                        text = f"{num:,.2f} ₽".replace(",", " ")
                        fg   = QColor("#059669")
                    elif pd.notna(num) and num < 0:
                        text = f"−{abs(num):,.2f} ₽".replace(",", " ")
                        fg   = QColor("#DC2626")
                    else:
                        text = ""
                        fg   = QColor("#374151")
                    item = _SortItem(text)
                    item.setForeground(fg)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight)
                elif col == "Дата" and pd.notna(val):
                    text = val.strftime("%d.%m.%Y")
                    item = _SortItem(text)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
                else:
                    text = "" if pd.isna(val) else str(val)
                    item = _SortItem(text)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)

                item.setBackground(row_color)
                item.setData(Qt.ItemDataRole.UserRole, df_idx)
                self.table.setItem(row_idx, col_idx, item)

        header = self.table.horizontalHeader()
        for col_idx, col in enumerate(columns):
            w = col_widths.get(col)
            if w:
                self.table.setColumnWidth(col_idx, w)
                header.setSectionResizeMode(col_idx, QHeaderView.ResizeMode.Interactive)
            else:
                header.setSectionResizeMode(col_idx, QHeaderView.ResizeMode.ResizeToContents)

        self.table.setSortingEnabled(True)
        self.table.blockSignals(False)

        if "Сумма" in df.columns:
            s = pd.to_numeric(df["Сумма"], errors="coerce")
            total_in  = s[s > 0].sum()
            total_out = s[s < 0].abs().sum()
        else:
            total_in = total_out = 0.0
        self.lbl_income.setText(f"✅  Поступления: {total_in:,.2f} ₽".replace(",", " "))
        self.lbl_expense.setText(f"🔴  Списания: {total_out:,.2f} ₽".replace(",", " "))
        self.status_label.setText(f"Показано записей: {len(df)}")


    # ------------------------------------------------------------------ #
    #  КОНТЕКСТНОЕ МЕНЮ (ПКМ)
    # ------------------------------------------------------------------ #
    def _show_context_menu(self, pos: QPoint):
        row = self.table.rowAt(pos.y())
        if row < 0:
            return
        self.table.selectRow(row)

        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background: #F8F9FA;
                border: 1px solid #D1D5DB;
                color: #374151;
                font-size: 13px;
                padding: 4px;
            }
            QMenu::item {
                padding: 8px 20px;
                border-radius: 4px;
            }
            QMenu::item:selected {
                background: #EEF2FF;
                color: #6366F1;
            }
            QMenu::separator {
                height: 1px;
                background: #E5E7EB;
                margin: 4px 8px;
            }
        """)

        act_dup = QAction("Дублировать строку", self)
        act_dup.triggered.connect(lambda: self._duplicate_row(row))
        menu.addAction(act_dup)

        menu.addSeparator()

        act_del = QAction("Удалить строку", self)
        act_del.triggered.connect(lambda: self._delete_row(row))
        menu.addAction(act_del)

        menu.exec(self.table.viewport().mapToGlobal(pos))

    def _duplicate_row(self, row: int):
        """Вставляет копию строки row сразу под ней."""
        col_count = self.table.columnCount()
        insert_at = row + 1

        # Получаем pandas-индекс исходной строки
        src_item0 = self.table.item(row, 0)
        src_df_idx = src_item0.data(Qt.ItemDataRole.UserRole) if src_item0 else None

        # Дублируем строку в df_full заранее, чтобы знать новый индекс
        # до того, как сортировка изменит позиции строк в таблице
        new_df_idx = None
        if self.df_full is not None and src_df_idx is not None and src_df_idx in self.df_full.index:
            new_df_idx = int(self.df_full.index.max()) + 1
            self.df_full.loc[new_df_idx] = self.df_full.loc[src_df_idx].copy()

        # ВАЖНО: отключаем сортировку перед вставкой строки.
        # При setSortingEnabled(True) вызов setItem() немедленно пересортировывает
        # таблицу внутри самого вызова (blockSignals не помогает — сортировка
        # не идёт через сигналы). Строка «улетает» на другую позицию, все
        # последующие item(row, col) читают чужие данные → краш.
        self.table.blockSignals(True)
        self.table.setSortingEnabled(False)
        self.table.insertRow(insert_at)
        for col in range(col_count):
            src_item = self.table.item(row, col)
            if src_item:
                new_item = _SortItem(src_item.text())
                new_item.setBackground(src_item.background())
                new_item.setForeground(src_item.foreground())
                new_item.setTextAlignment(src_item.textAlignment())
                if new_df_idx is not None:
                    new_item.setData(Qt.ItemDataRole.UserRole, new_df_idx)
                self.table.setItem(insert_at, col, new_item)
        self.table.setSortingEnabled(True)
        self.table.blockSignals(False)

        self.table.selectRow(insert_at)
        self._update_summary()

    def _delete_row(self, row: int):
        """Удаляет строку из таблицы с подтверждением."""
        reply = QMessageBox.question(
            self, "Удаление строки",
            "Удалить выбранную строку?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            item0 = self.table.item(row, 0)
            df_idx = item0.data(Qt.ItemDataRole.UserRole) if item0 else None
            self.table.removeRow(row)
            if self.df_full is not None and df_idx is not None and df_idx in self.df_full.index:
                self.df_full = self.df_full.drop(index=df_idx)
            self._update_summary()

    # ------------------------------------------------------------------ #
    #  ОБРАБОТКА РЕДАКТИРОВАНИЯ ЯЧЕЙКИ
    # ------------------------------------------------------------------ #
    def _on_item_changed(self, item: QTableWidgetItem):
        """Обновляет цвет и выравнивание после ручного редактирования."""
        # Защита от рекурсии при programmatic setBackground/setForeground
        if self.table.signalsBlocked():
            return

        col_idx = item.column()
        col_name = self.table.horizontalHeaderItem(col_idx)
        if col_name is None:
            return
        col = col_name.text()

        self.table.blockSignals(True)

        # Пересчитываем цвет всей строки если изменилась Категория
        if col == "Категория":
            cat = item.text().strip()
            row_color = CATEGORY_COLORS.get(cat, QColor(55, 55, 60))
            for c in range(self.table.columnCount()):
                cell = self.table.item(item.row(), c)
                if cell:
                    cell.setBackground(row_color)

        # Цвет и выравнивание для столбца Сумма
        if col == "Сумма":
            item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight)
            raw_num = item.text().replace(" ", "").replace("−", "-").replace("₽", "").replace(",", ".")
            try:
                num_val = float(raw_num)
                item.setForeground(QColor("#059669") if num_val > 0 else QColor("#DC2626"))
            except ValueError:
                item.setForeground(QColor("#374151"))

        self.table.blockSignals(False)

        # Записываем изменение обратно в df_full
        if self.df_full is not None:
            df_idx = item.data(Qt.ItemDataRole.UserRole)
            if df_idx is not None and df_idx in self.df_full.index:
                new_text = item.text().strip()
                if col == "Сумма":
                    raw = new_text.replace(" ", "").replace("−", "-").replace("₽", "").replace(",", ".")
                    try:
                        self.df_full.at[df_idx, "Сумма"] = float(raw) if raw else float("nan")
                    except ValueError:
                        pass
                elif col == "Дата":
                    try:
                        self.df_full.at[df_idx, col] = pd.to_datetime(new_text, dayfirst=True)
                    except Exception:
                        pass
                else:
                    self.df_full.at[df_idx, col] = new_text

        self._update_summary()

    def _update_summary(self):
        """Пересчитывает итоги по текущему содержимому таблицы."""
        col_headers = {
            self.table.horizontalHeaderItem(c).text(): c
            for c in range(self.table.columnCount())
            if self.table.horizontalHeaderItem(c)
        }
        total_in  = 0.0
        total_out = 0.0
        rows = self.table.rowCount()

        for r in range(rows):
            for name, c in col_headers.items():
                cell = self.table.item(r, c)
                if cell:
                    raw = cell.text().replace(" ", "").replace("₽", "").replace("−", "-").replace(",", ".")
                    try:
                        val = float(raw)
                    except ValueError:
                        continue
                    if name == "Сумма":
                        if val > 0:
                            total_in  += val
                        elif val < 0:
                            total_out += abs(val)

        self.lbl_income.setText(f"✅  Поступления: {total_in:,.2f} ₽".replace(",", " "))
        self.lbl_expense.setText(f"🔴  Списания: {total_out:,.2f} ₽".replace(",", " "))
        self.status_label.setText(f"Показано записей: {self.table.rowCount()}")


# ======================================================================= #
#  ВКЛАДКА СВОДКА
# ======================================================================= #

# Порядок участков для строк таблицы — динамически из snt_plots.json
def _plot_num_key(s: str):
    """Ключ сортировки: числовые участки по значению, остальные в конце."""
    try:
        return (0, int(s), s)
    except ValueError:
        return (1, 0, s)

def _load_plot_order() -> list[str]:
    """Возвращает список номеров участков из snt_plots.json, отсортированный по _plot_num_key."""
    try:
        if os.path.exists(_PLOTS_FILE):
            with open(_PLOTS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            nums = [str(e.get("num", "")) for e in data if e.get("num")]
            return sorted(set(nums), key=_plot_num_key)
    except Exception:
        pass
    return []


# ======================================================================= #
#  ВКЛАДКА «НОРМАТИВЫ»
# ======================================================================= #

class RatesWidget(QWidget):
    """Вкладка тарифов на электроэнергию (₽/кВт·ч)."""

    DATA_FILE = os.path.join(DATA_DIR, "snt_rates.json")

    def __init__(self):
        super().__init__()
        self._rates: list = self._load()  # [{"date": "YYYY-MM-DD", "rate": "X.XX", "note": "..."}]
        self._setup_ui()
        self._rebuild_table()

    def _load(self) -> list:
        try:
            if os.path.exists(self.DATA_FILE):
                with open(self.DATA_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            pass
        # Стартовые тарифы для СНТ (можно удалить)
        return []

    def _save(self):
        try:
            with open(self.DATA_FILE, "w", encoding="utf-8") as f:
                json.dump(self._rates, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def reload(self):
        self._rates = self._load()
        self._rebuild_table()

    def _setup_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(16)

        # Заголовок
        top = QHBoxLayout()
        title = QLabel("Нормативы (тарифы на электроэнергию)", objectName="pageTitle")
        top.addWidget(title)
        top.addStretch()
        btn_add = QPushButton("＋  Добавить тариф")
        btn_add.setObjectName("btnPrimary")
        btn_add.clicked.connect(self._add_rate)
        top.addWidget(btn_add)
        lay.addLayout(top)

        hint = QLabel(
            "Двойной клик по ячейке — редактировать.  "
            "ПКМ по строке — удалить.  "
            "Записи отсортированы по дате (новые сверху)."
        )
        hint.setStyleSheet("color: #9CA3AF; background: transparent; font-size: 11px;")
        lay.addWidget(hint)

        # Форма добавления (скрыта по умолчанию)
        self.form_frame = QFrame()
        self.form_frame.setObjectName("filterFrame")
        self.form_frame.setVisible(False)
        form_lay = QHBoxLayout(self.form_frame)
        form_lay.setContentsMargins(16, 12, 16, 12)
        form_lay.setSpacing(12)

        form_lay.addWidget(QLabel("Дата:", objectName="filterLabel"))
        self.inp_date = QDateEdit()
        self.inp_date.setObjectName("datePicker")
        self.inp_date.setCalendarPopup(True)
        self.inp_date.setDate(QDate.currentDate())
        self.inp_date.setDisplayFormat("dd.MM.yyyy")
        form_lay.addWidget(self.inp_date)

        form_lay.addWidget(QLabel("₽/кВт·ч:", objectName="filterLabel"))
        self.inp_rate = QLineEdit()
        self.inp_rate.setObjectName("searchInput")
        self.inp_rate.setPlaceholderText("например: 4.50")
        self.inp_rate.setFixedWidth(120)
        form_lay.addWidget(self.inp_rate)

        form_lay.addWidget(QLabel("Примечание:", objectName="filterLabel"))
        self.inp_note = QLineEdit()
        self.inp_note.setObjectName("searchInput")
        self.inp_note.setPlaceholderText("необязательно")
        form_lay.addWidget(self.inp_note, stretch=1)

        btn_ok = QPushButton("Добавить")
        btn_ok.setObjectName("btnPrimary")
        btn_ok.clicked.connect(self._confirm_add)
        form_lay.addWidget(btn_ok)

        btn_cancel = QPushButton("✕")
        btn_cancel.setObjectName("btnSecondary")
        btn_cancel.clicked.connect(lambda: self.form_frame.setVisible(False))
        form_lay.addWidget(btn_cancel)

        lay.addWidget(self.form_frame)

        # Таблица
        self.table = QTableWidget()
        self.table.setObjectName("summaryTable")
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Дата вступления в силу", "Тариф (₽/кВт·ч)", "Примечание"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(0, 200)
        self.table.setColumnWidth(1, 180)
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(False)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._context_menu)
        self.table.itemChanged.connect(self._on_cell_edited)
        lay.addWidget(self.table)

        self.status_lbl = QLabel("", objectName="statusLabel")
        lay.addWidget(self.status_lbl)

    def _rebuild_table(self):
        self.table.blockSignals(True)
        self.table.clearContents()

        # Сортировка: новые сверху
        rates = sorted(self._rates, key=lambda r: r.get("date", ""), reverse=True)

        self.table.setRowCount(len(rates))
        for r_idx, entry in enumerate(rates):
            # Дата
            raw_date = entry.get("date", "")
            try:
                from datetime import datetime
                d = datetime.strptime(raw_date, "%Y-%m-%d")
                display_date = d.strftime("%d.%m.%Y")
            except Exception:
                display_date = raw_date

            # Цвет: первая строка (самый актуальный тариф) — выделена
            is_current = (r_idx == 0)
            bg = "#0d2a0d" if is_current else "#F9FAFB"
            fg_date = "#059669" if is_current else "#374151"

            for c_idx, (text, fg) in enumerate([
                (display_date,           fg_date),
                (entry.get("rate", ""), "#6366F1" if is_current else "#374151"),
                (entry.get("note", ""), "#9CA3AF"),
            ]):
                it = QTableWidgetItem(text)
                it.setBackground(QColor(bg))
                it.setForeground(QColor(fg))
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                # Дата — не редактируем напрямую (только через форму добавления)
                if c_idx == 0:
                    it.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                self.table.setItem(r_idx, c_idx, it)

            self.table.setRowHeight(r_idx, 34)

        self.table.blockSignals(False)

        current = rates[0] if rates else None
        if current:
            self.status_lbl.setText(
                f"Актуальный тариф: {current.get('rate', '?')} ₽/кВт·ч  "
                f"(с {rates[0].get('date', '?')})  ·  всего записей: {len(rates)}"
            )
        else:
            self.status_lbl.setText("Нет записей — добавьте первый тариф")

    def _add_rate(self):
        self.inp_rate.clear()
        self.inp_note.clear()
        self.form_frame.setVisible(True)
        self.inp_rate.setFocus()

    def _confirm_add(self):
        rate_text = self.inp_rate.text().strip().replace(",", ".")
        if not rate_text:
            self.inp_rate.setFocus()
            return
        try:
            float(rate_text)
        except ValueError:
            self.inp_rate.setStyleSheet(self.inp_rate.styleSheet() + "border:1px solid #DC2626;")
            return

        date_str = self.inp_date.date().toString("yyyy-MM-dd")
        entry = {"date": date_str, "rate": rate_text, "note": self.inp_note.text().strip()}
        self._rates.append(entry)
        self._save()
        self.form_frame.setVisible(False)
        self._rebuild_table()

    def _on_cell_edited(self, item: QTableWidgetItem):
        if self.table.signalsBlocked():
            return
        # Пересчитываем индекс в отсортированном списке
        rates_sorted = sorted(self._rates, key=lambda r: r.get("date", ""), reverse=True)
        r_idx = item.row()
        if r_idx >= len(rates_sorted):
            return
        col = item.column()
        val = item.text().strip()
        entry = rates_sorted[r_idx]
        # Находим оригинальную запись в self._rates
        orig_idx = next(
            (i for i, e in enumerate(self._rates) if e is entry), None
        )
        if orig_idx is None:
            return
        if col == 1:
            self._rates[orig_idx]["rate"] = val.replace(",", ".")
        elif col == 2:
            self._rates[orig_idx]["note"] = val
        self._save()
        self._rebuild_table()

    def _context_menu(self, pos: QPoint):
        row = self.table.rowAt(pos.y())
        if row < 0:
            return
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu{background:#F8F9FA;border:1px solid #D1D5DB;color:#374151;
                  font-size:13px;padding:4px;}
            QMenu::item{padding:8px 20px;border-radius:4px;}
            QMenu::item:selected{background:#EEF2FF;color:#DC2626;}
        """)
        act_del = QAction("удалить запись", self)
        act_del.triggered.connect(lambda: self._delete_rate(row))
        menu.addAction(act_del)
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def _delete_rate(self, row: int):
        rates_sorted = sorted(self._rates, key=lambda r: r.get("date", ""), reverse=True)
        if row >= len(rates_sorted):
            return
        entry = rates_sorted[row]
        reply = QMessageBox.question(
            self, "Удаление тарифа",
            f"Удалить запись от {entry.get('date', '')} ({entry.get('rate', '')} ₽/кВт·ч)?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._rates = [e for e in self._rates if e is not entry]
            self._save()
            self._rebuild_table()


class VznosyRatesWidget(QWidget):
    """Управление периодами членских взносов."""

    DATA_FILE = os.path.join(DATA_DIR, "snt_vznosy_rates.json")

    def __init__(self):
        super().__init__()
        self._rates: list = self._load()
        self._setup_ui()
        self._rebuild_table()

    def _load(self) -> list:
        try:
            if os.path.exists(self.DATA_FILE):
                with open(self.DATA_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            pass
        return []

    def _save(self):
        try:
            with open(self.DATA_FILE, "w", encoding="utf-8") as f:
                json.dump(self._rates, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def reload(self):
        self._rates = self._load()
        self._rebuild_table()

    @staticmethod
    def _fmt_date(iso: str) -> str:
        try:
            from datetime import datetime
            return datetime.strptime(iso, "%Y-%m-%d").strftime("%d.%m.%Y")
        except Exception:
            return iso or "—"

    def _sorted_periods(self) -> list:
        """Периоды, отсортированные по date_from по убыванию (новые сверху)."""
        def _key(r):
            return r.get("date_from", r.get("date", ""))
        return sorted(self._rates, key=_key, reverse=True)

    def _setup_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(16)

        top = QHBoxLayout()
        title = QLabel("Периоды членских взносов", objectName="pageTitle")
        top.addWidget(title)
        top.addStretch()
        btn_add = QPushButton("＋  Добавить период")
        btn_add.setObjectName("btnPrimary")
        btn_add.clicked.connect(self._add_rate)
        top.addWidget(btn_add)
        lay.addLayout(top)

        hint = QLabel(
            "Каждый период — отдельная строка тарифа.  "
            "ПКМ по строке — удалить.  "
            "Двойной клик по сумме — редактировать."
        )
        hint.setStyleSheet("color: #9CA3AF; background: transparent; font-size: 11px;")
        lay.addWidget(hint)

        # ── Форма добавления периода ──────────────────────────────
        self.form_frame = QFrame()
        self.form_frame.setObjectName("filterFrame")
        self.form_frame.setVisible(False)
        form_lay = QHBoxLayout(self.form_frame)
        form_lay.setContentsMargins(16, 12, 16, 12)
        form_lay.setSpacing(10)

        form_lay.addWidget(QLabel("С:", objectName="filterLabel"))
        self.inp_date_from = QDateEdit()
        self.inp_date_from.setObjectName("datePicker")
        self.inp_date_from.setCalendarPopup(True)
        self.inp_date_from.setDate(QDate.currentDate())
        self.inp_date_from.setDisplayFormat("dd.MM.yyyy")
        self.inp_date_from.setFixedWidth(115)
        form_lay.addWidget(self.inp_date_from)

        form_lay.addWidget(QLabel("По:", objectName="filterLabel"))
        self.inp_date_to = QDateEdit()
        self.inp_date_to.setObjectName("datePicker")
        self.inp_date_to.setCalendarPopup(True)
        self.inp_date_to.setDate(QDate.currentDate())
        self.inp_date_to.setDisplayFormat("dd.MM.yyyy")
        self.inp_date_to.setFixedWidth(115)
        form_lay.addWidget(self.inp_date_to)

        self.chk_open_end = QCheckBox("Открытый")
        self.chk_open_end.setToolTip("Не указывать конечную дату (последний активный период)")
        self.chk_open_end.setStyleSheet(
            "QCheckBox{color:#374151;background:transparent;font-size:12px;}"
            "QCheckBox::indicator{width:15px;height:15px;}"
        )
        self.chk_open_end.stateChanged.connect(
            lambda s: self.inp_date_to.setEnabled(not bool(s))
        )
        form_lay.addWidget(self.chk_open_end)

        form_lay.addWidget(QLabel("Сумма (₽):", objectName="filterLabel"))
        self.inp_amount = QLineEdit()
        self.inp_amount.setObjectName("searchInput")
        self.inp_amount.setPlaceholderText("10000")
        self.inp_amount.setFixedWidth(100)
        form_lay.addWidget(self.inp_amount)

        self.chk_per_sqm = QCheckBox("₽/м²")
        self.chk_per_sqm.setToolTip("Сумма указана в рублях за м²")
        self.chk_per_sqm.setStyleSheet(
            "QCheckBox{color:#374151;background:transparent;font-size:12px;}"
            "QCheckBox::indicator{width:15px;height:15px;}"
        )
        self.chk_per_sqm.stateChanged.connect(self._on_toggle_per_sqm)
        form_lay.addWidget(self.chk_per_sqm)

        form_lay.addWidget(QLabel("₽/м²:", objectName="filterLabel"))
        self.inp_rate_sqm = QLineEdit()
        self.inp_rate_sqm.setObjectName("searchInput")
        self.inp_rate_sqm.setPlaceholderText("15.00")
        self.inp_rate_sqm.setFixedWidth(80)
        self.inp_rate_sqm.setEnabled(False)
        form_lay.addWidget(self.inp_rate_sqm)

        form_lay.addWidget(QLabel("Примечание:", objectName="filterLabel"))
        self.inp_note = QLineEdit()
        self.inp_note.setObjectName("searchInput")
        self.inp_note.setPlaceholderText("необязательно")
        form_lay.addWidget(self.inp_note, stretch=1)

        btn_ok = QPushButton("Сохранить")
        btn_ok.setObjectName("btnPrimary")
        btn_ok.clicked.connect(self._confirm_add)
        form_lay.addWidget(btn_ok)

        btn_cancel = QPushButton("✕")
        btn_cancel.setObjectName("btnSecondary")
        btn_cancel.clicked.connect(lambda: self.form_frame.setVisible(False))
        form_lay.addWidget(btn_cancel)

        lay.addWidget(self.form_frame)

        # ── Таблица периодов ──────────────────────────────────────
        self.table = QTableWidget()
        self.table.setObjectName("summaryTable")
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(
            ["Период с", "По (включительно)", "Сумма (₽)", "Цена за м² (₽)", "Примечание"]
        )
        hdr = self.table.horizontalHeader()
        for c, (mode, w) in enumerate([
            (QHeaderView.ResizeMode.Fixed, 130),
            (QHeaderView.ResizeMode.Fixed, 150),
            (QHeaderView.ResizeMode.Fixed, 130),
            (QHeaderView.ResizeMode.Fixed, 130),
            (QHeaderView.ResizeMode.Stretch, 0),
        ]):
            hdr.setSectionResizeMode(c, mode)
            if w:
                self.table.setColumnWidth(c, w)
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(False)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._context_menu)
        self.table.itemChanged.connect(self._on_cell_edited)
        lay.addWidget(self.table)

        self.status_lbl = QLabel("", objectName="statusLabel")
        lay.addWidget(self.status_lbl)

    def _on_toggle_per_sqm(self, state):
        on = bool(state)
        self.inp_amount.setEnabled(not on)
        self.inp_rate_sqm.setEnabled(on)
        if on:
            self.inp_amount.clear()
            self.inp_rate_sqm.setFocus()
        else:
            self.inp_rate_sqm.clear()
            self.inp_amount.setFocus()

    def _rebuild_table(self):
        self.table.blockSignals(True)
        self.table.clearContents()

        periods = self._sorted_periods()
        self.table.setRowCount(len(periods))

        for r_idx, entry in enumerate(periods):
            date_from = entry.get("date_from", entry.get("date", ""))
            date_to = entry.get("date_to", "")
            is_current = (r_idx == 0)
            is_per_sqm = bool(entry.get("per_sqm"))
            bg = "#0d2a0d" if is_current else "#F9FAFB"
            fg_date = "#059669" if is_current else "#374151"
            fg_value = "#6366F1" if is_current else "#374151"

            amount_text = "—" if is_per_sqm else str(entry.get("amount", ""))
            rate_sqm_text = str(entry.get("rate_sqm", "")) if is_per_sqm else "—"
            date_to_text = self._fmt_date(date_to) if date_to else "открытый"

            cells = [
                (self._fmt_date(date_from), fg_date,  True),   # read-only
                (date_to_text,             fg_date,  True),   # read-only
                (amount_text,              fg_value, is_per_sqm),
                (rate_sqm_text,            fg_value, not is_per_sqm),
                (entry.get("note", ""),    "#9CA3AF", False),
            ]
            for c_idx, (text, fg, read_only) in enumerate(cells):
                it = QTableWidgetItem(text)
                it.setBackground(QColor(bg))
                it.setForeground(QColor(fg))
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                if read_only:
                    it.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                self.table.setItem(r_idx, c_idx, it)

            self.table.setRowHeight(r_idx, 34)

        self.table.blockSignals(False)

        if periods:
            current = periods[0]
            date_from = current.get("date_from", current.get("date", "?"))
            date_to = current.get("date_to", "")
            period_str = self._fmt_date(date_from)
            if date_to:
                period_str += f" — {self._fmt_date(date_to)}"
            if current.get("per_sqm"):
                desc = f"{current.get('rate_sqm', '?')} ₽/м²"
            else:
                desc = f"{current.get('amount', '?')} ₽"
            self.status_lbl.setText(
                f"Актуальный период: {period_str}  ·  {desc}  ·  всего периодов: {len(periods)}"
            )
        else:
            self.status_lbl.setText("Нет периодов — добавьте первый")

    def _add_rate(self):
        self.inp_amount.clear()
        self.inp_rate_sqm.clear()
        self.inp_note.clear()
        self.chk_per_sqm.setChecked(False)
        self.chk_open_end.setChecked(False)
        self.inp_date_to.setEnabled(True)
        self.form_frame.setVisible(True)
        self.inp_amount.setFocus()

    def _confirm_add(self):
        per_sqm = self.chk_per_sqm.isChecked()
        if per_sqm:
            raw = self.inp_rate_sqm.text().strip().replace(",", ".")
            target = self.inp_rate_sqm
        else:
            raw = self.inp_amount.text().strip().replace(",", ".")
            target = self.inp_amount

        if not raw:
            target.setFocus()
            return
        try:
            v = float(raw)
            if v <= 0:
                raise ValueError
        except ValueError:
            target.setStyleSheet(target.styleSheet() + "border:1px solid #DC2626;")
            return

        date_from_str = self.inp_date_from.date().toString("yyyy-MM-dd")
        entry: dict = {
            "date_from": date_from_str,
            "amount": "" if per_sqm else raw,
            "per_sqm": per_sqm,
            "rate_sqm": raw if per_sqm else "",
            "note": self.inp_note.text().strip(),
        }
        if not self.chk_open_end.isChecked():
            entry["date_to"] = self.inp_date_to.date().toString("yyyy-MM-dd")

        self._rates.append(entry)
        self._save()
        self.form_frame.setVisible(False)
        self._rebuild_table()

    def _on_cell_edited(self, item: QTableWidgetItem):
        if self.table.signalsBlocked():
            return
        periods = self._sorted_periods()
        r_idx = item.row()
        if r_idx >= len(periods):
            return
        col = item.column()
        val = item.text().strip()
        entry = periods[r_idx]
        orig_idx = next((i for i, e in enumerate(self._rates) if e is entry), None)
        if orig_idx is None:
            return
        if col == 2 and not entry.get("per_sqm"):
            self._rates[orig_idx]["amount"] = val.replace(",", ".")
        elif col == 3 and entry.get("per_sqm"):
            self._rates[orig_idx]["rate_sqm"] = val.replace(",", ".")
        elif col == 4:
            self._rates[orig_idx]["note"] = val
        self._save()
        self._rebuild_table()

    def _context_menu(self, pos: QPoint):
        row = self.table.rowAt(pos.y())
        if row < 0:
            return
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu{background:#F8F9FA;border:1px solid #D1D5DB;color:#374151;
                  font-size:13px;padding:4px;}
            QMenu::item{padding:8px 20px;border-radius:4px;}
            QMenu::item:selected{background:#EEF2FF;color:#DC2626;}
        """)
        act_del = QAction("Удалить период", self)
        act_del.triggered.connect(lambda: self._delete_rate(row))
        menu.addAction(act_del)
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def _delete_rate(self, row: int):
        periods = self._sorted_periods()
        if row >= len(periods):
            return
        entry = periods[row]
        date_from = entry.get("date_from", entry.get("date", ""))
        date_to = entry.get("date_to", "")
        period_str = self._fmt_date(date_from)
        if date_to:
            period_str += f" — {self._fmt_date(date_to)}"
        if entry.get("per_sqm"):
            desc = f"{entry.get('rate_sqm', '')} ₽/м²"
        else:
            desc = f"{entry.get('amount', '')} ₽"
        reply = QMessageBox.question(
            self, "Удаление периода",
            f"Удалить период {period_str} ({desc})?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._rates = [e for e in self._rates if e is not entry]
            self._save()
            self._rebuild_table()


class PlotsWidget(QWidget):
    """Вкладка участков: ручное добавление и управление списком."""

    plotsUpdated = pyqtSignal()   # эмитится при любом изменении snt_plots.json

    DATA_FILE = os.path.join(DATA_DIR, "snt_plots.json")

    def __init__(self):
        super().__init__()
        self.setAutoFillBackground(True)
        self._plots: list = self._load()  # [{"num": "15", "owners": ["Иван Петров", ...]}, ...]
        self._setup_ui()
        self._rebuild_table()

    def _load(self) -> list:
        try:
            if os.path.exists(self.DATA_FILE):
                with open(self.DATA_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            pass
        return []

    def _save(self):
        try:
            with open(self.DATA_FILE, "w", encoding="utf-8") as f:
                json.dump(self._plots, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
        self.plotsUpdated.emit()

    def reload(self):
        self._plots = self._load()
        self._rebuild_table()
        self.plotsUpdated.emit()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(14)

        # Заголовок + кнопка
        top = QHBoxLayout()
        title = QLabel("Участки")
        title.setObjectName("pageTitle")
        top.addWidget(title)
        top.addStretch()
        btn_import = QPushButton("📥  Импорт из Excel")
        btn_import.setObjectName("btnSecondary")
        btn_import.clicked.connect(self._import_from_excel)
        top.addWidget(btn_import)
        btn_add = QPushButton("＋  Добавить участок")
        btn_add.setObjectName("btnPrimary")
        btn_add.clicked.connect(self._add_plot)
        top.addWidget(btn_add)
        layout.addLayout(top)

        self.status_label = QLabel("", objectName="statusLabel")
        layout.addWidget(self.status_label)

        self.table = QTableWidget(objectName="mainTable")
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(False)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._context_menu)
        layout.addWidget(self.table)

    def _rebuild_table(self):
        self.table.blockSignals(True)
        self.table.clearContents()

        plots_sorted = sorted(self._plots, key=lambda p: str(p.get("num", "")))

        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Участок", "Собственники", "Площадь, м²"])
        self.table.setRowCount(len(plots_sorted))

        for r_idx, plot in enumerate(plots_sorted):
            num_item = QTableWidgetItem(f"уч. {plot.get('num', '?')}")
            num_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            num_item.setForeground(QColor("#6366F1"))
            f = num_item.font(); f.setBold(True); num_item.setFont(f)
            self.table.setItem(r_idx, 0, num_item)

            owner_widget = self._build_owners_cell(plot)
            self.table.setCellWidget(r_idx, 1, owner_widget)

            area_raw = plot.get("area")
            try:
                area_v = float(area_raw) if area_raw not in (None, "") else None
            except (TypeError, ValueError):
                area_v = None
            area_item = QTableWidgetItem(f"{area_v:g}" if area_v is not None else "—")
            area_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            area_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            area_item.setForeground(QColor("#374151" if area_v is not None else "#9CA3AF"))
            self.table.setItem(r_idx, 2, area_item)

            self.table.setRowHeight(r_idx, 28)

        hdr = self.table.horizontalHeader()
        hdr.setStretchLastSection(False)
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)

        self.table.blockSignals(False)
        self.status_label.setText(f"Участков: {len(plots_sorted)}")

    def _build_owners_cell(self, plot: dict) -> QWidget:
        owners = plot.get("owners", []) or []
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(6, 0, 6, 0)
        layout.setSpacing(8)

        if owners:
            first_owner = owners[0]
            first_label = QLabel(first_owner)
            first_label.setStyleSheet("color:#374151;font-size:13px;")
            first_label.setToolTip("\n".join(owners))
            layout.addWidget(first_label, 1)

            extra = len(owners) - 1
            if extra > 0:
                btn_more = QPushButton(f"+{extra}")
                btn_more.setCursor(Qt.CursorShape.PointingHandCursor)
                btn_more.setStyleSheet(
                    "QPushButton{background:transparent;color:#82cfff;border:none;"
                    "font-weight:700;padding:0px;margin:0px;}"
                    "QPushButton:hover{text-decoration:underline;}"
                )
                btn_more.clicked.connect(lambda _, p=plot: self._show_owners_popup(p))
                layout.addWidget(btn_more, 0, Qt.AlignmentFlag.AlignRight)
        else:
            label = QLabel("—")
            label.setStyleSheet("color:#9CA3AF;font-size:13px;")
            layout.addWidget(label)

        return container

    def _show_owners_popup(self, plot: dict):
        dlg = OwnersPopup(plot.get("num", "?"), plot.get("owners", []), self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            updated = dlg.get_owners()
            if updated != plot.get("owners", []):
                idx = self._plots.index(plot)
                self._plots[idx] = {**plot, "owners": updated}
                self._save()
                self._rebuild_table()

    def _import_from_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Выберите файл Excel", "",
            "Excel файлы (*.xlsx *.xls *.xlsm)"
        )
        if not path:
            return

        try:
            df = pd.read_excel(path, dtype=str)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка чтения файла", str(e))
            return

        # Ищем столбцы по частичному совпадению
        col_num = None
        col_name = None
        col_area = None
        for col in df.columns:
            col_lower = str(col).lower().strip()
            if col_num is None and ("участк" in col_lower or col_lower in ("№", "n", "номер")):
                col_num = col
            if col_name is None and ("ф.и.о" in col_lower or "фио" in col_lower or "имя" in col_lower or col_lower == "ф.и.о."):
                col_name = col
            if col_area is None and ("площад" in col_lower or "кв.м" in col_lower or "м²" in col_lower or "м2" in col_lower):
                col_area = col

        if col_num is None or col_name is None:
            QMessageBox.warning(
                self, "Неверный формат",
                f"Не удалось найти нужные столбцы.\n"
                f"Ожидается: «№ участка» и «Ф.И.О.»\n"
                f"Найдены столбцы: {', '.join(str(c) for c in df.columns)}"
            )
            return

        # Группируем по номеру участка
        imported: dict[str, dict] = {}
        for _, row in df.iterrows():
            num = str(row[col_num]).strip()
            name = str(row[col_name]).strip()
            if not num or num.lower() in ("nan", "none", "") or not name or name.lower() in ("nan", "none", ""):
                continue
            entry = imported.setdefault(num, {"owners": [], "area": None})
            if name not in entry["owners"]:
                entry["owners"].append(name)
            if col_area is not None and entry["area"] is None:
                raw = str(row[col_area]).strip().replace(",", ".")
                if raw and raw.lower() not in ("nan", "none"):
                    try:
                        v = float(raw)
                        if v > 0:
                            entry["area"] = v
                    except ValueError:
                        pass

        if not imported:
            QMessageBox.warning(self, "Пустой файл", "В файле не найдено данных об участках.")
            return

        # Диалог выбора режима импорта
        msg = QMessageBox(self)
        msg.setWindowTitle("Импорт участков")
        msg.setText(
            f"Найдено {len(imported)} участков в файле.\n\n"
            "Как импортировать?"
        )
        btn_replace = msg.addButton("Заменить всё", QMessageBox.ButtonRole.DestructiveRole)
        btn_merge   = msg.addButton("Объединить",   QMessageBox.ButtonRole.AcceptRole)
        btn_cancel  = msg.addButton("Отмена",        QMessageBox.ButtonRole.RejectRole)
        msg.setDefaultButton(btn_merge)
        msg.exec()

        clicked = msg.clickedButton()
        if clicked is btn_cancel:
            return

        if clicked is btn_replace:
            new_plots = []
            for num, entry in imported.items():
                item = {"num": num, "owners": entry["owners"]}
                if entry["area"] is not None:
                    item["area"] = entry["area"]
                new_plots.append(item)
            self._plots = new_plots
        else:
            # Объединяем: добавляем новые участки, дополняем существующие
            existing = {p["num"]: p for p in self._plots}
            for num, entry in imported.items():
                owners = entry["owners"]
                area = entry["area"]
                if num in existing:
                    current_owners = existing[num].get("owners", [])
                    for o in owners:
                        if o not in current_owners:
                            current_owners.append(o)
                    existing[num]["owners"] = current_owners
                    if area is not None and existing[num].get("area") in (None, "", 0):
                        existing[num]["area"] = area
                else:
                    item = {"num": num, "owners": owners}
                    if area is not None:
                        item["area"] = area
                    existing[num] = item
            self._plots = list(existing.values())

        self._save()
        self._rebuild_table()
        QMessageBox.information(
            self, "Импорт завершён",
            f"Импортировано {len(imported)} участков."
        )

    def _add_plot(self):
        dlg = PlotEditDialog(plot_data=None, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            result = dlg.get_result()
            if result:
                self._plots.append(result)
                self._save()
                self._rebuild_table()

    def _context_menu(self, pos: QPoint):
        row = self.table.rowAt(pos.y())
        if row < 0:
            return
        
        plots_sorted = sorted(self._plots, key=lambda p: str(p.get("num", "")))
        if row >= len(plots_sorted):
            return
        
        plot = plots_sorted[row]
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu{background:#F8F9FA;border:1px solid #D1D5DB;color:#374151;
                  font-size:13px;padding:4px;}
            QMenu::item{padding:8px 20px;border-radius:4px;}
            QMenu::item:selected{background:#EEF2FF;color:#DC2626;}
        """)
        
        act_edit = QAction("✏️  Редактировать", self)
        act_edit.triggered.connect(lambda: self._edit_plot(row, plot))
        menu.addAction(act_edit)
        
        act_del = QAction("Удалить", self)
        act_del.triggered.connect(lambda: self._delete_plot(row, plot))
        menu.addAction(act_del)
        
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def _edit_plot(self, row: int, plot: dict):
        dlg = PlotEditDialog(plot_data=plot, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            result = dlg.get_result()
            if result:
                # Обновляем оригинальный объект в списке
                idx = self._plots.index(plot)
                self._plots[idx] = result
                self._save()
                self._rebuild_table()

    def _delete_plot(self, row: int, plot: dict):
        reply = QMessageBox.question(
            self, "Удаление участка",
            f"Удалить участок № {plot.get('num', '?')}?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._plots = [p for p in self._plots if p is not plot]
            self._save()
            self._rebuild_table()


# ======================================================================= #
#  ВКЛАДКА «УЧАСТКИ»
# ======================================================================= #            

class OwnersPopup(QDialog):
    """Диалог просмотра/редактирования списка собственников участка."""

    def __init__(self, plot_num: str, owners: list[str], parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Собственники — уч. {plot_num}")
        self.setMinimumWidth(420)
        self.setModal(True)
        self._owners = list(owners)
        self._inputs: list[QLineEdit] = []
        self._setup_ui()
        self._apply_styles()

    def _setup_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 20, 20, 20)
        lay.setSpacing(10)

        title = QLabel("Список собственников")
        title.setStyleSheet("font-size:14px;font-weight:700;color:#111827;")
        lay.addWidget(title)

        # Прокручиваемая область с полями
        self._scroll_widget = QWidget()
        self._form_lay = QVBoxLayout(self._scroll_widget)
        self._form_lay.setSpacing(6)
        self._form_lay.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(self._scroll_widget)
        scroll.setStyleSheet(
            "QScrollArea{background:#F8F9FA;border:1px solid #E5E7EB;border-radius:6px;}"
        )
        scroll.setMinimumHeight(140)
        scroll.setMaximumHeight(300)
        lay.addWidget(scroll)

        for name in self._owners:
            self._add_owner_row(name)

        btn_add = QPushButton("＋  Добавить собственника")
        btn_add.setObjectName("btnSecondary")
        btn_add.clicked.connect(lambda: self._add_owner_row(""))
        lay.addWidget(btn_add)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        btns.setStyleSheet(
            "QPushButton{background:#4F46E5;color:white;border:none;border-radius:6px;"
            "padding:7px 18px;font-size:13px;font-weight:600;}"
            "QPushButton:hover{background:#6366F1;}"
            "QPushButton[text='Cancel']{background:#E5E7EB;color:#6B7280;}"
        )
        lay.addWidget(btns)

    def _add_owner_row(self, name: str):
        row_widget = QWidget()
        row_widget.setStyleSheet("background:transparent;")
        rlay = QHBoxLayout(row_widget)
        rlay.setContentsMargins(6, 2, 6, 2)
        rlay.setSpacing(6)

        inp = QLineEdit(name)
        inp.setPlaceholderText("Фамилия Имя Отчество")
        inp.setStyleSheet(
            "background:#F8F9FA;border:1px solid #D1D5DB;border-radius:5px;"
            "color:#374151;padding:6px 10px;font-size:13px;"
        )
        self._inputs.append(inp)
        rlay.addWidget(inp, stretch=1)

        btn_del = QPushButton("✕")
        btn_del.setFixedSize(28, 28)
        btn_del.setStyleSheet(
            "QPushButton{background:#2a1a1a;border:1px solid #5a2a2a;"
            "border-radius:5px;color:#DC2626;font-size:13px;}"
            "QPushButton:hover{background:#3a2020;}"
        )
        btn_del.clicked.connect(lambda _, w=row_widget, i=inp: self._remove_row(w, i))
        rlay.addWidget(btn_del)

        self._form_lay.addWidget(row_widget)

    def _remove_row(self, row_widget: QWidget, inp: QLineEdit):
        if inp in self._inputs:
            self._inputs.remove(inp)
        row_widget.setParent(None)
        row_widget.deleteLater()

    def get_owners(self) -> list[str]:
        return [inp.text().strip() for inp in self._inputs if inp.text().strip()]

    def _apply_styles(self):
        self.setStyleSheet(
            "QDialog{background:#FFFFFF;color:#374151;}"
            "QLabel{background:transparent;color:#374151;}"
        )


class PlotEditDialog(QDialog):
    """Диалог добавления / редактирования участка."""

    def __init__(self, plot_data: dict | None = None, parent=None):
        super().__init__(parent)
        self._is_edit = plot_data is not None
        self._plot_data = plot_data or {}
        self.setWindowTitle("Редактировать участок" if self._is_edit else "Новый участок")
        self.setMinimumWidth(460)
        self.setModal(True)
        self._owner_inputs: list[QLineEdit] = []
        self._setup_ui()
        self._apply_styles()

    def _setup_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(24, 24, 24, 20)
        lay.setSpacing(14)

        form = QFormLayout()
        form.setSpacing(10)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        # Номер участка
        self.inp_num = QLineEdit(str(self._plot_data.get("num", "")))
        self.inp_num.setPlaceholderText("например: 15 или 15/207")
        if self._is_edit:
            self.inp_num.setReadOnly(True)
            self.inp_num.setStyleSheet(
                "background:#F3F4F6;border:1px solid #E5E7EB;"
                "border-radius:5px;color:#9CA3AF;padding:7px 10px;"
            )
        form.addRow("Номер участка:", self.inp_num)

        # Площадь участка (м²) — опционально, используется для расчёта ЧВ за м²
        area_raw = self._plot_data.get("area")
        area_text = ""
        if area_raw not in (None, "", 0):
            try:
                area_text = f"{float(area_raw):g}"
            except (TypeError, ValueError):
                area_text = str(area_raw)
        self.inp_area = QLineEdit(area_text)
        self.inp_area.setPlaceholderText("например: 612 (необязательно)")
        form.addRow("Площадь, м²:", self.inp_area)
        lay.addLayout(form)

        # Собственники
        own_label = QLabel("Собственники:")
        own_label.setStyleSheet("color:#9CA3AF;")
        lay.addWidget(own_label)

        self._owners_container = QWidget()
        self._owners_container.setStyleSheet("background:transparent;")
        self._owners_vlay = QVBoxLayout(self._owners_container)
        self._owners_vlay.setSpacing(6)
        self._owners_vlay.setContentsMargins(0, 0, 0, 0)

        existing_owners = self._plot_data.get("owners", [""])
        if not existing_owners:
            existing_owners = [""]
        for name in existing_owners:
            self._add_owner_field(name)

        lay.addWidget(self._owners_container)

        btn_add_owner = QPushButton("＋  Добавить собственника")
        btn_add_owner.setObjectName("btnSecondary")
        btn_add_owner.clicked.connect(lambda: self._add_owner_field(""))
        lay.addWidget(btn_add_owner)

        # Разделитель
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color:#E5E7EB;background:#E5E7EB;max-height:1px;")
        lay.addWidget(sep)

        # Кнопки OK / Cancel
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Сохранить")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("Отмена")
        btns.accepted.connect(self._on_accept)
        btns.rejected.connect(self.reject)
        lay.addWidget(btns)

    def _add_owner_field(self, name: str):
        row = QWidget()
        row.setStyleSheet("background:transparent;")
        rlay = QHBoxLayout(row)
        rlay.setContentsMargins(0, 0, 0, 0)
        rlay.setSpacing(6)

        inp = QLineEdit(name)
        inp.setPlaceholderText("Фамилия Имя Отчество")
        self._owner_inputs.append(inp)
        rlay.addWidget(inp, stretch=1)

        btn = QPushButton("✕")
        btn.setFixedSize(28, 28)
        btn.setStyleSheet(
            "QPushButton{background:#2a1a1a;border:1px solid #5a2a2a;"
            "border-radius:5px;color:#DC2626;font-size:12px;}"
            "QPushButton:hover{background:#3a2020;}"
        )
        btn.clicked.connect(lambda _, r=row, i=inp: self._remove_owner_field(r, i))
        rlay.addWidget(btn)
        self._owners_vlay.addWidget(row)

    def _remove_owner_field(self, row: QWidget, inp: QLineEdit):
        if len(self._owner_inputs) <= 1:
            inp.clear()
            return
        if inp in self._owner_inputs:
            self._owner_inputs.remove(inp)
        row.setParent(None)
        row.deleteLater()

    def _on_accept(self):
        num = self.inp_num.text().strip()
        if not num:
            QMessageBox.warning(self, "Ошибка", "Укажите номер участка")
            return
        owners = [i.text().strip() for i in self._owner_inputs if i.text().strip()]

        area_raw = self.inp_area.text().strip().replace(",", ".")
        area_val: float | None = None
        if area_raw:
            try:
                area_val = float(area_raw)
                if area_val <= 0:
                    raise ValueError("non-positive")
            except ValueError:
                QMessageBox.warning(self, "Ошибка",
                                    "Площадь должна быть положительным числом")
                return

        result = {"num": num, "owners": owners}
        if area_val is not None:
            result["area"] = area_val
        self._result = result
        self.accept()

    def get_result(self) -> dict:
        return getattr(self, "_result", {})

    def _apply_styles(self):
        self.setStyleSheet("""
            QDialog { background: #FFFFFF; color: #374151; }
            QLabel  { background: transparent; color: #374151; font-size: 13px; }
            QLineEdit {
                background: #F8F9FA; border: 1px solid #D1D5DB;
                border-radius: 5px; color: #374151; padding: 7px 10px; font-size: 13px;
            }
            QLineEdit:focus { border: 1px solid #6366F1; }
            QPushButton#btnSecondary {
                background: #E5E7EB; color: #6B7280; border: 1px solid #D1D5DB;
                border-radius: 6px; padding: 7px 14px; font-size: 13px;
            }
            QPushButton#btnSecondary:hover { background: #E5E7EB; color: #374151; }
            QDialogButtonBox QPushButton {
                background: #4F46E5; color: white; border: none;
                border-radius: 6px; padding: 8px 20px; font-size: 13px; font-weight: 600;
            }
            QDialogButtonBox QPushButton:hover { background: #6366F1; }
            QDialogButtonBox QPushButton[text='Отмена'] {
                background: #E5E7EB; color: #6B7280;
            }
        """)


class DocCell(QWidget):
    """
    Ячейка документа: иконка-статус (✔️/—) + кнопка скрепки для прикрепления файла.
    Эмитит сигнал при изменении.
    """
    changed = pyqtSignal()

    def __init__(self, plot_num: str, doc_key: str,
                 file_path: str = "", parent=None):
        super().__init__(parent)
        self._plot = plot_num
        self._doc_key = doc_key
        self._path = file_path

        lay = QHBoxLayout(self)
        lay.setContentsMargins(6, 2, 6, 2)
        lay.setSpacing(6)

        self.lbl_status = QLabel()
        self.lbl_status.setFixedWidth(20)
        self.lbl_status.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(self.lbl_status)

        self.btn_attach = QPushButton()
        self.btn_attach.setFixedSize(28, 28)
        self.btn_attach.setToolTip("Прикрепить / заменить файл")
        self.btn_attach.clicked.connect(self._on_attach)
        lay.addWidget(self.btn_attach)

        # Кнопка открыть файл
        self.btn_open = QPushButton("↗️")
        self.btn_open.setFixedSize(28, 28)
        self.btn_open.setToolTip("Открыть прикреплённый файл")
        self.btn_open.clicked.connect(self._on_open)
        lay.addWidget(self.btn_open)

        lay.addStretch()
        self._refresh()

    def _refresh(self):
        has = bool(self._path and os.path.exists(self._path))
        if has:
            self.lbl_status.setText("✔️")
            self.lbl_status.setStyleSheet("color:#059669;font-size:14px;font-weight:700;")
            self.btn_attach.setText("📎")
            self.btn_attach.setStyleSheet(
                "QPushButton{background:#0d3b1a;border:1px solid #2e7d32;"
                "border-radius:5px;font-size:13px;}"
                "QPushButton:hover{background:#1b5e20;}"
            )
            self.btn_open.setEnabled(True)
            self.btn_open.setStyleSheet(
                "QPushButton{background:#F0F2F5;border:1px solid #D1D5DB;"
                "border-radius:5px;color:#6366F1;font-size:12px;}"
                "QPushButton:hover{background:#E5E7EB;}"
            )
        else:
            self.lbl_status.setText("—")
            self.lbl_status.setStyleSheet("color:#6B7280;font-size:14px;font-weight:700;")
            self.btn_attach.setText("📎")
            self.btn_attach.setStyleSheet(
                "QPushButton{background:#F0F2F5;border:1px solid #D1D5DB;"
                "border-radius:5px;font-size:13px;}"
                "QPushButton:hover{background:#E5E7EB;}"
            )
            self.btn_open.setEnabled(False)
            self.btn_open.setStyleSheet(
                "QPushButton{background:#0d1720;border:1px solid #1b2a3c;"
                "border-radius:5px;color:#9CA3AF;font-size:12px;}"
            )

    def _on_attach(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Прикрепить файл", "", "Все файлы (*.*)"
        )
        if not path:
            return
        self._path = path
        self._refresh()
        self.changed.emit()

    def _on_open(self):
        if not self._path or not os.path.exists(self._path):
            QMessageBox.warning(self, "Ошибка", "Файл не найден")
            return
        try:
            os.startfile(self._path)
        except Exception:
            QMessageBox.warning(self, "Ошибка", "Не удалось открыть файл")

    def get_path(self) -> str:
        return self._path


class DocsWidget(QWidget):
    """Вкладка документов: таблица по участку и типу документа."""

    DATA_FILE = os.path.join(DATA_DIR, "snt_docs.json")
    DOC_TYPES = [
        "Паспорт",
        "Договор",
        "Схема участка",
        "Прочие документы",
    ]

    def __init__(self):
        super().__init__()
        self.setAutoFillBackground(True)
        self._docs = self._load()
        self._cells: dict[tuple[str, str], DocCell] = {}
        self._setup_ui()
        self._rebuild_table()

    def _load(self) -> dict:
        try:
            if os.path.exists(self.DATA_FILE):
                with open(self.DATA_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            pass
        return {}

    def _save(self):
        try:
            with open(self.DATA_FILE, "w", encoding="utf-8") as f:
                json.dump(self._docs, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def reload(self):
        """Перечитывает snt_docs.json и перестраивает таблицу."""
        self._docs = self._load()
        self._cells.clear()
        self._rebuild_table()

    def refresh_plots(self):
        """Перестраивает таблицу при изменении списка участков."""
        self._cells.clear()
        self._rebuild_table()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(14)

        top_bar = QHBoxLayout()
        title = QLabel("Документы")
        title.setObjectName("pageTitle")
        top_bar.addWidget(title)
        top_bar.addStretch()

        btn_save = QPushButton("Сохранить")
        btn_save.setObjectName("btnPrimary")
        btn_save.clicked.connect(self._save)
        top_bar.addWidget(btn_save)
        layout.addLayout(top_bar)

        self.status_label = QLabel("Документы не загружены", objectName="statusLabel")
        layout.addWidget(self.status_label)

        self.table = QTableWidget(objectName="mainTable")
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(False)
        layout.addWidget(self.table)

    def _rebuild_table(self):
        self.table.blockSignals(True)
        self.table.clearContents()

        plot_order = _load_plot_order()
        rows = len(plot_order)
        cols = 1 + len(self.DOC_TYPES)
        self.table.setRowCount(rows)
        self.table.setColumnCount(cols)

        headers = ["Участок"] + self.DOC_TYPES
        self.table.setHorizontalHeaderLabels(headers)

        for r_idx, plot in enumerate(plot_order):
            plot_item = QTableWidgetItem(f"уч. {plot}")
            plot_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            plot_item.setForeground(QColor("#6366F1"))
            f = plot_item.font(); f.setBold(True); plot_item.setFont(f)
            self.table.setItem(r_idx, 0, plot_item)

            plot_docs = self._docs.get(str(plot), {})
            for c_idx, doc_key in enumerate(self.DOC_TYPES, start=1):
                cell = DocCell(str(plot), doc_key, plot_docs.get(doc_key, ""), self)
                cell.changed.connect(
                    lambda _, p=str(plot), d=doc_key, w=cell: self._on_doc_changed(p, d, w)
                )
                self.table.setCellWidget(r_idx, c_idx, cell)
                self._cells[(str(plot), doc_key)] = cell
            self.table.setRowHeight(r_idx, 34)

        self.table.blockSignals(False)
        self._update_status()

    def _on_doc_changed(self, plot: str, doc_key: str, cell: DocCell):
        self._docs.setdefault(str(plot), {})[doc_key] = cell.get_path()
        self._save()
        self._update_status()

    def _update_status(self):
        total = 0
        attached = 0
        for plot in _load_plot_order():
            for doc_key in self.DOC_TYPES:
                path = self._docs.get(str(plot), {}).get(doc_key, "")
                total += 1
                if path:
                    attached += 1
        self.status_label.setText(
            f"Документов: {attached} из {total} прикреплено"
        )


from ui.map_widget import MapWidget
from ui.home_widget import HomeWidget



# ======================================================================= #
#  ВКЛАДКА «ДОЛГИ ПО ЭЛЕКТРОЭНЕРГИИ»
# ======================================================================= #

class _NumItem(QTableWidgetItem):
    """QTableWidgetItem с числовой сортировкой."""
    def __init__(self, text: str, value: float):
        super().__init__(text)
        self._value = value

    def __lt__(self, other):
        try:
            return self._value < other._value
        except AttributeError:
            return super().__lt__(other)


class MeterReplacementDialog(QDialog):
    """Регистрация замены счётчика для участка."""

    def __init__(self, plot: str, parent=None,
                 prev_value: float | None = None):
        super().__init__(parent)
        self.setWindowTitle(f"Замена счётчика — уч. {plot}")
        self.setMinimumWidth(420)
        self.setModal(True)
        self._plot = plot
        self._result: dict | None = None

        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 20, 20, 18)
        lay.setSpacing(12)

        title = QLabel(f"Замена счётчика на участке {plot}")
        title.setStyleSheet("font-size:14px;font-weight:700;color:#111827;")
        lay.addWidget(title)

        form = QFormLayout()
        form.setSpacing(8)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        self.inp_date = QDateEdit(calendarPopup=True)
        self.inp_date.setDisplayFormat("dd.MM.yyyy")
        self.inp_date.setDate(QDate.currentDate())
        form.addRow("Дата замены:", self.inp_date)

        self.inp_old = QLineEdit("" if prev_value is None else f"{prev_value:g}")
        self.inp_old.setPlaceholderText("конечное показание старого счётчика")
        form.addRow("Старый счётчик (конечн.):", self.inp_old)

        self.inp_new = QLineEdit("0")
        self.inp_new.setPlaceholderText("начальное показание нового счётчика")
        form.addRow("Новый счётчик (нач.):", self.inp_new)

        self.inp_note = QLineEdit()
        self.inp_note.setPlaceholderText("замена по сроку поверки, срыв пломбы и т.д.")
        form.addRow("Примечание:", self.inp_note)

        lay.addLayout(form)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Сохранить")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("Отмена")
        btns.accepted.connect(self._on_accept)
        btns.rejected.connect(self.reject)
        lay.addWidget(btns)

        self.setStyleSheet("""
            QDialog { background: #FFFFFF; color: #374151; }
            QLabel { background: transparent; color: #374151; font-size: 13px; }
            QLineEdit, QDateEdit {
                background: #F8F9FA; border: 1px solid #D1D5DB;
                border-radius: 5px; color: #374151; padding: 6px 8px; font-size: 13px;
            }
            QLineEdit:focus, QDateEdit:focus { border: 1px solid #6366F1; }
            QDialogButtonBox QPushButton {
                background: #4F46E5; color: white; border: none;
                border-radius: 6px; padding: 7px 18px; font-size: 13px; font-weight: 600;
            }
            QDialogButtonBox QPushButton:hover { background: #6366F1; }
            QDialogButtonBox QPushButton[text='Отмена'] {
                background: #E5E7EB; color: #6B7280;
            }
        """)

    def _on_accept(self):
        try:
            old_v = float(self.inp_old.text().strip().replace(",", "."))
            new_v = float(self.inp_new.text().strip().replace(",", "."))
        except ValueError:
            QMessageBox.warning(self, "Ошибка", "Показания должны быть числами")
            return
        if old_v < 0 or new_v < 0:
            QMessageBox.warning(self, "Ошибка", "Показания не могут быть отрицательными")
            return
        self._result = {
            "date": self.inp_date.date().toString("yyyy-MM-dd"),
            "old_final": f"{old_v:g}",
            "new_initial": f"{new_v:g}",
            "note": self.inp_note.text().strip(),
        }
        self.accept()

    def get_result(self) -> dict | None:
        return self._result


class PlotCardDialog(QDialog):
    """Карточка участка: сводка показаний/начислений/платежей + ввод и правка показаний."""

    MONTH_NAMES = ["янв", "фев", "мар", "апр", "май", "июн",
                   "июл", "авг", "сен", "окт", "ноя", "дек"]

    def __init__(self, plot: str, df, parent=None, as_of: date | None = None):
        super().__init__(parent)
        self._plot = str(plot)
        self._df = df
        self._as_of = as_of or date.today()
        self._value_edits: dict[tuple[int, int], QLineEdit] = {}
        self.setWindowTitle(f"Участок {plot} — карточка")
        self.setMinimumSize(960, 620)
        self.setModal(False)

        self._setup_ui()
        self._rebuild()

    # ── UI ───────────────────────────────────────────────────────────────
    def _setup_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(18, 18, 18, 16)
        lay.setSpacing(10)

        owners = energy.owners_map().get(self._plot, [])
        owners_text = ", ".join(owners) if owners else "владельцы не указаны"
        head = QLabel(f"<b>Участок {self._plot}</b>  ·  {owners_text}")
        head.setStyleSheet("font-size:14px;color:#111827;background:transparent;")
        lay.addWidget(head)

        self.summary_lbl = QLabel("")
        self.summary_lbl.setStyleSheet("color:#9CA3AF;background:transparent;font-size:12px;")
        lay.addWidget(self.summary_lbl)

        # ── Панель ввода нового показания ──────────────────────────
        entry = QFrame(objectName="entryBox")
        eh = QHBoxLayout(entry)
        eh.setContentsMargins(12, 8, 12, 8)
        eh.setSpacing(8)

        lbl = QLabel("Передать показание:")
        lbl.setStyleSheet("color:#6B7280;background:transparent;font-size:12px;")
        eh.addWidget(lbl)

        self.cb_month = QComboBox()
        self.cb_month.setObjectName("filterCombo")
        for i, m in enumerate(self.MONTH_NAMES, start=1):
            self.cb_month.addItem(m, i)
        today = date.today()
        self.cb_month.setCurrentIndex(today.month - 1)
        self.cb_month.setFixedWidth(80)
        eh.addWidget(self.cb_month)

        self.cb_year = QComboBox()
        self.cb_year.setObjectName("filterCombo")
        for y in range(today.year - 5, today.year + 2):
            self.cb_year.addItem(str(y), y)
        self.cb_year.setCurrentText(str(today.year))
        self.cb_year.setFixedWidth(90)
        eh.addWidget(self.cb_year)

        self.le_value = QLineEdit()
        self.le_value.setObjectName("searchInput")
        self.le_value.setPlaceholderText("значение, кВт·ч")
        self.le_value.setFixedWidth(160)
        self.le_value.returnPressed.connect(self._on_add_reading)
        eh.addWidget(self.le_value)

        btn_add = QPushButton("Сохранить", objectName="btnPrimary")
        btn_add.clicked.connect(self._on_add_reading)
        eh.addWidget(btn_add)
        eh.addStretch()
        lay.addWidget(entry)

        # ── Таблица ────────────────────────────────────────────────
        self.table = QTableWidget(objectName="summaryTable")
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels([
            "Дата / Месяц", "Показание", "Расход (кВт·ч)",
            "Тариф", "Начислено", "Оплачено",
            "Изм. баланса", "Баланс нараст.", "",
        ])
        hdr = self.table.horizontalHeader()
        for c, w in enumerate([90, 150, 110, 75, 110, 110, 110, 130, 40]):
            hdr.setSectionResizeMode(c, QHeaderView.ResizeMode.Interactive)
            self.table.setColumnWidth(c, w)
        hdr.setStretchLastSection(False)
        lay.addWidget(self.table, 1)

        # ── Подсказка по аномалиям ─────────────────────────────────
        hint = QHBoxLayout()
        hint.setSpacing(16)
        for color, text in [
            ("#DC2626", "■  показание < предыдущего"),
            ("#ffd54f", "■  аномально большой расход / замена счётчика"),
        ]:
            lb = QLabel(text)
            lb.setStyleSheet(f"color:{color};background:transparent;font-size:11px;")
            hint.addWidget(lb)
        hint.addStretch()
        lay.addLayout(hint)

        # ── Кнопки внизу ───────────────────────────────────────────
        bottom = QHBoxLayout()
        self.btn_replace = QPushButton("Зарегистрировать замену счётчика")
        self.btn_replace.setObjectName("btnSecondary")
        self.btn_replace.clicked.connect(self._on_replace)
        bottom.addWidget(self.btn_replace)

        self.btn_pdf = QPushButton("📄  Сохранить PDF-квитанцию")
        self.btn_pdf.setObjectName("btnSecondary")
        self.btn_pdf.clicked.connect(self._on_pdf)
        bottom.addWidget(self.btn_pdf)

        bottom.addStretch()

        btn_close = QPushButton("Закрыть")
        btn_close.setObjectName("btnPrimary")
        btn_close.clicked.connect(self.accept)
        bottom.addWidget(btn_close)
        lay.addLayout(bottom)

        self.setStyleSheet("""
            QDialog { background: #FFFFFF; color: #374151; }
            QLabel  { background: transparent; }
            QFrame#entryBox {
                background: #F8F9FA; border: 1px solid #E5E7EB; border-radius: 8px;
            }
            QPushButton#btnPrimary {
                background: #4F46E5; color: white; border: none; border-radius: 6px;
                padding: 8px 18px; font-size: 13px; font-weight: 600;
            }
            QPushButton#btnPrimary:hover  { background: #6366F1; }
            QPushButton#btnSecondary {
                background: #E5E7EB; color: #6B7280; border: 1px solid #D1D5DB;
                border-radius: 6px; padding: 7px 14px; font-size: 13px;
            }
            QPushButton#btnSecondary:hover { background: #E5E7EB; color: #374151; }
            QLineEdit#searchInput, QComboBox#filterCombo {
                background: #F8F9FA; border: 1px solid #D1D5DB; border-radius: 6px;
                color: #374151; padding: 6px 10px; font-size: 13px;
            }
            QLineEdit#searchInput:focus { border: 1px solid #6366F1; }
            QComboBox#filterCombo::drop-down { border: none; width: 18px; }
            QComboBox QAbstractItemView {
                background: #F8F9FA; border: 1px solid #D1D5DB;
                color: #374151; selection-background-color: #EEF2FF;
            }
            QTableWidget#summaryTable {
                background: #F8F9FA; border: 1px solid #E5E7EB; border-radius: 8px;
                gridline-color: #F3F4F6; color: #374151; font-size: 12px;
                selection-background-color: #EEF2FF; selection-color: #111827;
            }
            QTableWidget#summaryTable QHeaderView::section {
                background: #F9FAFB; color: #6366F1; border: none;
                border-right: 1px solid #E5E7EB; border-bottom: 2px solid #6366F1;
                padding: 6px 8px; font-size: 12px; font-weight: 600;
            }
        """)

    # ── Перестройка таблицы ──────────────────────────────────────────────
    def _rebuild(self):
        self._value_edits.clear()

        meters = energy.load_meters()
        rates = energy.load_rates()
        repls = energy.load_replacements()
        baseline = energy.load_baseline()
        if self._df is not None and not self._df.empty:
            baseline["start_date"] = self._df["Дата"].min().date().isoformat()

        base = energy._to_float(baseline.get("balances", {}).get(self._plot)) or 0.0
        base_start = energy._parse_iso(baseline.get("start_date", ""))

        # ── Хронология: каждое снятие показания + каждый платёж — отдельная строка
        charges = energy.all_charges(self._plot, meters, rates, repls, up_to=self._as_of)
        payments = [
            p for p in energy.payments_breakdown(self._plot, self._df)
            if p["date"] is not None
            and (base_start is None or p["date"] >= base_start)
            and p["date"] <= self._as_of
        ]
        events: list[tuple[str, date, dict]] = []
        for c in charges:
            events.append(("charge", energy.reading_date(c["year"], c["month"]), c))
        for p in payments:
            events.append(("payment", p["date"], p))
        for repl in repls.get(self._plot, []):
            d = energy._parse_iso(repl.get("date", ""))
            if d and d <= self._as_of:
                events.append(("replacement", d, repl))
        # charge < replacement < payment при совпадении даты
        _kind_order = {"charge": 0, "replacement": 1, "payment": 2}
        events.sort(key=lambda e: (e[1], _kind_order.get(e[0], 9)))

        anomaly_map: dict[tuple[int, int], str] = {}
        for a in energy.anomalies(self._plot, meters, repls):
            if a.type in ("drop", "spike"):
                anomaly_map[(a.year, a.month)] = a.type

        self.table.setRowCount(len(events) + (1 if base != 0 else 0))
        r0 = 0
        cum = base
        if base != 0:
            label = "Начальное сальдо"
            if base_start:
                label += f" ({base_start.isoformat()})"
            self._set_row(0, [
                (label, None), ("—", None), ("—", None), ("—", None),
                ("—", None), ("—", None),
                (fmt_money(base), None),
                (fmt_money(base), self._debt_color(base)),
                ("", None),
            ], bold=True)
            r0 = 1

        for i, (kind, evdate, payload) in enumerate(events):
            r = r0 + i
            if kind == "charge":
                c = payload
                y, m = c["year"], c["month"]
                label = f"{m:02d}.{y}"
                kwh_text = f"{c['kwh']:.0f}" if c["kwh"] is not None else "—"
                rate_text = f"{c['rate']:.2f}" if c["rate"] is not None else "—"
                amount = c["amount"] or 0.0
                charged_text = fmt_money(c["amount"]) if c["amount"] is not None else "—"
                cum += amount
                mbal_text = fmt_money(amount) if amount else "0 ₽"
                self._set_row(r, [
                    (label, None),
                    None,                              # «Показание» — редактируемое поле
                    (kwh_text, None),
                    (rate_text, None),
                    (charged_text, "#f9a825" if c["amount"] else None),
                    ("—", None),
                    (mbal_text, None),
                    (fmt_money(cum), self._debt_color(cum)),
                    None,                              # кнопка удаления
                ])
                self._install_value_editor(r, y, m, c["value"], anomaly_map.get((y, m)))
                self._install_delete_button(r, y, m)
            elif kind == "payment":
                p = payload
                paid = p["amount"]
                cum -= paid
                label = f"💳  {p['date'].strftime('%d.%m.%Y')}"
                if p.get("mixed"):
                    label += " ⅟₂"   # подсказка что платёж пополам с членскими
                self._set_row(r, [
                    (label, "#059669"),
                    ("—", None),
                    ("—", None),
                    ("—", None),
                    ("—", None),
                    (fmt_money(paid), "#059669"),
                    (fmt_money(-paid), None),
                    (fmt_money(cum), self._debt_color(cum)),
                    ("", None),
                ])
                # tooltip с назначением платежа на всю строку
                if p.get("purpose"):
                    for col in range(self.table.columnCount()):
                        it = self.table.item(r, col)
                        if it is not None:
                            it.setToolTip(p["purpose"])
            else:  # replacement
                repl = payload
                old_f = energy._to_float(repl.get("old_final"))
                new_i = energy._to_float(repl.get("new_initial"))
                reading_text = (
                    f"{old_f:g} → {new_i:g}"
                    if old_f is not None and new_i is not None else "—"
                )
                self._set_row(r, [
                    (f"[замена] {evdate.strftime('%d.%m.%Y')}", "#ffd54f"),
                    (reading_text, "#ffd54f"),
                    ("—", None), ("—", None), ("—", None), ("—", None), ("—", None),
                    (fmt_money(cum), self._debt_color(cum)),
                    None,
                ])
                note = repl.get("note", "").strip()
                tip = f"Замена счётчика: конечное {old_f:g}, начальное {new_i:g}"
                if note:
                    tip += f"\nПричина: {note}"
                for col in range(self.table.columnCount()):
                    it = self.table.item(r, col)
                    if it is not None:
                        it.setToolTip(tip)
                self._install_delete_replacement_button(r, repl.get("date", ""))

        # Итоги — через energy.balance(), чтобы совпадали с вкладкой «Долги»
        bal = energy.balance(self._plot, self._as_of, meters, rates, repls, baseline, self._df)
        if events or base != 0:
            self.summary_lbl.setText(
                f"Начислено всего: {fmt_money(bal.baseline + bal.charged)}  ·  "
                f"оплачено всего: {fmt_money(bal.paid)}  ·  "
                f"итоговый баланс: {fmt_money(bal.debt)}  ·  "
                f"на {self._as_of.strftime('%d.%m.%Y')}"
            )
        else:
            self.summary_lbl.setText(
                "Показаний и платежей по этому участку пока нет — внесите первое выше."
            )

    # ── Помощники строк ──────────────────────────────────────────────────
    def _set_row(self, r: int, cells: list, bold: bool = False):
        for c, cell in enumerate(cells):
            if cell is None:
                # Placeholder под ячейку-виджет: пустой item, чтобы сохранить выделение строки
                ph = QTableWidgetItem("")
                ph.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                self.table.setItem(r, c, ph)
                continue
            text, color = cell
            it = QTableWidgetItem(text)
            it.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            if color:
                it.setForeground(QColor(color))
            if bold:
                f = it.font(); f.setBold(True); it.setFont(f)
            it.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            self.table.setItem(r, c, it)
        self.table.setRowHeight(r, 30)

    def _install_value_editor(self, r: int, year: int, month: int,
                              value: float, anomaly: str | None):
        edit = QLineEdit(f"{value:g}")
        edit.setAlignment(Qt.AlignmentFlag.AlignCenter)
        edit.setStyleSheet(self._value_edit_style(anomaly))
        if anomaly == "drop":
            edit.setToolTip("Показание меньше предыдущего")
        elif anomaly == "spike":
            edit.setToolTip("Расход существенно больше обычного")
        edit.editingFinished.connect(
            lambda y=year, m=month, e=edit: self._on_value_committed(y, m, e)
        )
        self._value_edits[(year, month)] = edit
        self.table.setCellWidget(r, 1, edit)

    def _install_delete_button(self, r: int, year: int, month: int):
        btn = QPushButton("✕")
        btn.setFixedSize(26, 22)
        btn.setToolTip(f"Удалить показание за {month:02d}.{year}")
        btn.setStyleSheet(
            "QPushButton{background:#2a1318;color:#DC2626;border:1px solid #6e2a30;"
            "border-radius:4px;font-size:12px;font-weight:bold;}"
            "QPushButton:hover{background:#4a1a22;color:#ffcccc;}"
        )
        btn.clicked.connect(lambda _, y=year, m=month: self._on_delete_reading(y, m))
        self.table.setCellWidget(r, 8, btn)

    def _install_delete_replacement_button(self, r: int, repl_date: str):
        btn = QPushButton("✕")
        btn.setFixedSize(26, 22)
        btn.setToolTip(f"Удалить запись о замене счётчика от {repl_date}")
        btn.setStyleSheet(
            "QPushButton{background:#2a2200;color:#ffd54f;border:1px solid #7a6000;"
            "border-radius:4px;font-size:12px;font-weight:bold;}"
            "QPushButton:hover{background:#4a3c00;color:#ffe57f;}"
        )
        btn.clicked.connect(lambda _, d=repl_date: self._on_delete_replacement(d))
        self.table.setCellWidget(r, 8, btn)

    def _on_delete_replacement(self, repl_date: str):
        reply = QMessageBox.question(
            self, "Удалить замену счётчика",
            f"Удалить запись о замене счётчика от {repl_date} на уч. {self._plot}?\n"
            "Расчёт расхода электроэнергии будет пересчитан.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        repls = energy.load_replacements()
        plot_repls = repls.get(self._plot, [])
        repls[self._plot] = [r for r in plot_repls if r.get("date") != repl_date]
        if not repls[self._plot]:
            del repls[self._plot]
        energy.save_replacements(repls)
        self._rebuild()

    @staticmethod
    def _value_edit_style(anomaly: str | None) -> str:
        if anomaly == "drop":
            return ("background:#2a0d0d;border:1px solid #DC2626;border-radius:3px;"
                    "color:#DC2626;font-size:12px;font-weight:700;padding:3px 6px;")
        if anomaly == "spike":
            return ("background:#2a1f0d;border:1px solid #f9a825;border-radius:3px;"
                    "color:#ffd54f;font-size:12px;padding:3px 6px;")
        return ("background:#F0F2F5;border:1px solid #D1D5DB;border-radius:3px;"
                "color:#374151;font-size:12px;padding:3px 6px;")

    # ── Сохранение ───────────────────────────────────────────────────────
    def _on_add_reading(self):
        month = self.cb_month.currentData()
        year = self.cb_year.currentData()
        val = self.le_value.text().strip().replace(",", ".")
        if not val:
            return
        try:
            num = float(val)
        except ValueError:
            QMessageBox.warning(self, "Показание", "Значение должно быть числом.")
            return
        if num < 0:
            QMessageBox.warning(self, "Показание", "Значение не может быть отрицательным.")
            return
        self._store_reading(year, month, num)
        self.le_value.clear()
        self.le_value.setFocus()

    def _on_value_committed(self, year: int, month: int, edit: QLineEdit):
        text = edit.text().strip().replace(",", ".")
        meters = energy.load_meters()
        key = f"{self._plot}:{year}:{month}"
        old_raw = meters.get(key, "")
        if text == "" and old_raw == "":
            return
        if text == old_raw.strip().replace(",", "."):
            return
        if text == "":
            # очистка ячейки = удаление показания
            if self._confirm_delete(year, month):
                meters.pop(key, None)
                energy.save_meters(meters)
                self._rebuild()
            else:
                self._rebuild()  # восстановить старое значение в виджете
            return
        try:
            num = float(text)
        except ValueError:
            QMessageBox.warning(self, "Показание", "Значение должно быть числом.")
            self._rebuild()
            return
        if num < 0:
            QMessageBox.warning(self, "Показание", "Значение не может быть отрицательным.")
            self._rebuild()
            return
        self._store_reading(year, month, num)

    def _store_reading(self, year: int, month: int, value: float):
        meters = energy.load_meters()
        key = f"{self._plot}:{year}:{month}"
        # сохраняем как целое если без дробной части
        meters[key] = str(int(value)) if float(value).is_integer() else f"{value:g}"
        energy.save_meters(meters)
        self._rebuild()

    def _on_delete_reading(self, year: int, month: int):
        if not self._confirm_delete(year, month):
            return
        meters = energy.load_meters()
        meters.pop(f"{self._plot}:{year}:{month}", None)
        energy.save_meters(meters)
        self._rebuild()

    def _confirm_delete(self, year: int, month: int) -> bool:
        reply = QMessageBox.question(
            self, "Удалить показание",
            f"Удалить показание за {month:02d}.{year} на уч. {self._plot}?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        return reply == QMessageBox.StandardButton.Yes

    # ── Прочее ───────────────────────────────────────────────────────────
    @staticmethod
    def _debt_color(v: float) -> str | None:
        if v > 0:
            return "#DC2626"
        if v < 0:
            return "#059669"
        return None

    def _on_replace(self):
        meters = energy.load_meters()
        readings = energy.plot_readings(self._plot, meters)
        last_val = readings[-1][2] if readings else None
        dlg = MeterReplacementDialog(self._plot, self, prev_value=last_val)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        result = dlg.get_result()
        if not result:
            return
        repls = energy.load_replacements()
        repls.setdefault(self._plot, []).append(result)
        repls[self._plot].sort(key=lambda r: r.get("date", ""))
        energy.save_replacements(repls)
        self._rebuild()
        QMessageBox.information(self, "Замена счётчика",
                                "Замена сохранена. Расчёт пересчитан.")

    def _on_pdf(self):
        try:
            from core import receipt
        except ImportError:
            QMessageBox.information(self, "Квитанции",
                                    "Модуль квитанций ещё не подключён.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить квитанцию", f"Уч_{self._plot}.pdf",
            "PDF (*.pdf)"
        )
        if not path:
            return
        try:
            receipt.save_plot_receipt_pdf(self._plot, self._df, path)
            QMessageBox.information(self, "Квитанция", f"Сохранено:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить:\n{e}")


class VznosyAdjustmentDialog(QDialog):
    """Диалог добавления ручного платежа или переопределения начисления ЧВ."""

    KIND_LABELS = {
        "payment_manual": "Ручной платёж",
        "charge_override": "Переопределение начисления за период",
        "exempt_period": "Освобождение от взноса за период",
    }

    def __init__(self, plot: str, parent=None,
                 default_kind: str = "payment_manual"):
        super().__init__(parent)
        self._plot = plot
        self._result: dict | None = None
        self.setWindowTitle(f"Корректировка ЧВ — уч. {plot}")
        self.setMinimumWidth(480)
        self.setModal(True)

        from core import vznosy as _vznosy
        self._periods = _vznosy.build_periods(_vznosy.load_rates())

        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 20, 20, 18)
        lay.setSpacing(12)

        title = QLabel(f"Корректировка ЧВ на участке {plot}")
        title.setStyleSheet("font-size:14px;font-weight:700;color:#111827;")
        lay.addWidget(title)

        form = QFormLayout()
        form.setSpacing(8)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        self.cb_kind = QComboBox()
        for key, label in self.KIND_LABELS.items():
            self.cb_kind.addItem(label, key)
        idx = self.cb_kind.findData(default_kind)
        if idx >= 0:
            self.cb_kind.setCurrentIndex(idx)
        self.cb_kind.currentIndexChanged.connect(self._on_kind_changed)
        form.addRow("Тип:", self.cb_kind)

        self.inp_date = QDateEdit(calendarPopup=True)
        self.inp_date.setDisplayFormat("dd.MM.yyyy")
        self.inp_date.setDate(QDate.currentDate())
        form.addRow("Дата:", self.inp_date)

        self.cb_period = QComboBox()
        self._fill_period_combo()
        form.addRow("Период:", self.cb_period)

        self.inp_amount = QLineEdit()
        self.inp_amount.setPlaceholderText("например: 5000")
        form.addRow("Сумма, ₽:", self.inp_amount)

        self.inp_note = QLineEdit()
        self.inp_note.setPlaceholderText("необязательно")
        form.addRow("Примечание:", self.inp_note)

        lay.addLayout(form)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Сохранить")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("Отмена")
        btns.accepted.connect(self._on_accept)
        btns.rejected.connect(self.reject)
        lay.addWidget(btns)

        self.setStyleSheet("""
            QDialog { background: #FFFFFF; color: #374151; }
            QLabel { background: transparent; color: #374151; font-size: 13px; }
            QLineEdit, QDateEdit, QComboBox, QSpinBox {
                background: #F8F9FA; border: 1px solid #D1D5DB;
                border-radius: 5px; color: #374151; padding: 6px 8px; font-size: 13px;
            }
            QLineEdit:focus, QDateEdit:focus { border: 1px solid #6366F1; }
            QDialogButtonBox QPushButton {
                background: #4F46E5; color: white; border: none;
                border-radius: 6px; padding: 7px 18px; font-size: 13px; font-weight: 600;
            }
            QDialogButtonBox QPushButton:hover { background: #6366F1; }
            QDialogButtonBox QPushButton[text='Отмена'] {
                background: #E5E7EB; color: #6B7280;
            }
        """)

        self._on_kind_changed()

    def _fill_period_combo(self):
        self.cb_period.clear()
        for r in reversed(self._periods):   # новые сверху
            pf = r.get("date_from", "")
            pt = r.get("date_to", "")
            try:
                from datetime import datetime as _dt
                pf_label = _dt.strptime(pf, "%Y-%m-%d").strftime("%d.%m.%Y")
            except Exception:
                pf_label = pf
            try:
                from datetime import datetime as _dt
                pt_label = _dt.strptime(pt, "%Y-%m-%d").strftime("%d.%m.%Y")
            except Exception:
                pt_label = pt or "открытый"
            self.cb_period.addItem(f"{pf_label} — {pt_label}", pf)
        if not self._periods:
            self.cb_period.addItem("Нет периодов", "")

    def _on_kind_changed(self):
        kind = self.cb_kind.currentData()
        is_payment = (kind == "payment_manual")
        self.cb_period.setEnabled(not is_payment)
        if kind == "payment_manual":
            self.inp_amount.setEnabled(True)
            self.inp_amount.setPlaceholderText("например: 5000")
        elif kind == "charge_override":
            self.inp_amount.setEnabled(True)
            self.inp_amount.setPlaceholderText("новая сумма за период, ₽")
        else:  # exempt_period
            self.inp_amount.setEnabled(False)
            self.inp_amount.setText("0")

    def _on_accept(self):
        kind = self.cb_kind.currentData()
        result = {
            "date": self.inp_date.date().toString("yyyy-MM-dd"),
            "kind": kind,
            "note": self.inp_note.text().strip(),
        }
        if kind in ("charge_override", "exempt_period"):
            result["period_from"] = self.cb_period.currentData() or ""

        if kind == "exempt_period":
            result["amount"] = "0"
        else:
            raw = self.inp_amount.text().strip().replace(",", ".")
            try:
                v = float(raw)
                if v < 0:
                    raise ValueError
            except ValueError:
                QMessageBox.warning(self, "Ошибка",
                                    "Сумма должна быть неотрицательным числом")
                return
            result["amount"] = f"{v:g}"

        self._result = result
        self.accept()

    def get_result(self) -> dict | None:
        return self._result


class VznosyCardDialog(QDialog):
    """Карточка участка по членским взносам: годы, начисления, платежи, корректировки."""

    def __init__(self, plot: str, df, parent=None, as_of: date | None = None):
        super().__init__(parent)
        self._plot = str(plot)
        self._df = df
        self._as_of = as_of or date.today()
        self.setWindowTitle(f"Участок {plot} — карточка ЧВ")
        self.setMinimumSize(1010, 640)
        self.setModal(False)
        self._rebuilding = False   # защита от рекурсии при установке чекбоксов

        self._setup_ui()
        self._rebuild()

    def _setup_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(18, 18, 18, 16)
        lay.setSpacing(10)

        from core import vznosy
        owners = energy.owners_map().get(self._plot, [])
        owners_text = ", ".join(owners) if owners else "владельцы не указаны"
        area = vznosy.plot_area_map().get(self._plot)
        area_text = f"{area:g} м²" if area is not None else "площадь не указана"

        self.head = QLabel(
            f"<b>Участок {self._plot}</b>  ·  {owners_text}  ·  {area_text}"
        )
        self.head.setStyleSheet("font-size:14px;color:#111827;background:transparent;")
        lay.addWidget(self.head)

        self.summary_lbl = QLabel("")
        self.summary_lbl.setStyleSheet("color:#9CA3AF;background:transparent;font-size:12px;")
        lay.addWidget(self.summary_lbl)

        self.warning_lbl = QLabel("")
        self.warning_lbl.setStyleSheet(
            "color:#DC2626;background:#2a1318;border:1px solid #6e2a30;"
            "border-radius:6px;padding:8px 12px;font-size:12px;"
        )
        self.warning_lbl.setVisible(False)
        lay.addWidget(self.warning_lbl)

        # ── Главная таблица: разбивка по периодам ──────────────────────
        years_lbl = QLabel("Разбивка по периодам", objectName="filterLabel")
        years_lbl.setStyleSheet("color:#6366F1;background:transparent;font-size:12px;margin-top:4px;")
        lay.addWidget(years_lbl)

        self.table = QTableWidget(objectName="summaryTable")
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            "Период", "Тариф", "Начислено", "Оплачено за период",
            "Корректировка", "Баланс нараст.", "Не учитывать",
        ])
        hdr = self.table.horizontalHeader()
        for c, w in enumerate([200, 240, 120, 140, 190, 130, 100]):
            hdr.setSectionResizeMode(c, QHeaderView.ResizeMode.Interactive)
            self.table.setColumnWidth(c, w)
        hdr.setStretchLastSection(False)
        lay.addWidget(self.table, 1)

        # ── Список ручных операций и корректировок ────────────────
        adj_lbl = QLabel("Ручные операции и корректировки")
        adj_lbl.setStyleSheet("color:#6366F1;background:transparent;font-size:12px;margin-top:8px;")
        lay.addWidget(adj_lbl)

        self.adj_table = QTableWidget(objectName="summaryTable")
        self.adj_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.adj_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.adj_table.verticalHeader().setVisible(False)
        self.adj_table.setColumnCount(6)
        self.adj_table.setHorizontalHeaderLabels([
            "Дата", "Тип", "Период", "Сумма", "Примечание", "",
        ])
        ahdr = self.adj_table.horizontalHeader()
        for c, w in enumerate([110, 200, 80, 120, 380, 40]):
            ahdr.setSectionResizeMode(c, QHeaderView.ResizeMode.Interactive)
            self.adj_table.setColumnWidth(c, w)
        ahdr.setStretchLastSection(False)
        self.adj_table.setMaximumHeight(180)
        lay.addWidget(self.adj_table)

        # ── Кнопки внизу ──────────────────────────────────────────
        bottom = QHBoxLayout()
        btn_pay = QPushButton("Ручной платёж", objectName="btnSecondary")
        btn_pay.clicked.connect(lambda: self._add_adjustment("payment_manual"))
        bottom.addWidget(btn_pay)

        btn_over = QPushButton("🛠️  Переопределить начисление за период",
                               objectName="btnSecondary")
        btn_over.clicked.connect(lambda: self._add_adjustment("charge_override"))
        bottom.addWidget(btn_over)

        btn_pdf = QPushButton("📄  Сохранить PDF-квитанцию", objectName="btnSecondary")
        btn_pdf.clicked.connect(self._on_pdf)
        bottom.addWidget(btn_pdf)

        bottom.addStretch()
        btn_close = QPushButton("Закрыть", objectName="btnPrimary")
        btn_close.clicked.connect(self.accept)
        bottom.addWidget(btn_close)
        lay.addLayout(bottom)

        self.setStyleSheet("""
            QDialog { background: #FFFFFF; color: #374151; }
            QLabel  { background: transparent; }
            QPushButton#btnPrimary {
                background: #4F46E5; color: white; border: none; border-radius: 6px;
                padding: 8px 18px; font-size: 13px; font-weight: 600;
            }
            QPushButton#btnPrimary:hover  { background: #6366F1; }
            QPushButton#btnSecondary {
                background: #E5E7EB; color: #6B7280; border: 1px solid #D1D5DB;
                border-radius: 6px; padding: 7px 14px; font-size: 13px;
            }
            QPushButton#btnSecondary:hover { background: #E5E7EB; color: #374151; }
            QTableWidget#summaryTable {
                background: #F8F9FA; border: 1px solid #E5E7EB; border-radius: 8px;
                gridline-color: #F3F4F6; color: #374151; font-size: 12px;
                selection-background-color: #EEF2FF; selection-color: #111827;
            }
            QTableWidget#summaryTable QHeaderView::section {
                background: #F9FAFB; color: #6366F1; border: none;
                border-right: 1px solid #E5E7EB; border-bottom: 2px solid #6366F1;
                padding: 6px 8px; font-size: 12px; font-weight: 600;
            }
        """)

    def _rebuild(self):
        from core import vznosy
        rates = vznosy.load_rates()
        adj = vznosy.load_adjustments()
        area = vznosy.plot_area_map().get(self._plot)
        periods = vznosy.build_periods(rates)
        bal = vznosy.balance_for_plot(self._plot, area, self._as_of,
                                       rates, adj, self._df)
        py = vznosy.paid_by_period(self._plot, self._df, self._as_of, periods, adj)

        # Обновить заголовок (площадь могла измениться)
        owners = energy.owners_map().get(self._plot, [])
        owners_text = ", ".join(owners) if owners else "владельцы не указаны"
        area_text = f"{area:g} м²" if area is not None else "площадь не указана"
        self.head.setText(
            f"<b>Участок {self._plot}</b>  ·  {owners_text}  ·  {area_text}"
        )

        # Баннер о площади
        if bal.area_missing_warning:
            self.warning_lbl.setText(
                "⚠ Для этого участка не указана площадь — тариф «за м²» не применён. "
                "Откройте вкладку «Участки» и укажите площадь."
            )
            self.warning_lbl.setVisible(True)
        else:
            self.warning_lbl.setVisible(False)

        # Главная таблица периодов
        self.table.setRowCount(len(bal.breakdown))
        self._rebuilding = True
        try:
            cum = 0.0
            for r, y in enumerate(bal.breakdown):
                # Метка периода
                if y.period_to:
                    period_label = (f"{y.period_from.strftime('%d.%m.%Y')}"
                                    f"—{y.period_to.strftime('%d.%m.%Y')}")
                else:
                    period_label = f"{y.period_from.strftime('%d.%m.%Y')}—..."

                # Тариф
                if y.tariff is None:
                    tariff_text = "—"
                    tariff_color = "#9CA3AF"
                elif y.tariff.get("per_sqm"):
                    rate = y.tariff.get("rate_sqm", "?")
                    if y.area_missing:
                        tariff_text = f"{rate} ₽/м²  (площадь не указана)"
                        tariff_color = "#DC2626"
                    else:
                        tariff_text = f"{rate} ₽/м² · {area:g} м² → {fmt_money(y.amount)}"
                        tariff_color = "#374151"
                else:
                    tariff_text = f"{y.tariff.get('amount', '?')} ₽"
                    tariff_color = "#374151"

                # Начислено
                if y.ignored:
                    amount_text = "—"
                    amount_color = "#9CA3AF"
                elif y.amount is None:
                    amount_text = "—"
                    amount_color = "#DC2626" if y.area_missing else "#9CA3AF"
                else:
                    amount_text = fmt_money(y.amount)
                    amount_color = "#f9a825" if y.amount > 0 else "#9CA3AF"

                # Оплачено за период
                period_key = y.period_from.isoformat()
                paid_period = py.get(period_key, 0.0)
                paid_text = fmt_money(paid_period) if paid_period else "—"
                paid_color = "#059669" if paid_period else "#9CA3AF"

                # Корректировка
                note_text = "—"
                note_color = "#9CA3AF"
                if y.ignored:
                    note_text = "не учитывается"
                    note_color = "#546e7a"
                elif y.overridden:
                    ov = vznosy._period_override(self._plot, period_key,
                                                 y.period_from.year, adj)
                    kind = ov.get("kind", "") if ov else ""
                    kind_label = "освобождён" if kind in ("exempt_period", "exempt_year") else "переопределено"
                    if ov and ov.get("note"):
                        note_text = f"{kind_label}: {ov['note']}"
                    else:
                        note_text = kind_label
                    note_color = "#ffd54f"

                # Накопительный баланс
                if not y.ignored:
                    if y.amount is not None:
                        cum += y.amount
                    cum -= paid_period
                cum_text = fmt_money(cum) if not y.ignored else "—"
                cum_color = "#DC2626" if cum > 0 else ("#059669" if cum < 0 else None)

                self._set_year_row(r, [
                    (period_label, None),
                    (tariff_text, tariff_color),
                    (amount_text, amount_color),
                    (paid_text, paid_color),
                    (note_text, note_color),
                    (cum_text, cum_color),
                ], ignored=y.ignored)

                # Чекбокс «Не учитывать»
                cb_widget = QWidget()
                cb_widget.setStyleSheet("background: transparent;")
                cb_layout = QHBoxLayout(cb_widget)
                cb_layout.setContentsMargins(0, 0, 0, 0)
                cb_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
                cb = QCheckBox()
                cb.setChecked(y.ignored)
                cb.setToolTip(
                    "Снять флаг — период снова учитывается" if y.ignored
                    else "Поставить флаг — период будет исключён из расчёта"
                )
                cb.setStyleSheet("""
                    QCheckBox::indicator { width: 16px; height: 16px; }
                    QCheckBox::indicator:unchecked {
                        border: 2px solid #D1D5DB; border-radius: 3px;
                        background: #F8F9FA;
                    }
                    QCheckBox::indicator:checked {
                        border: 2px solid #DC2626; border-radius: 3px;
                        background: #4a1a22;
                    }
                """)
                cb.toggled.connect(
                    lambda checked, pk=period_key: self._toggle_ignore_period(pk, checked)
                )
                cb_layout.addWidget(cb)
                self.table.setCellWidget(r, 6, cb_widget)
        finally:
            self._rebuilding = False

        # Таблица корректировок
        plot_adjs = list(adj.get(self._plot, []) or [])
        plot_adjs_idx = list(enumerate(plot_adjs))
        plot_adjs_idx.sort(key=lambda t: t[1].get("date", ""))
        self.adj_table.setRowCount(len(plot_adjs_idx))
        for r, (orig_idx, a) in enumerate(plot_adjs_idx):
            kind = a.get("kind", "")
            kind_label = {
                "payment_manual": "Ручной платёж",
                "charge_override": "Переопр. начисления",
                "exempt_period": "Освобождение",
                "exempt_year": "Освобождение",
                "ignore_period": "Не учитывать",
                "ignore_year": "Не учитывать",
            }.get(kind, kind)
            date_text = a.get("date", "")
            try:
                from datetime import datetime as _dt
                date_text = _dt.strptime(date_text, "%Y-%m-%d").strftime("%d.%m.%Y")
            except Exception:
                pass
            # Период: period_from (новый формат) или year (старый)
            pf = a.get("period_from", "")
            if pf:
                try:
                    from datetime import datetime as _dt
                    period_text = _dt.strptime(pf, "%Y-%m-%d").strftime("%d.%m.%Y")
                except Exception:
                    period_text = pf
            elif a.get("year"):
                period_text = str(a["year"])
            else:
                period_text = "—"
            amount_v = energy._to_float(a.get("amount")) or 0.0
            amount_text = fmt_money(amount_v)
            note_text = a.get("note", "")

            for c, text in enumerate([date_text, kind_label, period_text,
                                       amount_text, note_text]):
                it = QTableWidgetItem(text)
                it.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                if c in (0, 2, 3):
                    it.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                self.adj_table.setItem(r, c, it)

            btn = QPushButton("✕")
            btn.setFixedSize(26, 22)
            btn.setToolTip("Удалить корректировку")
            btn.setStyleSheet(
                "QPushButton{background:#2a1318;color:#DC2626;border:1px solid #6e2a30;"
                "border-radius:4px;font-size:12px;font-weight:bold;}"
                "QPushButton:hover{background:#4a1a22;color:#ffcccc;}"
            )
            btn.clicked.connect(lambda _, i=orig_idx: self._delete_adjustment(i))
            self.adj_table.setCellWidget(r, 5, btn)
            self.adj_table.setRowHeight(r, 28)

        # Сводка
        self.summary_lbl.setText(
            f"Начислено: {fmt_money(bal.charged)}  ·  "
            f"оплачено: {fmt_money(bal.paid)}  ·  "
            f"баланс: {fmt_money(bal.debt)}  ·  "
            f"на {self._as_of.strftime('%d.%m.%Y')}"
        )

    def _set_year_row(self, r: int, cells: list, *, ignored: bool = False):
        dim_color = QColor("#3a4a56")   # приглушённый цвет для игнорируемых строк
        for c, (text, color) in enumerate(cells):
            it = QTableWidgetItem(text)
            it.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            if ignored:
                it.setForeground(dim_color)
            elif color:
                it.setForeground(QColor(color))
            it.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            self.table.setItem(r, c, it)
        self.table.setRowHeight(r, 28)

    def _add_adjustment(self, kind: str):
        dlg = VznosyAdjustmentDialog(self._plot, self, default_kind=kind)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        result = dlg.get_result()
        if not result:
            return
        from core import vznosy
        adj = vznosy.load_adjustments()
        adj.setdefault(self._plot, []).append(result)
        vznosy.save_adjustments(adj)
        self._rebuild()

    def _delete_adjustment(self, idx: int):
        from core import vznosy
        adj = vznosy.load_adjustments()
        items = adj.get(self._plot, [])
        if 0 <= idx < len(items):
            reply = QMessageBox.question(
                self, "Удалить корректировку",
                "Удалить эту корректировку?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
            items.pop(idx)
            if items:
                adj[self._plot] = items
            else:
                adj.pop(self._plot, None)
            vznosy.save_adjustments(adj)
            self._rebuild()

    def _toggle_ignore_period(self, period_from: str, ignore: bool):
        """Устанавливает или снимает флаг «Не учитывать» для периода.

        Удаляет как новые (ignore_period + period_from) так и старые
        (ignore_year + year) записи, чтобы снятие галки работало
        при загрузке проекта, сохранённого в старом формате.
        """
        if self._rebuilding:
            return
        from core import vznosy
        try:
            from datetime import datetime as _dt
            period_year = _dt.strptime(period_from, "%Y-%m-%d").year
        except Exception:
            period_year = None

        adj = vznosy.load_adjustments()
        items = adj.setdefault(self._plot, [])

        def _is_ignore_for_this_period(a: dict) -> bool:
            kind = a.get("kind")
            if kind == vznosy.KIND_IGNORE_PERIOD:
                return a.get("period_from") == period_from
            if kind == vznosy.KIND_IGNORE_YEAR and period_year is not None:
                try:
                    return int(a.get("year", 0)) == period_year
                except (TypeError, ValueError):
                    pass
            return False

        adj[self._plot] = [a for a in items if not _is_ignore_for_this_period(a)]
        if ignore:
            adj[self._plot].append({
                "kind": vznosy.KIND_IGNORE_PERIOD,
                "period_from": period_from,
                "date": date.today().isoformat(),
                "amount": None,
                "note": "Не учитывать",
            })
        if not adj[self._plot]:
            adj.pop(self._plot, None)
        vznosy.save_adjustments(adj)
        self._rebuild()

    def _on_pdf(self):
        try:
            from core import receipt
        except ImportError:
            QMessageBox.information(self, "Квитанции",
                                    "Модуль квитанций ещё не подключён.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить квитанцию",
            f"Уч_{self._plot}_ЧВ_{self._as_of.isoformat()}.pdf",
            "PDF (*.pdf)"
        )
        if not path:
            return
        try:
            receipt.save_vznosy_receipt_pdf(self._plot, self._df, path,
                                            as_of=self._as_of)
            QMessageBox.information(self, "Квитанция", f"Сохранено:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить:\n{e}")


class EnergyDebtWidget(QWidget):
    """Вкладка контроля долгов по электроэнергии."""

    def __init__(self):
        super().__init__()
        self.setAutoFillBackground(True)
        self._df = None
        self._setup_ui()

    def _setup_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(12)

        top = QHBoxLayout()
        title = QLabel("Долги по электроэнергии", objectName="pageTitle")
        top.addWidget(title)
        top.addStretch()

        top.addWidget(QLabel("на дату:", objectName="filterLabel"))
        self.date_as_of = QDateEdit(calendarPopup=True, objectName="datePicker",
                                    displayFormat="dd.MM.yyyy")
        self.date_as_of.setDate(QDate.currentDate())
        self.date_as_of.setMaximumDate(QDate.currentDate())
        self.date_as_of.dateChanged.connect(self._rebuild)
        top.addWidget(self.date_as_of)

        self.search = QLineEdit(objectName="searchInput")
        self.search.setPlaceholderText("Поиск по № участка или ФИО")
        self.search.setFixedWidth(280)
        self.search.textChanged.connect(self._apply_filter)
        top.addWidget(self.search)

        self.cb_only_debt = QComboBox(objectName="filterCombo")
        self.cb_only_debt.addItems(["Все участки", "Только должники", "Только аванс/0"])
        self.cb_only_debt.currentIndexChanged.connect(self._apply_filter)
        top.addWidget(self.cb_only_debt)

        btn = QPushButton("🔄  Пересчитать", objectName="btnPrimary")
        btn.clicked.connect(self._rebuild)
        top.addWidget(btn)

        self.btn_mass_pdf = QPushButton("📄  Квитанции должникам", objectName="btnSecondary")
        self.btn_mass_pdf.clicked.connect(self._export_debtor_receipts)
        top.addWidget(self.btn_mass_pdf)

        btn_rates = QPushButton("📐  Нормативы", objectName="btnSecondary")
        btn_rates.clicked.connect(self._open_rates_dialog)
        top.addWidget(btn_rates)

        btn_excel = QPushButton("Экспорт в Excel", objectName="btnSecondary")
        btn_excel.clicked.connect(self._export_excel)
        top.addWidget(btn_excel)

        lay.addLayout(top)

        # Легенда
        legend = QHBoxLayout()
        legend.setSpacing(20)
        for color, text in [
            ("#059669", "■  без долга / аванс"),
            ("#f9a825", "■  небольшой долг"),
            ("#ef6c00", "■  средний"),
            ("#DC2626", "■  крупный"),
        ]:
            lb = QLabel(text)
            lb.setStyleSheet(
                f"color:{color};background:transparent;font-size:11px;"
            )
            legend.addWidget(lb)
        legend.addStretch()
        lay.addLayout(legend)

        self.table = QTableWidget(objectName="summaryTable")
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(True)
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels([
            "Участок", "Владелец", "Последнее показание", "Дата показ.",
            "Начислено", "Оплачено", "Стартовое", "Долг / Аванс", "Без оплаты, мес.",
        ])
        hdr = self.table.horizontalHeader()
        for c, w in enumerate([85, 240, 140, 105, 120, 120, 100, 130, 110]):
            hdr.setSectionResizeMode(c, QHeaderView.ResizeMode.Interactive)
            self.table.setColumnWidth(c, w)
        hdr.setStretchLastSection(True)
        self.table.cellDoubleClicked.connect(self._open_card)
        lay.addWidget(self.table, 1)

        # Сверка с поставщиком
        self.recon_lbl = QLabel("", objectName="statusLabel")
        self.recon_lbl.setWordWrap(True)
        self.recon_lbl.setStyleSheet(
            "background:#F9FAFB;border:1px solid #E5E7EB;border-radius:6px;"
            "padding:10px 14px;color:#374151;font-size:12px;"
        )
        lay.addWidget(self.recon_lbl)

        self.status_lbl = QLabel("Загрузите выписку на вкладке «Детализация»",
                                  objectName="statusLabel")
        lay.addWidget(self.status_lbl)

        self.rates = RatesWidget()

    def _open_rates_dialog(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Нормативы — тарифы на электроэнергию")
        dlg.resize(700, 500)
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.addWidget(self.rates)
        dlg.exec()
        lay.removeWidget(self.rates)
        self.rates.setParent(self)  # type: ignore[arg-type]

    def refresh(self, df):
        self._df = df
        self._rebuild()

    def get_debts(self) -> dict:
        """Для карты: {plot: {debt, color}}."""
        if not hasattr(self, "_last_debts"):
            return {}
        return self._last_debts

    def _plot_list(self) -> list[str]:
        plots = energy.load_plots()
        nums = [str(p.get("num", "")) for p in plots if p.get("num")]
        # Plus all plots with readings even if not in registry
        meters = energy.load_meters()
        for key in meters:
            parts = key.split(":")
            if parts and parts[0] not in nums:
                nums.append(parts[0])
        # сортировка: сначала числовые по возрастанию, потом сложные
        def _key(s):
            try:
                return (0, int(s))
            except ValueError:
                return (1, s)
        return sorted(set(nums), key=_key)

    def _rebuild(self):
        as_of = self.date_as_of.date().toPyDate()
        meters = energy.load_meters()
        rates = energy.load_rates()
        repls = energy.load_replacements()
        baseline = energy.load_baseline()
        if self._df is not None and not self._df.empty:
            baseline["start_date"] = self._df["Дата"].min().date().isoformat()
        owners = energy.owners_map()
        plots = self._plot_list()

        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(plots))

        total_debt = 0.0
        total_charged = 0.0
        total_paid = 0.0
        debt_count = 0
        avg_monthly: list[float] = []
        debts_map: dict[str, dict] = {}

        for r, plot in enumerate(plots):
            bal = energy.balance(plot, as_of, meters, rates, repls, baseline, self._df)
            owner = ", ".join(owners.get(plot, [])) or "—"
            last_reading_text = "—"
            last_date_text = "—"
            if bal.last_reading:
                ly, lm, lv = bal.last_reading
                last_reading_text = f"{lv:g}"
                last_date_text = f"{lm:02d}.{ly}"

            # Средний месячный платёж для оценки уровня долга (для цвета на карте)
            if bal.charged > 0:
                count_charged_months = sum(
                    1 for c in energy.all_charges(plot, meters, rates, repls, as_of)
                    if c["amount"] is not None
                )
                if count_charged_months:
                    avg_monthly.append(bal.charged / count_charged_months)

            color = energy.debt_color(bal.debt, monthly_avg=300.0)
            debts_map[plot] = {"debt": bal.debt, "color": color,
                                "charged": bal.charged, "paid": bal.paid}

            total_debt += bal.debt
            total_charged += bal.charged
            total_paid += bal.paid
            if bal.debt > 0.5:
                debt_count += 1

            try:
                plot_sort = float(str(plot).split(",")[0])
            except ValueError:
                plot_sort = 0.0
            try:
                reading_sort = float(last_reading_text) if last_reading_text != "—" else -1.0
            except ValueError:
                reading_sort = -1.0
            date_sort = float(bal.last_reading[0] * 100 + bal.last_reading[1]) if bal.last_reading else -1.0

            cells = [
                (f"уч. {plot}", plot_sort, None, "#6366F1", True),
                (owner, owner, None, "#374151", False),
                (last_reading_text, reading_sort, None, "#374151", False),
                (last_date_text, date_sort, None, "#9CA3AF", False),
                (fmt_money(bal.charged), bal.charged, None, "#f9a825" if bal.charged else "#3a5a7a", False),
                (fmt_money(bal.paid), bal.paid, None, "#059669" if bal.paid else "#3a5a7a", False),
                (fmt_money(bal.baseline) if bal.baseline else "—", bal.baseline, None,
                 "#c97c7c" if bal.baseline else "#3a5a7a", False),
                (fmt_money(bal.debt), bal.debt, color, "#ffffff", True),
                ("—" if bal.months_without_payment is None else str(bal.months_without_payment),
                 bal.months_without_payment or 0, None,
                 "#DC2626" if (bal.months_without_payment or 0) > 3 else "#374151", False),
            ]
            for c, (text, value, bg, fg, bold) in enumerate(cells):
                if isinstance(value, (int, float)):
                    it = _NumItem(text, float(value))
                else:
                    it = QTableWidgetItem(text)
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                if bg:
                    it.setBackground(QColor(bg))
                if fg:
                    it.setForeground(QColor(fg))
                if bold:
                    f = it.font(); f.setBold(True); it.setFont(f)
                it.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                self.table.setItem(r, c, it)
            self.table.setRowHeight(r, 28)

        self.table.setSortingEnabled(True)
        self._last_debts = debts_map

        self.status_lbl.setText(
            f"Участков: {len(plots)}  ·  должников: {debt_count}  ·  "
            f"общий долг: {fmt_money(total_debt)}"
        )

        # Сверка
        try:
            df = self._df
            if df is None or df.empty:
                date_from = baseline_start = energy._parse_iso(baseline.get("start_date", "")) or date(as_of.year, 1, 1)
                date_to = as_of
            else:
                date_from = max(
                    df["Дата"].min().date(),
                    energy._parse_iso(baseline.get("start_date", "")) or df["Дата"].min().date(),
                )
                date_to = as_of
            common = energy.load_common_meter()
            rec = energy.reconcile(date_from, date_to, plots,
                                    meters, rates, repls, common, df)
            extras = ""
            if rec.common_kwh is not None:
                extras = (f"  ·  общий счётчик: {rec.common_kwh:.0f} кВт·ч"
                          f"  ·  частные: {rec.private_kwh:.0f} кВт·ч"
                          f"  ·  потери: {rec.loss_kwh:.0f} кВт·ч"
                          + (f" ({fmt_money(rec.loss_rub)})" if rec.loss_rub else ""))
            self.recon_lbl.setText(
                f"<b>Сверка с поставщиком</b> ({rec.period_from} — {rec.period_to}):  "
                f"начислено садоводам {fmt_money(rec.charged_total)}  ·  "
                f"собрано {fmt_money(rec.collected_total)}  ·  "
                f"уплачено в Пермэнергосбыт {fmt_money(rec.paid_to_supplier)}  ·  "
                f"расхождение {fmt_money(rec.collected_total - rec.paid_to_supplier)}"
                + extras
            )
        except Exception as e:
            self.recon_lbl.setText(f"Сверка недоступна: {e}")

        self._apply_filter()

    def _apply_filter(self):
        text = self.search.text().strip().lower()
        mode = self.cb_only_debt.currentText()
        for r in range(self.table.rowCount()):
            plot_item = self.table.item(r, 0)
            owner_item = self.table.item(r, 1)
            debt_item = self.table.item(r, 7)
            if not plot_item or not debt_item:
                continue
            visible = True
            if text:
                hay = (plot_item.text() + " " + (owner_item.text() if owner_item else "")).lower()
                visible = text in hay
            if visible and mode == "Только должники":
                visible = isinstance(debt_item, _NumItem) and debt_item._value > 0.5
            elif visible and mode == "Только аванс/0":
                visible = isinstance(debt_item, _NumItem) and debt_item._value <= 0.5
            self.table.setRowHidden(r, not visible)

    def _open_card(self, row: int, _col: int):
        plot_item = self.table.item(row, 0)
        if not plot_item:
            return
        plot = plot_item.text().replace("уч. ", "").strip()
        as_of = self.date_as_of.date().toPyDate()
        dlg = PlotCardDialog(plot, self._df, self, as_of=as_of)
        dlg.exec()
        # после возможной правки показаний / замены счётчика — пересчитать
        self._rebuild()

    def _export_debtor_receipts(self):
        if not getattr(self, "_last_debts", None):
            QMessageBox.information(self, "Квитанции", "Сначала загрузите выписку.")
            return
        debtors = [(p, info) for p, info in self._last_debts.items()
                   if info["debt"] > 0.5]
        if not debtors:
            QMessageBox.information(self, "Квитанции", "Должников нет — квитанции не нужны.")
            return
        folder = QFileDialog.getExistingDirectory(self, "Папка для квитанций")
        if not folder:
            return
        try:
            from core import receipt
        except ImportError as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось импортировать модуль квитанций:\n{e}")
            return

        owners = energy.owners_map()
        as_of = self.date_as_of.date().toPyDate()
        ok = 0
        errors = []
        for plot, info in debtors:
            owner = (owners.get(plot, [""])[0] or "").split()
            surname = owner[0] if owner else ""
            fname = f"Уч_{plot}"
            if surname:
                safe_surname = re.sub(r"[^\w\-]", "_", surname)
                fname += f"_{safe_surname}"
            fname += f"_{as_of.isoformat()}.pdf"
            fname = fname.replace("/", "-")
            path = os.path.join(folder, fname)
            try:
                receipt.save_plot_receipt_pdf(plot, self._df, path, as_of=as_of)
                ok += 1
            except Exception as e:
                errors.append(f"уч. {plot}: {e}")
        if errors:
            QMessageBox.warning(
                self, "Квитанции",
                f"Создано: {ok}\nОшибки ({len(errors)}):\n" + "\n".join(errors[:10])
            )
        else:
            QMessageBox.information(
                self, "Квитанции",
                f"✅  Сформировано {ok} квитанций в:\n{folder}"
            )

    def _export_excel(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Экспорт таблицы", "электроэнергия_долги.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        if not path.endswith(".xlsx"):
            path += ".xlsx"

        headers = [
            self.table.horizontalHeaderItem(c).text()
            for c in range(self.table.columnCount())
        ]
        rows = []
        for r in range(self.table.rowCount()):
            if self.table.isRowHidden(r):
                continue
            rows.append([
                (self.table.item(r, c).text() if self.table.item(r, c) else "")
                for c in range(self.table.columnCount())
            ])

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Долги по электроэнергии"
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            for row in rows:
                ws.append(row)
            wb.save(path)
            QMessageBox.information(self, "Экспорт завершён", f"Файл сохранён:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка экспорта", str(e))


class VznosyDebtWidget(QWidget):
    """Вкладка контроля долгов по членским взносам."""

    def __init__(self):
        super().__init__()
        self.setAutoFillBackground(True)
        self._df = None
        self._last_debts: dict = {}
        self._setup_ui()
        self._rebuild()

    def _setup_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(12)

        top = QHBoxLayout()
        title = QLabel("Долги по членским взносам", objectName="pageTitle")
        top.addWidget(title)
        top.addStretch()

        top.addWidget(QLabel("на дату:", objectName="filterLabel"))
        self.date_as_of = QDateEdit(calendarPopup=True, objectName="datePicker",
                                    displayFormat="dd.MM.yyyy")
        self.date_as_of.setDate(QDate.currentDate())
        self.date_as_of.setMaximumDate(QDate.currentDate())
        self.date_as_of.dateChanged.connect(self._rebuild)
        top.addWidget(self.date_as_of)

        self.search = QLineEdit(objectName="searchInput")
        self.search.setPlaceholderText("Поиск по № участка или ФИО")
        self.search.setFixedWidth(280)
        self.search.textChanged.connect(self._apply_filter)
        top.addWidget(self.search)

        self.cb_only_debt = QComboBox(objectName="filterCombo")
        self.cb_only_debt.addItems(["Все участки", "Только должники", "Только оплачено/аванс"])
        self.cb_only_debt.currentIndexChanged.connect(self._apply_filter)
        top.addWidget(self.cb_only_debt)

        btn = QPushButton("🔄  Пересчитать", objectName="btnPrimary")
        btn.clicked.connect(self._rebuild)
        top.addWidget(btn)

        self.btn_mass_pdf = QPushButton("📄  Квитанции должникам", objectName="btnSecondary")
        self.btn_mass_pdf.clicked.connect(self._export_debtor_receipts)
        top.addWidget(self.btn_mass_pdf)

        btn_rates = QPushButton("📅  Периоды", objectName="btnSecondary")
        btn_rates.clicked.connect(self._open_rates_dialog)
        top.addWidget(btn_rates)

        btn_excel = QPushButton("Экспорт в Excel", objectName="btnSecondary")
        btn_excel.clicked.connect(self._export_excel)
        top.addWidget(btn_excel)

        lay.addLayout(top)

        # Легенда
        legend = QHBoxLayout()
        legend.setSpacing(20)
        for color, text in [
            ("#059669", "■  без долга / аванс"),
            ("#f9a825", "■  небольшой долг"),
            ("#ef6c00", "■  средний"),
            ("#DC2626", "■  крупный"),
        ]:
            lb = QLabel(text)
            lb.setStyleSheet(
                f"color:{color};background:transparent;font-size:11px;"
            )
            legend.addWidget(lb)
        legend.addStretch()
        lay.addLayout(legend)

        self.table = QTableWidget(objectName="summaryTable")
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(True)
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            "Участок", "Владелец", "Площадь, м²",
            "Начислено", "Оплачено", "Долг / Аванс", "Лет без оплаты",
        ])
        hdr = self.table.horizontalHeader()
        for c, w in enumerate([85, 280, 100, 130, 130, 140, 110]):
            hdr.setSectionResizeMode(c, QHeaderView.ResizeMode.Interactive)
            self.table.setColumnWidth(c, w)
        hdr.setStretchLastSection(True)
        self.table.cellDoubleClicked.connect(self._open_card)
        lay.addWidget(self.table, 1)

        self.status_lbl = QLabel("Загрузите выписку на вкладке «Детализация»",
                                  objectName="statusLabel")
        lay.addWidget(self.status_lbl)

        self.rates = VznosyRatesWidget()

    def _open_rates_dialog(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Периоды членских взносов")
        dlg.resize(720, 500)
        dlg_lay = QVBoxLayout(dlg)
        dlg_lay.setContentsMargins(0, 0, 0, 0)
        dlg_lay.addWidget(self.rates)
        dlg.exec()
        dlg_lay.removeWidget(self.rates)
        self.rates.setParent(self)  # type: ignore[arg-type]
        # Тарифы могли поменяться — пересчитаем
        self._rebuild()

    def refresh(self, df):
        self._df = df
        self._rebuild()

    def _plot_list(self) -> list[str]:
        plots = energy.load_plots()
        nums = [str(p.get("num", "")) for p in plots if p.get("num")]
        def _key(s):
            try:
                return (0, int(s))
            except ValueError:
                return (1, s)
        return sorted(set(nums), key=_key)

    def _rebuild(self):
        from core import vznosy
        as_of = self.date_as_of.date().toPyDate()
        rates = vznosy.load_rates()
        adj = vznosy.load_adjustments()
        areas = vznosy.plot_area_map()
        owners = energy.owners_map()
        plots = self._plot_list()

        # средняя годовая сумма по тарифу — для подсветки уровня долга
        annual_avg = 0.0
        if rates:
            sums = []
            for r in rates:
                v = energy._to_float(r.get("amount"))
                if v is not None:
                    sums.append(v)
            if sums:
                annual_avg = sum(sums) / len(sums)

        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(plots))

        total_debt = 0.0
        total_charged = 0.0
        total_paid = 0.0
        debt_count = 0
        debts_map: dict[str, dict] = {}

        for r, plot in enumerate(plots):
            area = areas.get(plot)
            bal = vznosy.balance_for_plot(plot, area, as_of, rates, adj, self._df)
            owner = ", ".join(owners.get(plot, [])) or "—"

            area_text = f"{area:g}" if area is not None else "—"
            area_color = "#DC2626" if bal.area_missing_warning else (
                "#374151" if area is not None else "#9CA3AF"
            )
            area_tip = ("Не указана площадь — начисление по тарифу за м² невозможно"
                        if bal.area_missing_warning else "")

            color = vznosy.debt_color(bal.debt, annual_avg=annual_avg)
            debts_map[plot] = {"debt": bal.debt, "color": color,
                                "charged": bal.charged, "paid": bal.paid}

            total_debt += bal.debt
            total_charged += bal.charged
            total_paid += bal.paid
            if bal.debt > 0.5:
                debt_count += 1

            years_unpaid_text = "—" if bal.years_unpaid is None else str(bal.years_unpaid)
            years_unpaid_color = "#374151"
            if bal.years_unpaid and bal.years_unpaid >= 2:
                years_unpaid_color = "#DC2626" if bal.years_unpaid >= 3 else "#ffd54f"

            try:
                plot_sort = float(str(plot).split(",")[0])
            except ValueError:
                plot_sort = 0.0

            cells = [
                (f"уч. {plot}", plot_sort, None, "#6366F1", True, ""),
                (owner, owner, None, "#374151", False, ""),
                (area_text, area if area is not None else 0.0, None, area_color, False, area_tip),
                (fmt_money(bal.charged), bal.charged, None,
                 "#f9a825" if bal.charged else "#3a5a7a", False, ""),
                (fmt_money(bal.paid), bal.paid, None,
                 "#059669" if bal.paid else "#3a5a7a", False, ""),
                (fmt_money(bal.debt), bal.debt, color, "#ffffff", True, ""),
                (years_unpaid_text, float(bal.years_unpaid or 0), None,
                 years_unpaid_color, False, ""),
            ]
            for c, (text, value, bg, fg, bold, tip) in enumerate(cells):
                if isinstance(value, (int, float)):
                    it = _NumItem(text, float(value))
                else:
                    it = QTableWidgetItem(text)
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                if bg:
                    it.setBackground(QColor(bg))
                if fg:
                    it.setForeground(QColor(fg))
                if bold:
                    f = it.font(); f.setBold(True); it.setFont(f)
                if tip:
                    it.setToolTip(tip)
                it.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                self.table.setItem(r, c, it)
            self.table.setRowHeight(r, 28)

        self.table.setSortingEnabled(True)
        self._last_debts = debts_map

        self.status_lbl.setText(
            f"Участков: {len(plots)}  ·  должников: {debt_count}  ·  "
            f"начислено всего: {fmt_money(total_charged)}  ·  "
            f"оплачено всего: {fmt_money(total_paid)}  ·  "
            f"общий долг: {fmt_money(total_debt)}"
        )

        self._apply_filter()

    def _apply_filter(self):
        text = self.search.text().strip().lower()
        mode = self.cb_only_debt.currentText()
        for r in range(self.table.rowCount()):
            plot_item = self.table.item(r, 0)
            owner_item = self.table.item(r, 1)
            debt_item = self.table.item(r, 5)
            if not plot_item or not debt_item:
                continue
            visible = True
            if text:
                hay = (plot_item.text() + " " + (owner_item.text() if owner_item else "")).lower()
                visible = text in hay
            if visible and mode == "Только должники":
                visible = isinstance(debt_item, _NumItem) and debt_item._value > 0.5
            elif visible and mode == "Только оплачено/аванс":
                visible = isinstance(debt_item, _NumItem) and debt_item._value <= 0.5
            self.table.setRowHidden(r, not visible)

    def _open_card(self, row: int, _col: int):
        plot_item = self.table.item(row, 0)
        if not plot_item:
            return
        plot = plot_item.text().replace("уч. ", "").strip()
        as_of = self.date_as_of.date().toPyDate()
        dlg = VznosyCardDialog(plot, self._df, self, as_of=as_of)
        dlg.exec()
        self._rebuild()

    def _export_debtor_receipts(self):
        if not self._last_debts:
            QMessageBox.information(self, "Квитанции", "Сначала загрузите выписку.")
            return
        debtors = [(p, info) for p, info in self._last_debts.items()
                   if info["debt"] > 0.5]
        if not debtors:
            QMessageBox.information(self, "Квитанции", "Должников нет — квитанции не нужны.")
            return
        folder = QFileDialog.getExistingDirectory(self, "Папка для квитанций")
        if not folder:
            return
        try:
            from core import receipt
        except ImportError as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось импортировать модуль квитанций:\n{e}")
            return

        owners = energy.owners_map()
        as_of = self.date_as_of.date().toPyDate()
        ok = 0
        errors = []
        for plot, _info in debtors:
            owner = (owners.get(plot, [""])[0] or "").split()
            surname = owner[0] if owner else ""
            fname = f"Уч_{plot}_ЧВ"
            if surname:
                safe_surname = re.sub(r"[^\w\-]", "_", surname)
                fname += f"_{safe_surname}"
            fname += f"_{as_of.isoformat()}.pdf"
            fname = fname.replace("/", "-")
            path = os.path.join(folder, fname)
            try:
                receipt.save_vznosy_receipt_pdf(plot, self._df, path, as_of=as_of)
                ok += 1
            except Exception as e:
                errors.append(f"уч. {plot}: {e}")
        if errors:
            QMessageBox.warning(
                self, "Квитанции",
                f"Создано: {ok}\nОшибки ({len(errors)}):\n" + "\n".join(errors[:10])
            )
        else:
            QMessageBox.information(
                self, "Квитанции",
                f"✅  Сформировано {ok} квитанций в:\n{folder}"
            )

    def _export_excel(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Экспорт таблицы", "членские_взносы_долги.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        if not path.endswith(".xlsx"):
            path += ".xlsx"

        headers = [
            self.table.horizontalHeaderItem(c).text()
            for c in range(self.table.columnCount())
        ]
        rows = []
        for r in range(self.table.rowCount()):
            if self.table.isRowHidden(r):
                continue
            rows.append([
                (self.table.item(r, c).text() if self.table.item(r, c) else "")
                for c in range(self.table.columnCount())
            ])

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Долги по членским взносам"
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            for row in rows:
                ws.append(row)
            wb.save(path)
            QMessageBox.information(self, "Экспорт завершён", f"Файл сохранён:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка экспорта", str(e))

class _TitleBar(QWidget):
    """Custom frameless-window title bar: title text + min/max/close buttons."""

    def __init__(self, window: "MainWindow"):
        super().__init__(window)
        self._window = window
        self._drag_pos = None
        self.setObjectName("titleBar")
        self.setFixedHeight(32)

        lyt = QHBoxLayout(self)
        lyt.setContentsMargins(16, 0, 0, 0)
        lyt.setSpacing(0)

        lyt.addStretch()

        icon_font = QFont("Segoe MDL2 Assets")
        icon_font.setPixelSize(10)

        for obj_name, char, slot in [
            ("btnWinMin",   "", self._minimize),
            ("btnWinMax",   "", self._toggle_max),
            ("btnWinClose", "", window.close),
        ]:
            btn = QPushButton(char, objectName=obj_name)
            btn.setFixedSize(46, 32)
            btn.setFont(icon_font)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.clicked.connect(slot)
            lyt.addWidget(btn)

        self._btn_max = self.findChild(QPushButton, "btnWinMax")

    def paintEvent(self, event):
        opt = QStyleOption()
        opt.initFrom(self)
        p = QPainter(self)
        self.style().drawPrimitive(QStyle.PrimitiveElement.PE_Widget, opt, p, self)

    def _minimize(self):
        self._window.showMinimized()

    def _toggle_max(self):
        if self._window.isMaximized():
            self._window.showNormal()
            if self._btn_max:
                self._btn_max.setText("")
        else:
            self._window.showMaximized()
            if self._btn_max:
                self._btn_max.setText("")

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self._drag_pos = event.globalPosition().toPoint() - self._window.pos()
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._drag_pos is not None and event.buttons() == Qt.MouseButton.LeftButton:
            if self._window.isMaximized():
                self._window.showNormal()
                if self._btn_max:
                    self._btn_max.setText("")
                self._drag_pos = QPoint(self._window.width() // 3, self.height() // 2)
            self._window.move(event.globalPosition().toPoint() - self._drag_pos)
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        self._drag_pos = None
        super().mouseReleaseEvent(event)

    def mouseDoubleClickEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self._toggle_max()
        super().mouseDoubleClickEvent(event)


class _NavButton(QWidget):
    """Пункт левого сайдбара: глиф Material Icons + подпись."""
    nav_clicked = pyqtSignal(int)

    def __init__(self, icon_char: str, label: str, page_idx: int, parent=None):
        super().__init__(parent)
        self._page_idx = page_idx
        self.setObjectName("navBtn")
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setAttribute(Qt.WidgetAttribute.WA_Hover)
        self.setFixedHeight(38)

        lyt = QHBoxLayout(self)
        lyt.setContentsMargins(6, 0, 15, 0)
        lyt.setSpacing(7)

        self._icon = QLabel(icon_char, objectName="navIcon")
        icon_font = QFont("Material Symbols Rounded")
        icon_font.setPixelSize(25)
        self._icon.setFont(icon_font)
        self._icon.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)

        self._lbl = QLabel(label, objectName="navLabel")
        self._lbl.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)

        lyt.addWidget(self._icon)
        lyt.addWidget(self._lbl)
        lyt.addStretch()

    def paintEvent(self, event):
        opt = QStyleOption()
        opt.initFrom(self)
        p = QPainter(self)
        self.style().drawPrimitive(QStyle.PrimitiveElement.PE_Widget, opt, p, self)

    def set_active(self, active: bool):
        prop = "true" if active else "false"
        for w in (self, self._icon, self._lbl):
            w.setProperty("active", prop)
            w.style().unpolish(w)
            w.style().polish(w)

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.nav_clicked.emit(self._page_idx)
        super().mousePressEvent(event)


class _ActionButton(QWidget):
    """Кнопка действия в нижней части сайдбара: значок + подпись."""
    clicked = pyqtSignal()

    def __init__(self, icon_char: str, label: str, parent=None):
        super().__init__(parent)
        self.setObjectName("btnNavAction")
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setAttribute(Qt.WidgetAttribute.WA_Hover)
        self.setFixedHeight(38)

        lyt = QHBoxLayout(self)
        lyt.setContentsMargins(10, 0, 15, 0)
        lyt.setSpacing(7)

        self._icon = QLabel(icon_char, objectName="navIcon")
        icon_font = QFont("Material Symbols Rounded")
        icon_font.setPixelSize(17)
        self._icon.setFont(icon_font)
        self._icon.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)

        self._lbl = QLabel(label, objectName="actionLabel")
        self._lbl.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)

        lyt.addWidget(self._icon)
        lyt.addWidget(self._lbl)
        lyt.addStretch()

    def paintEvent(self, event):
        opt = QStyleOption()
        opt.initFrom(self)
        p = QPainter(self)
        self.style().drawPrimitive(QStyle.PrimitiveElement.PE_Widget, opt, p, self)

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.clicked.emit()
        super().mousePressEvent(event)


class _BrandText(QWidget):
    """Надпись «МОЙ / САДОВОД / Бухгалтерский учет для СНТ».

    Отрисовка вручную через QPainter: положение каждой строки считается
    по tightBoundingRect (реальные пиксели глифов), поэтому межстрочные
    зазоры точны и не зависят от капризов метрик конкретного шрифта.
    Зазоры _GAP_* — явные константы, их легко подправить.
    """

    _COLOR_TITLE = QColor("#07414F")
    _COLOR_SUB   = QColor("#7A8A95")
    _GAP_TITLE   = 2     # зазор между МОЙ и САДОВОД, px
    _GAP_SUB     = 0     # зазор между САДОВОД и подписью, px
    _PAD_X       = 0     # горизонтальный отступ-страховка (запас под овершут)
    _PAD_TOP     = 10    # верхний отступ — сдвигает текст вниз относительно логотипа
    _PAD_Y       = 0     # нижний отступ-страховка

    def __init__(self, family: str, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)

        def _font(px: int, spacing: float, weight: int = 400) -> QFont:
            f = QFont(family)
            f.setPixelSize(px)
            f.setWeight(QFont.Weight(weight))
            f.setLetterSpacing(QFont.SpacingType.AbsoluteSpacing, spacing)
            return f

        # (текст, шрифт, цвет, оптическая коррекция X)
        # nudge<0 сдвигает строку левее: круглая «С» зрительно кажется
        # правее плоской «М», поэтому «САДОВОД» слегка выносим влево.
        self._lines = [
            ("МОЙ",                        _font(14, 0.0, 600), self._COLOR_TITLE,  0),
            ("САДОВОД",                    _font(20, 0.0, 600), self._COLOR_TITLE,  0),
            ("Бухгалтерский учет для СНТ", _font(10, 0.0),      self._COLOR_SUB,    0),
        ]
        self._gaps = [self._GAP_TITLE, self._GAP_SUB]
        self._layout_lines()

    def _layout_lines(self):
        """Считает draw_x / baseline каждой строки и итоговый размер."""
        placed = []
        y = self._PAD_TOP
        right_edge = 0
        last_bottom = y
        for i, (text, font, color, nudge) in enumerate(self._lines):
            fm = QFontMetrics(font)
            tbr = fm.tightBoundingRect(text)
            # tbr.top() отрицателен (над базовой линией) → baseline ниже
            baseline = y - tbr.top()
            # левый край реальных глифов в _PAD_X + оптическая коррекция
            draw_x = self._PAD_X - tbr.left() + nudge
            placed.append((text, font, color, draw_x, baseline))
            last_bottom = baseline + tbr.bottom() + 1
            right_edge = max(right_edge, draw_x + tbr.left() + tbr.width())
            if i < len(self._gaps):
                y = last_bottom + self._gaps[i]
        self._placed = placed
        self.setFixedSize(right_edge + self._PAD_X + 6,
                          last_bottom + self._PAD_Y)

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        p.setRenderHint(QPainter.RenderHint.TextAntialiasing)
        for text, font, color, draw_x, baseline in self._placed:
            p.setFont(font)
            p.setPen(color)
            p.drawText(int(draw_x), int(baseline), text)


class _RoundedFrame(QFrame):
    """QFrame, который обрезает все дочерние виджеты по скруглённому прямоугольнику.

    Использует QBitmap-маску (setMask), которая физически ограничивает рендеринг
    всего дерева виджетов внутри скруглённой области — в отличие от QSS border-radius,
    влияющего только на собственный фон виджета.
    """
    _RADIUS = 14

    def resizeEvent(self, event):
        super().resizeEvent(event)
        bm = QBitmap(self.size())
        bm.fill(Qt.GlobalColor.color0)          # всё «прозрачно»
        bp = QPainter(bm)
        bp.setRenderHint(QPainter.RenderHint.Antialiasing)
        path = QPainterPath()
        path.addRoundedRect(QRectF(self.rect()), self._RADIUS, self._RADIUS)
        bp.fillPath(path, Qt.GlobalColor.color1)  # скруглённая область «видна»
        bp.end()
        self.setMask(bm)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.WindowType.Window)
        self.setWindowTitle("Мой Садовод")
        self.setMinimumSize(1280, 720)
        self.resize(1500, 860)
        self._setup_ui()
        self._apply_styles()

    # ── Native resize support on Windows ─────────────────────────────────

    def showEvent(self, event):
        super().showEvent(event)
        if sys.platform == "win32":
            self._restore_win_resize()

    def _restore_win_resize(self):
        """Ensure full WS_OVERLAPPEDWINDOW style + DWM frame for animations."""
        try:
            import ctypes
            hwnd = int(self.winId())
            GWL_STYLE = -16
            style = ctypes.windll.user32.GetWindowLongW(hwnd, GWL_STYLE)
            style |= (0x00040000   # WS_THICKFRAME
                    | 0x00C00000   # WS_CAPTION
                    | 0x00080000   # WS_SYSMENU
                    | 0x00020000   # WS_MINIMIZEBOX
                    | 0x00010000)  # WS_MAXIMIZEBOX
            ctypes.windll.user32.SetWindowLongW(hwnd, GWL_STYLE, style)

            class _MARGINS(ctypes.Structure):
                _fields_ = [("l", ctypes.c_int), ("r", ctypes.c_int),
                             ("t", ctypes.c_int), ("b", ctypes.c_int)]
            ctypes.windll.dwmapi.DwmExtendFrameIntoClientArea(
                hwnd, ctypes.byref(_MARGINS(-1, -1, -1, -1))
            )

            SWP_FLAGS = 0x0020 | 0x0002 | 0x0001 | 0x0004
            ctypes.windll.user32.SetWindowPos(hwnd, 0, 0, 0, 0, 0, SWP_FLAGS)
        except Exception:
            pass

    def nativeEvent(self, event_type, message):
        """Handle WM_NCHITTEST / WM_NCCALCSIZE for frameless resize support."""
        if sys.platform == "win32" and event_type == b"windows_generic_MSG":
            import ctypes

            class _MSG(ctypes.Structure):
                _fields_ = [
                    ("hWnd",    ctypes.c_void_p),
                    ("message", ctypes.c_uint),
                    ("wParam",  ctypes.c_size_t),
                    ("lParam",  ctypes.c_ssize_t),
                    ("time",    ctypes.c_uint),
                    ("pt_x",    ctypes.c_int),
                    ("pt_y",    ctypes.c_int),
                ]

            try:
                msg = ctypes.cast(int(message), ctypes.POINTER(_MSG)).contents
            except Exception:
                return False, 0

            WM_NCCALCSIZE = 0x0083
            WM_NCHITTEST  = 0x0084

            if msg.message == WM_NCCALCSIZE and msg.wParam:
                # Collapse non-client area so native chrome is invisible
                return True, 0

            if msg.message == WM_NCHITTEST:
                HTCLIENT      = 1
                HTCAPTION     = 2
                HTLEFT        = 10; HTRIGHT      = 11
                HTTOP         = 12; HTTOPLEFT    = 13
                HTTOPRIGHT    = 14; HTBOTTOM     = 15
                HTBOTTOMLEFT  = 16; HTBOTTOMRIGHT = 17

                x = ctypes.c_int16(msg.lParam & 0xFFFF).value
                y = ctypes.c_int16((msg.lParam >> 16) & 0xFFFF).value
                pos = self.mapFromGlobal(QPoint(x, y))
                px, py = pos.x(), pos.y()
                w, h = self.width(), self.height()
                m = 6  # resize margin in pixels

                on_l = px < m;      on_r = px > w - m - 1
                on_t = py < m;      on_b = py > h - m - 1

                if on_t and on_l:  return True, HTTOPLEFT
                if on_t and on_r:  return True, HTTOPRIGHT
                if on_b and on_l:  return True, HTBOTTOMLEFT
                if on_b and on_r:  return True, HTBOTTOMRIGHT
                if on_l:           return True, HTLEFT
                if on_r:           return True, HTRIGHT
                if on_t:           return True, HTTOP
                if on_b:           return True, HTBOTTOM

                # Title bar: HTCAPTION for native drag/snap, HTCLIENT for buttons
                tb_h = self._title_bar.height()
                btn_w = 3 * 46  # three 46px window buttons on the right
                if py < tb_h:
                    if px >= w - btn_w:
                        return True, HTCLIENT
                    return True, HTCAPTION

        return False, 0

    def _setup_ui(self):
        central = QWidget()
        central.setAutoFillBackground(True)
        self.setCentralWidget(central)

        # Outer layout: title bar + body (sidebar | content)
        outer = QVBoxLayout(central)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        self._title_bar = _TitleBar(self)
        outer.addWidget(self._title_bar)

        # Body: левый сайдбар + область контента
        body = QWidget(objectName="bodyArea")
        body.setAutoFillBackground(True)
        body_lyt = QHBoxLayout(body)
        body_lyt.setContentsMargins(0, 0, 8, 8)
        body_lyt.setSpacing(0)

        # ── Левый сайдбар навигации ──────────────────────────────────────
        sidebar = QWidget(objectName="sideNav")
        sidebar.setAutoFillBackground(True)
        sidebar.setFixedWidth(250)
        side_lyt = QVBoxLayout(sidebar)
        side_lyt.setContentsMargins(16, 0, 16, 16)
        side_lyt.setSpacing(4)

        # Шапка сайдбара: логотип + текстовый блок «МОЙ / САДОВОД»
        header = QWidget()
        header_lyt = QHBoxLayout(header)
        header_lyt.setContentsMargins(0, 0, 0, 0)
        header_lyt.setSpacing(4)

        _logo_file = Path(__file__).parent / "resources" / "images" / "logo.png"
        if _logo_file.exists():
            _pix = QPixmap(str(_logo_file))
            if not _pix.isNull():
                _logo_pix = _pix.scaledToHeight(
                    52, Qt.TransformationMode.SmoothTransformation
                )
                _lbl_logo = QLabel(objectName="navLogo")
                _lbl_logo.setPixmap(_logo_pix)
                _lbl_logo.setFixedSize(_logo_pix.width(), _logo_pix.height())
                header_lyt.addWidget(_lbl_logo, alignment=Qt.AlignmentFlag.AlignVCenter)

        # Текстовый блок «МОЙ / САДОВОД / подпись» — собственная отрисовка
        _installed = set(QFontDatabase.families())
        _brand_family = next(
            (f for f in ("Geologica", "KOT-Eitai Gothic Bold", "KOT-Eitai Gothic",
                         "Montserrat", "Segoe UI") if f in _installed),
            "Segoe UI",
        )
        _brand = _BrandText(_brand_family)
        header_lyt.addWidget(_brand, alignment=Qt.AlignmentFlag.AlignVCenter)
        header_lyt.addStretch()
        side_lyt.addWidget(header)
        side_lyt.addSpacing(18)

        self._nav_buttons: list[_NavButton] = []
        for icon, label, idx in [
            (chr(0xe587), "Главная",            0),
            (chr(0xf191), "Детализация",        1),
            (chr(0xeaec), "Членские взносы",    2),
            (chr(0xec1c), "Электричество",      4),
            (chr(0xf8ee), "Список участков",    3),
        ]:
            btn = _NavButton(icon, label, idx)
            btn.nav_clicked.connect(self._nav_click)
            self._nav_buttons.append(btn)
            side_lyt.addWidget(btn)

        side_lyt.addStretch()

        btn_save_proj = _ActionButton(chr(0xf09b), "Сохранить базу СНТ")
        btn_save_proj.clicked.connect(self._save_project)
        side_lyt.addWidget(btn_save_proj)

        btn_load_proj = _ActionButton(chr(0xf090), "Загрузить базу СНТ")
        btn_load_proj.clicked.connect(self._load_project)
        side_lyt.addWidget(btn_load_proj)

        body_lyt.addWidget(sidebar)

        # Content stack
        self.stack = QStackedWidget(objectName="contentArea")
        self.home        = HomeWidget()
        self.detail      = DetailWidget()
        self.vznosy_debt = VznosyDebtWidget()
        self.plots       = PlotsWidget()
        self.energy_debt = EnergyDebtWidget()
        for tab in (self.home, self.detail, self.vznosy_debt, self.plots,
                    self.energy_debt):
            tab.setAutoFillBackground(True)
        self.stack.addWidget(self.home)         # 0
        self.stack.addWidget(self.detail)       # 1
        self.stack.addWidget(self.vznosy_debt)  # 2
        self.stack.addWidget(self.plots)        # 3
        self.stack.addWidget(self.energy_debt)  # 4
        content_frame = _RoundedFrame(objectName="contentFrame")
        content_frame.setAutoFillBackground(True)
        cf_lyt = QVBoxLayout(content_frame)
        cf_lyt.setContentsMargins(0, 0, 0, 0)
        cf_lyt.setSpacing(0)
        cf_lyt.addWidget(self.stack)
        body_lyt.addWidget(content_frame, stretch=1)

        outer.addWidget(body, stretch=1)

        # Подписки на загрузку выписки
        self.detail.dataLoaded.connect(self.vznosy_debt.refresh)
        self.detail.dataLoaded.connect(self.energy_debt.refresh)
        # При изменении списка участков (в т.ч. площади) пересчитать ЧВ
        self.plots.plotsUpdated.connect(
            lambda: self.vznosy_debt.refresh(self.detail.df_full)
        )

        # При изменении данных вкладки «Участки» — обновить все зависимые вкладки
        self.plots.plotsUpdated.connect(self.detail.refresh_plot_column)

        self._nav_click(0)  # initial page: Главная

    def _nav_click(self, page_idx: int):
        for btn in self._nav_buttons:
            btn.set_active(btn._page_idx == page_idx)
        self.stack.setCurrentIndex(page_idx)
        if page_idx == 2:
            self.vznosy_debt.refresh(self.detail.df_full)
        elif page_idx == 4:
            self.energy_debt.refresh(self.detail.df_full)

    # ── Сохранение / загрузка проекта ────────────────────────────────────

    _PROJECT_JSON_FILES = [
        "snt_plots.json", "snt_rates.json",
        "snt_vznosy_rates.json", "snt_vznosy_adjustments.json",
        "snt_map_plots.json", "snt_map_image.json",
        "snt_meters.json", "snt_meters_years.json",
        "snt_meter_replacements.json", "snt_energy_baseline.json",
        "snt_common_meter.json",
    ]

    def _save_project(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить проект СНТ", "", "Проект СНТ (*.snt)")
        if not path:
            return
        if not path.endswith(".snt"):
            path += ".snt"

        data_dir = Path(DATA_DIR)
        errors = []
        try:
            with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
                for fname in self._PROJECT_JSON_FILES:
                    src = data_dir / fname
                    if src.exists():
                        zf.write(src, f"data/{fname}")

                # сохраняем данные вкладки «Детализация»
                if self.detail.df_full is not None:
                    try:
                        json_str = self.detail.df_full.to_json(
                            orient="records", force_ascii=False)
                        zf.writestr("data/detail_transactions.json", json_str)
                    except Exception as e:
                        errors.append(f"Детализация: {e}")

                # включаем файл карты, если он локальный
                map_cfg = data_dir / "snt_map_image.json"
                if map_cfg.exists():
                    try:
                        with open(map_cfg, encoding="utf-8") as f:
                            img_path = json.load(f).get("path", "")
                        if img_path and Path(img_path).is_file():
                            ext = Path(img_path).suffix
                            zf.write(img_path, f"map_image{ext}")
                    except Exception as e:
                        errors.append(f"Изображение карты: {e}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить проект:\n{e}")
            return

        msg = f"Проект сохранён:\n{path}"
        if errors:
            msg += "\n\nПредупреждения:\n" + "\n".join(errors)
        QMessageBox.information(self, "Сохранено", msg)

    def _load_project(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Загрузить проект СНТ", "", "Проект СНТ (*.snt)")
        if not path:
            return

        reply = QMessageBox.question(
            self, "Загрузка проекта",
            "Текущие данные будут заменены данными из файла.\nПродолжить?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        data_dir = Path(DATA_DIR)
        data_dir.mkdir(exist_ok=True)
        try:
            with zipfile.ZipFile(path, "r") as zf:
                names = zf.namelist()

                # извлекаем JSON-файлы данных (кроме транзакций — они в памяти)
                for name in names:
                    if name.startswith("data/") and name.endswith(".json"):
                        fname = name[5:]
                        if fname and fname != "detail_transactions.json":
                            dest = data_dir / fname
                            dest.write_bytes(zf.read(name))

                # восстанавливаем данные вкладки «Детализация»
                detail_df = None
                if "data/detail_transactions.json" in names:
                    try:
                        from io import StringIO
                        json_str = zf.read("data/detail_transactions.json").decode("utf-8")
                        detail_df = pd.read_json(StringIO(json_str), orient="records")
                        detail_df["Дата"] = pd.to_datetime(
                            detail_df["Дата"], unit="ms", errors="coerce")
                    except Exception as e:
                        QMessageBox.warning(
                            self, "Предупреждение",
                            f"Не удалось загрузить данные Детализации:\n{e}")

                # извлекаем изображение карты
                map_name = next(
                    (n for n in names if n.startswith("map_image.")), None)
                if map_name:
                    ext = Path(map_name).suffix
                    img_dest = data_dir / f"map_image{ext}"
                    img_dest.write_bytes(zf.read(map_name))
                    map_cfg = data_dir / "snt_map_image.json"
                    with open(map_cfg, "w", encoding="utf-8") as f:
                        json.dump({"path": str(img_dest.resolve())}, f)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить проект:\n{e}")
            return

        # перезагружаем все виджеты из новых файлов
        self.plots.reload()
        self.energy_debt.rates.reload()
        self.vznosy_debt.rates.reload()

        if detail_df is not None:
            self.detail.load_dataframe(detail_df)
        else:
            self.energy_debt.refresh(self.detail.df_full)
            self.vznosy_debt.refresh(self.detail.df_full)

        QMessageBox.information(self, "Загружено", "Проект успешно загружен.")

    def _apply_styles(self):
        self.setStyleSheet("""
            /* ── Global ───────────────────────────────────────── */
            QMainWindow { background: #E9EDF3; }

            /* ── Custom title bar ────────────────────────────── */
            QWidget#titleBar {
                background: #E9EDF3;
            }
            QPushButton#btnWinMin, QPushButton#btnWinMax {
                background: transparent; border: none;
                color: #1A1A1A;
            }
            QPushButton#btnWinMin:hover  { background: rgba(0,0,0,9%); color: #1A1A1A; }
            QPushButton#btnWinMax:hover  { background: rgba(0,0,0,9%); color: #1A1A1A; }
            QPushButton#btnWinMin:pressed  { background: rgba(0,0,0,16%); }
            QPushButton#btnWinMax:pressed  { background: rgba(0,0,0,16%); }
            QPushButton#btnWinClose {
                background: transparent; border: none;
                color: #1A1A1A;
            }
            QPushButton#btnWinClose:hover   { background: #C42B1C; color: #FFFFFF; }
            QPushButton#btnWinClose:pressed { background: #B22418; color: #FFFFFF; }

            /* ── Left navigation sidebar ──────────────────────── */
            QWidget#sideNav {
                background: #E9EDF3;
            }
            QLabel#navLogo { background: transparent; }
            QWidget#navBtn { background: transparent; border-radius: 8px; }
            QWidget#navBtn:hover { background: #DDE2EC; }
            QWidget#navBtn[active="true"] { background: #07414F; }
            QLabel#navIcon  { color: #6B7686; background: transparent; }
            QLabel#navLabel { color: #3C4654; background: transparent; font-size: 14px; font-weight: 550; }
            QWidget#navBtn[active="true"] QLabel#navIcon  { color: #FFFFFF; }
            QWidget#navBtn[active="true"] QLabel#navLabel {
                color: #FFFFFF; font-weight: 600;
            }

            QWidget#btnNavAction {
                background: #FFFFFF;
                border: 1px solid #D5DCE4; border-radius: 8px;
            }
            QWidget#btnNavAction:hover { background: #D7DCE8; }
            QWidget#btnNavAction:pressed { background: #CBD2E0; }
            QLabel#actionLabel { color: #3C4654; background: transparent; font-size: 13px; font-weight: 600; }

            /* ── Content area ─────────────────────────────────── */
            QWidget#bodyArea { background: #E9EDF3; }
            QFrame#contentFrame {
                background: #F4F6FA;
                border-radius: 14px;
            }
            QStackedWidget#contentArea { background: transparent; }

            /* ── Page titles ─────────────────────────────────── */
            QLabel#pageTitle {
                font-size: 20px; font-weight: 700;
                color: #1F2937; background: transparent;
            }

            /* ── Dashboard «Главная» ──────────────────────────── */
            QScrollArea#homeScroll { background: #F0F3F9; border: none; }
            QWidget#homeContent { background: #F0F3F9; }
            QFrame#dashCard {
                background: #FFFFFF; border: 1px solid #E3E8EE; border-radius: 14px;
            }
            QLabel#cardTitleGreen {
                color: #57A05C; background: transparent;
                font-size: 17px; font-weight: 700;
            }
            QLabel#cardTitle {
                color: #1F2937; background: transparent;
                font-size: 15px; font-weight: 700;
            }
            QWidget#statCard {
                background: #F6F8FA; border: 1px solid #E6EAEF; border-radius: 10px;
            }
            QLabel#statIcon { color: #2F7D55; background: transparent; }
            QLabel#statCaption { color: #6B7280; background: transparent; font-size: 12px; }
            QLabel#statValue {
                color: #1F2937; background: transparent;
                font-size: 22px; font-weight: 700;
            }
            QWidget#activityItem { background: transparent; border-radius: 8px; }
            QWidget#activityItem:hover { background: #F1F5F9; }
            QLabel#activityCheck { color: #2E9E5B; background: transparent; }
            QLabel#activityTitle { color: #1F2937; background: transparent; font-size: 12px; }
            QLabel#activityDate { color: #9AA3AE; background: transparent; font-size: 11px; }
            QLabel#footerText { color: #9AA3AE; background: transparent; font-size: 12px; }
            QPushButton#chartPeriodBtn {
                background: #F6F8FA; color: #3C4654;
                border: 1px solid #D5DCE4; border-radius: 6px;
                padding: 5px 12px; font-size: 12px;
            }
            QPushButton#chartPeriodBtn:hover { background: #EEF1F5; }
            QPushButton#chartMenuBtn {
                background: #F6F8FA; color: #6B7280;
                border: 1px solid #D5DCE4; border-radius: 6px;
            }
            QPushButton#chartMenuBtn:hover { background: #EEF1F5; }

            /* ── Filter bar ──────────────────────────────────── */
            QFrame#filterFrame {
                background: #FFFFFF; border: 1px solid #E3E8EE; border-radius: 8px;
            }
            QLabel#filterLabel { color: #9AA3AE; background: transparent; font-size: 13px; }

            /* ── Inputs ──────────────────────────────────────── */
            QLineEdit#searchInput {
                background: #FFFFFF; border: 1px solid #D5DCE4; border-radius: 6px;
                color: #1F2937; padding: 7px 12px; font-size: 13px;
            }
            QLineEdit#searchInput:focus { border: 1px solid #07414F; }
            QComboBox#filterCombo {
                background: #FFFFFF; border: 1px solid #D5DCE4; border-radius: 6px;
                color: #1F2937; padding: 7px 10px; font-size: 13px;
            }
            QComboBox#filterCombo::drop-down { border: none; width: 18px; }
            QComboBox QAbstractItemView {
                background: #FFFFFF; border: 1px solid #D5DCE4;
                color: #1F2937; selection-background-color: #C9D8E2;
                selection-color: #07414F;
            }
            QDateEdit#datePicker {
                background: #FFFFFF; border: 1px solid #D5DCE4; border-radius: 6px;
                color: #1F2937; padding: 7px 10px; font-size: 13px;
            }
            QDateEdit#datePicker::drop-down { border: none; width: 18px; }

            /* ── Buttons ─────────────────────────────────────── */
            QPushButton#btnPrimary {
                background: #2F7D55; color: #FFFFFF; border: none; border-radius: 6px;
                padding: 8px 18px; font-size: 13px; font-weight: 600;
            }
            QPushButton#btnPrimary:hover   { background: #379061; }
            QPushButton#btnPrimary:pressed { background: #266645; }
            QPushButton#btnSecondary {
                background: #FFFFFF; color: #3C4654;
                border: 1px solid #D5DCE4; border-radius: 6px;
                padding: 8px 14px; font-size: 13px;
            }
            QPushButton#btnSecondary:hover { background: #F0F3F7; color: #1F2937; }

            /* ── Tables ──────────────────────────────────────── */
            QTableWidget#mainTable {
                background: #F9FAFB; border: 1px solid #D8DDE6; border-radius: 8px;
                gridline-color: #CDD3DC; color: #1F2937; font-size: 12px;
                selection-background-color: #C9D8E2; selection-color: #07414F;
                alternate-background-color: #E8ECF4;
            }
            QTableWidget#mainTable QHeaderView::section {
                background: #E9EDF5; color: #4B5563; border: none;
                border-right: 1px solid #CDD3DC; border-bottom: 2px solid #C4CBD7;
                padding: 8px 10px; font-size: 12px; font-weight: 600;
            }
            QTableWidget#mainTable::item {
                padding: 5px 10px; border-bottom: 1px solid #D8DDE6;
            }
            QTableWidget#mainTable::item:alternate {
                background: #E8ECF4;
            }
            QTableWidget#mainTable::item:hover {
                background: #DDE4EE;
            }
            QTableWidget#mainTable::item:selected {
                background: #C9D8E2; color: #07414F;
            }

            QTableWidget#summaryTable {
                background: #FFFFFF; border: 1px solid #E3E8EE; border-radius: 8px;
                gridline-color: #EAEDF1; color: #1F2937; font-size: 12px;
                selection-background-color: #E4F0E9; selection-color: #2F7D55;
            }
            QTableWidget#summaryTable QHeaderView::section {
                background: #EEF1F5; color: #6B7280; border: none;
                border-right: 1px solid #E3E8EE; border-bottom: 2px solid #E3E8EE;
                padding: 8px 10px; font-size: 12px; font-weight: 600;
            }
            QTableWidget#summaryTable::item {
                padding: 4px 10px; border-bottom: 1px solid #EAEDF1;
            }

            /* ── Viewport backgrounds (Qt QSS quirk fix) ────── */
            QAbstractScrollArea { background: #F0F3F9; }
            QAbstractScrollArea > QWidget { background: #F0F3F9; }

            /* ── Scrollbars ──────────────────────────────────── */
            QScrollBar:vertical {
                background: #E6E9EE; width: 8px; border-radius: 4px; border: none;
            }
            QScrollBar::handle:vertical {
                background: #C3CAD3; border-radius: 4px; min-height: 30px;
            }
            QScrollBar::handle:vertical:hover { background: #97A1AE; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
            QScrollBar:horizontal {
                background: #E6E9EE; height: 8px; border: none;
            }
            QScrollBar::handle:horizontal { background: #C3CAD3; border-radius: 4px; }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width: 0; }

            /* ── Status / summary labels ─────────────────────── */
            QLabel#statusLabel  { color: #9AA3AE; background: transparent; font-size: 12px; }
            QLabel#summaryIncome  {
                color: #2E9E5B; background: transparent; font-size: 13px; font-weight: 600;
            }
            QLabel#summaryExpense {
                color: #DC2626; background: transparent; font-size: 13px; font-weight: 600;
            }
        """)


def main():
    os.makedirs(DATA_DIR, exist_ok=True)
    _ensure_fonts()

    if sys.platform == "win32":
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("snt.helper.app")

    app = QApplication(sys.argv)
    app.setApplicationName("СНТ Финансовый учёт")

    _icon_path = Path(__file__).parent / "resources" / "images" / "logo_2.ico"
    if _icon_path.exists():
        from PyQt6.QtGui import QIcon
        app.setWindowIcon(QIcon(str(_icon_path)))

    fonts_dir = Path(__file__).parent / "resources" / "fonts"
    for font_file in list(fonts_dir.glob("*.ttf")) + list(fonts_dir.glob("*.OTF")) + list(fonts_dir.glob("*.otf")):
        QFontDatabase.addApplicationFont(str(font_file))

    base_font = QFont("Segoe UI", 10)
    base_font.setStyleStrategy(
        QFont.StyleStrategy.PreferAntialias | QFont.StyleStrategy.PreferQuality
    )
    base_font.setHintingPreference(QFont.HintingPreference.PreferNoHinting)
    app.setFont(base_font)

    palette = app.palette()
    palette.setColor(QPalette.ColorRole.Window,      QColor("#F0F3F9"))
    palette.setColor(QPalette.ColorRole.Base,        QColor("#FFFFFF"))
    palette.setColor(QPalette.ColorRole.AlternateBase, QColor("#E8ECF4"))
    palette.setColor(QPalette.ColorRole.WindowText,  QColor("#1F2937"))
    palette.setColor(QPalette.ColorRole.Text,        QColor("#1F2937"))
    palette.setColor(QPalette.ColorRole.Button,      QColor("#F0F3F9"))
    palette.setColor(QPalette.ColorRole.ButtonText,  QColor("#1F2937"))
    palette.setColor(QPalette.ColorRole.Highlight,   QColor("#07414F"))
    palette.setColor(QPalette.ColorRole.HighlightedText, QColor("#FFFFFF"))
    app.setPalette(palette)

    window = MainWindow()
    if _icon_path.exists():
        from PyQt6.QtGui import QIcon
        window.setWindowIcon(QIcon(str(_icon_path)))
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
