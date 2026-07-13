import hashlib
import json
import os
import time

import pandas as pd
from PyQt6.QtCore import (
    Qt, QDate, QEvent, QPoint, QRect, QRectF, QSize, QModelIndex,
    QAbstractItemModel, QTimer, pyqtSignal,
)
from PyQt6.QtGui import (
    QAction, QBitmap, QBrush, QColor, QFont, QFontMetrics, QPainter, QPen,
    QPolygon, QRegion,
)
from PyQt6.QtWidgets import (
    QAbstractItemView, QApplication, QButtonGroup, QComboBox, QDateEdit, QDialog,
    QFileDialog, QFormLayout, QFrame, QGridLayout,
    QHBoxLayout, QHeaderView, QLabel, QLineEdit, QMenu,
    QPushButton, QScrollArea, QStyle, QStyleFactory, QStyledItemDelegate, QStyleOptionViewItem,
    QTreeView, QVBoxLayout, QWidget, QCompleter,
)

from ui.categorization import (
    CATEGORY_COLORS, ALL_CATEGORIES, apply_categorization,
    save_user_categories, save_user_category_color, rename_user_category,
    delete_user_category, PROTECTED_CATEGORIES, NO_CATEGORY_LABEL,
)
from ui.plot_detection import apply_plot_column, _PLOTS_FILE, load_plot_numbers
from ui import icons
from ui.buttons import LinkButton, PrimaryButton, SecondaryButton
from ui.plots_widget import _FilterTabButton
from ui.common import (
    AppTooltip as _AppTooltip,
    CalendarArrowFlip,
    ClipFrame as _ClipFrame,
    NoJumpDateEdit,
    TooltipFilter as _TooltipFilter,
    TREE_STYLE as _TREE_STYLE_COMMON,
    style_date_popup,
)
from ui.dialogs import (
    AlertDialog as _AlertDialog,
    BaseDialog as _FramelessDialog,
    ConfirmDialog as _ConfirmDialog,
    PromptDialog as _BasePromptDialog,
    exec_dialog as _exec_dialog,
)
from ui.theme import C, FS, menu_qss, scrollbar_qss


# =========================================================================== #
#  Вспомогательные функции для DataFrame                                      #
# =========================================================================== #

def _merge_to_summa(df: "pd.DataFrame") -> "pd.DataFrame":
    """Если в DataFrame есть Поступление/Списание — объединяет их в Сумма.
    Если Сумма уже есть — ничего не делает."""
    if "Сумма" in df.columns:
        return df
    if "Поступление" not in df.columns and "Списание" not in df.columns:
        return df
    inc = pd.to_numeric(df["Поступление"], errors="coerce") if "Поступление" in df.columns \
        else pd.Series(0.0, index=df.index)
    exp = pd.to_numeric(df["Списание"],    errors="coerce") if "Списание"    in df.columns \
        else pd.Series(0.0, index=df.index)
    summa = inc.fillna(0) - exp.fillna(0)
    summa = summa.where(summa != 0, other=float("nan"))
    pos = list(df.columns).index("Поступление") if "Поступление" in df.columns else len(df.columns)
    df = df.drop(columns=[c for c in ("Поступление", "Списание") if c in df.columns])
    df.insert(min(pos, len(df.columns)), "Сумма", summa)
    return df


def _compute_hash(row: dict) -> str:
    """SHA-256 (12 символов) по ключевым полям — стабильный ID операции."""
    parts = [
        str(row.get("Дата", ""))[:10],
        str(row.get("Сумма", "")),
        str(row.get("Контрагент") or "").strip(),
        str(row.get("Назначение") or "").strip(),
    ]
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()[:12]


def _ensure_meta_columns(df: "pd.DataFrame") -> "pd.DataFrame":
    """Добавляет _hash и _breakdown, если их ещё нет."""
    df = df.copy()
    if "_hash" not in df.columns:
        df["_hash"] = df.apply(lambda r: _compute_hash(r.to_dict()), axis=1)
    if "_breakdown" not in df.columns:
        df["_breakdown"] = None
    return df


def _parse_breakdown(value) -> list:
    """Читает разбивку операции (хранится в колонке _breakdown как JSON-строка).
    Возвращает список словарей вида {Назначение, Сумма, Категория}."""
    if isinstance(value, list):
        return value
    if isinstance(value, str) and value.strip():
        try:
            data = json.loads(value)
            return data if isinstance(data, list) else []
        except (ValueError, TypeError):
            return []
    return []


def _dump_breakdown(items: list) -> str:
    """Сериализует разбивку в JSON-строку для хранения в _breakdown."""
    if not items:
        return ""
    return json.dumps(items, ensure_ascii=False)


# =========================================================================== #
#  Парсинг и форматирование значений                                          #
# =========================================================================== #

def _to_num(v):
    """Приводит значение к float или None (для пустых/нечисловых)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return None if (isinstance(v, float) and pd.isna(v)) else float(v)
    try:
        return float(
            str(v).replace(" ", "").replace(" ", "")
                  .replace("−", "-").replace("₽", "").replace(",", ".")
        )
    except ValueError:
        return None


def _parse_money(text: str) -> float:
    n = _to_num(text)
    return float("nan") if n is None else n


def _fmt_money(num: float) -> str:
    if num > 0:
        return f"{num:,.2f} ₽".replace(",", " ")
    if num < 0:
        return f"−{abs(num):,.2f} ₽".replace(",", " ")
    return ""


def _num_edit_str(num: float) -> str:
    """Чистое строковое представление числа для редактора (без ₽ и пробелов)."""
    return ("%g" % num)


def _to_ts(v):
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    try:
        ts = pd.Timestamp(v)
        return None if pd.isna(ts) else ts
    except (ValueError, TypeError):
        return None


def _parse_date(text: str):
    try:
        ts = pd.to_datetime(text, dayfirst=True)
        return None if pd.isna(ts) else ts
    except (ValueError, TypeError):
        return None


# =========================================================================== #
#  Узел дерева и модель                                                        #
# =========================================================================== #

# Роль, по которой делегат понимает, что ячейка была отредактирована вручную.
MANUAL_ROLE = Qt.ItemDataRole.UserRole + 1
# Роль: операция помечена как повторно импортированная (см. DetailWidget._dup_pending).
DUP_ROLE = Qt.ItemDataRole.UserRole + 2

_DEFAULT_ROW_COLOR = QColor(55, 55, 60)
# Заглушка для родительской строки, у которой есть дочерние строки-распределения.
_MULTI_OP_LABEL = "Мультиоперация"
_MULTI_CAT_LABEL = "Несколько категорий"
_MULTI_PLOT_LABEL = "Несколько участков"
# Служебный столбец кнопки редактирования. Хранится последним в модели.
_EDIT_COL = "\x00edit"
# Служебный столбец чекбокса выбора строк.
_CHECK_COL = "\x00check"
# Колонки, отображаемые в строке-ветке (split). «Контрагент» — копия из операции.
_SPLIT_DISPLAY_COLS = ("Контрагент", "Сумма", "Категория", "Участок")
# Из них реально редактируемые в ветке.
_SPLIT_EDIT_COLS = ("Сумма", "Категория", "Участок")


class _Node:
    """Узел дерева. kind="op" — операция (строка выписки),
    kind="split" — распределение внутри операции."""

    __slots__ = ("kind", "data", "df_idx", "parent", "children", "orig_idx")

    def __init__(self, kind: str, data: dict, df_idx=None, parent=None):
        self.kind = kind
        self.data = data          # словарь {имя_колонки: значение}
        self.df_idx = df_idx      # индекс в df_full (только для операций)
        self.parent = parent
        self.children: list[_Node] = []
        self.orig_idx = 0         # позиция среди братьев ДО сортировки — see _apply_sort_internal

    def row(self) -> int:
        if self.parent is not None:
            return self.parent.children.index(self)
        return 0


class OperationsTreeModel(QAbstractItemModel):
    """Двухуровневая модель: операции -> распределения по категориям.

    Сумма операции НЕ вычисляется из детей — она остаётся исходным значением
    выписки. Дети-распределения это аннотация (на что пошли деньги)."""

    # Эмитится при ручном редактировании ячейки: (узел, имя_колонки).
    cellEdited = pyqtSignal(object, str)

    def __init__(self, manual_cells: set, dup_pending: dict, parent=None):
        super().__init__(parent)
        self._manual = manual_cells
        self._dup_pending = dup_pending
        self._columns: list[str] = []
        self._root = _Node("root", {})
        self._sort_col: int = -1    # -1 = сортировка не активна (3-е состояние)
        self._sort_order = Qt.SortOrder.AscendingOrder

    # -- наполнение --------------------------------------------------------- #
    def load(self, columns: list[str], records: list[tuple]):
        """records: список (df_idx, data_dict, breakdown_list)."""
        self.beginResetModel()
        # Чекбокс выбора — первым столбцом (как в «Участках»), карандаш
        # редактирования — по-прежнему последним.
        self._columns = [_CHECK_COL] + list(columns) + [_EDIT_COL]
        root = _Node("root", {})
        for df_idx, data, breakdown in records:
            op = _Node("op", data, df_idx=df_idx, parent=root)
            op.orig_idx = len(root.children)
            for b in breakdown:
                child = _Node("split", dict(b), parent=op)
                child.orig_idx = len(op.children)
                op.children.append(child)
            root.children.append(op)
        self._root = root
        self._apply_sort_internal()
        self.endResetModel()

    def top_nodes(self) -> list:
        return self._root.children

    def category_column(self) -> int:
        return self._columns.index("Категория") if "Категория" in self._columns else -1

    def columns(self) -> list:
        return self._columns

    def index_for_df_idx(self, df_idx) -> QModelIndex:
        for i, op in enumerate(self._root.children):
            if op.df_idx == df_idx:
                return self.createIndex(i, 0, op)
        return QModelIndex()

    # -- ядро дерева -------------------------------------------------------- #
    def index(self, row, column, parent=QModelIndex()):
        if not self.hasIndex(row, column, parent):
            return QModelIndex()
        parent_node = parent.internalPointer() if parent.isValid() else self._root
        if 0 <= row < len(parent_node.children):
            return self.createIndex(row, column, parent_node.children[row])
        return QModelIndex()

    def parent(self, index):
        if not index.isValid():
            return QModelIndex()
        node = index.internalPointer()
        p = node.parent
        if p is None or p is self._root:
            return QModelIndex()
        return self.createIndex(p.row(), 0, p)

    def rowCount(self, parent=QModelIndex()):
        if parent.column() > 0:
            return 0
        node = parent.internalPointer() if parent.isValid() else self._root
        return len(node.children)

    def columnCount(self, parent=QModelIndex()):
        return len(self._columns)

    # -- чтение ------------------------------------------------------------- #
    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        node = index.internalPointer()
        col = self._columns[index.column()]

        # Служебные столбцы — всё рисует делегат, модель данных не отдаёт.
        if col in (_EDIT_COL, _CHECK_COL):
            return None

        if role == MANUAL_ROLE:
            return node.kind == "op" and (node.df_idx, col) in self._manual

        if role == DUP_ROLE:
            return node.kind == "op" and node.df_idx in self._dup_pending

        if role in (Qt.ItemDataRole.DisplayRole, Qt.ItemDataRole.EditRole):
            if node.kind == "split" and col not in _SPLIT_DISPLAY_COLS:
                return ""
            if role == Qt.ItemDataRole.DisplayRole and node.kind == "op" and node.children:
                if col == "Участок":
                    plots = set()
                    for ch in node.children:
                        p = str(ch.data.get("Участок", "")).strip()
                        if p:
                            plots.add(p)
                    if len(plots) == 1:
                        return plots.pop()
                    if len(plots) == 0:
                        return _MULTI_OP_LABEL
                    return _MULTI_PLOT_LABEL
                if col == "Категория":
                    cats = set()
                    for ch in node.children:
                        c = str(ch.data.get("Категория", "")).strip()
                        if c:
                            cats.add(c)
                    if len(cats) == 1:
                        return cats.pop()
                    if len(cats) == 0:
                        return _MULTI_OP_LABEL
                    return _MULTI_CAT_LABEL
            val = node.data.get(col)
            if col == "Сумма":
                num = _to_num(val)
                if role == Qt.ItemDataRole.EditRole:
                    return "" if num is None else _num_edit_str(num)
                return "" if not num else _fmt_money(num)
            if col == "Дата":
                ts = _to_ts(val)
                return "" if ts is None else ts.strftime("%d.%m.%Y")
            if val is None or (isinstance(val, float) and pd.isna(val)):
                return ""
            return str(val)

        if role == Qt.ItemDataRole.ForegroundRole:
            if col == "Сумма":
                num = _to_num(node.data.get(col))
                if num is not None and num > 0:
                    return QColor("#059669")
                if num is not None and num < 0:
                    return QColor("#DC2626")
                return QColor("#374151")

        if role == Qt.ItemDataRole.BackgroundRole:
            if node.kind == "split":
                return None
            if node.kind == "op" and node.df_idx in self._dup_pending:
                return QColor("#FFF3CD")
            cat = str(node.data.get("Категория", ""))
            return CATEGORY_COLORS.get(cat, _DEFAULT_ROW_COLOR)

        if role == Qt.ItemDataRole.ToolTipRole:
            if node.kind == "op" and node.df_idx in self._dup_pending:
                return ("Эта операция встретилась ещё раз при повторном импорте.\n"
                        "ПКМ по строке — восстановить исходные данные или "
                        "оставить как есть.")

        if role == Qt.ItemDataRole.TextAlignmentRole:
            if col == "Сумма":
                return int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            return int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        return None

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if orientation == Qt.Orientation.Horizontal and role == Qt.ItemDataRole.DisplayRole:
            if 0 <= section < len(self._columns):
                name = self._columns[section]
                if name == _EDIT_COL:   return chr(0xE3C9)
                if name == _CHECK_COL:  return ""
                if name == "Участок":   return "№ уч."
                return name
        return None

    # -- редактирование ----------------------------------------------------- #
    def flags(self, index):
        if not index.isValid():
            return Qt.ItemFlag.NoItemFlags
        node = index.internalPointer()
        col = self._columns[index.column()]
        f = Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable
        if col in (_EDIT_COL, _CHECK_COL):
            return f
        if node.kind == "op":
            # Категория и Участок недоступны для редактирования у строк с дочерними
            if col in ("Категория", "Участок") and node.children:
                pass
            else:
                f |= Qt.ItemFlag.ItemIsEditable
        elif col in _SPLIT_EDIT_COLS:
            f |= Qt.ItemFlag.ItemIsEditable
        return f

    def setData(self, index, value, role=Qt.ItemDataRole.EditRole):
        if role != Qt.ItemDataRole.EditRole or not index.isValid():
            return False
        if not (self.flags(index) & Qt.ItemFlag.ItemIsEditable):
            return False
        node = index.internalPointer()
        col = self._columns[index.column()]
        text = str(value).strip()

        if col == "Сумма":
            node.data["Сумма"] = _parse_money(text)
        elif col == "Дата":
            ts = _parse_date(text)
            if ts is None:
                return False
            node.data["Дата"] = ts
        else:
            node.data[col] = text

        # Перерисовываем всю строку — фон зависит от категории, которая могла измениться.
        left = self.index(index.row(), 0, index.parent())
        right = self.index(index.row(), self.columnCount() - 1, index.parent())
        self.dataChanged.emit(left, right)
        self.cellEdited.emit(node, col)
        return True

    # -- сортировка --------------------------------------------------------- #
    def sort(self, column, order=Qt.SortOrder.AscendingOrder):
        if 0 <= column < len(self._columns) and self._columns[column] in (_EDIT_COL, _CHECK_COL):
            return
        self._sort_col = column
        self._sort_order = order
        self.beginResetModel()
        self._apply_sort_internal()
        self.endResetModel()

    def _apply_sort_internal(self):
        if not (0 <= self._sort_col < len(self._columns)):
            # Сортировка не активна (3-е состояние — сброс) — возвращаем
            # исходный порядок вставки вместо того, чтобы оставить строки
            # в порядке последней применённой сортировки.
            self._root.children.sort(key=lambda n: n.orig_idx)
            for op in self._root.children:
                op.children.sort(key=lambda n: n.orig_idx)
            return
        col = self._columns[self._sort_col]
        if col == _EDIT_COL:
            return
        reverse = self._sort_order == Qt.SortOrder.DescendingOrder

        def key(node):
            v = node.data.get(col)
            if col == "Сумма":
                num = _to_num(v)
                return (num is None, num if num is not None else 0.0)
            if col == "Дата":
                ts = _to_ts(v)
                return (ts is None, ts if ts is not None else pd.Timestamp.min)
            return (False, str(v or "").lower())

        self._root.children.sort(key=key, reverse=reverse)
        for op in self._root.children:
            op.children.sort(key=key, reverse=reverse)


# =========================================================================== #
#  Делегаты                                                                   #
# =========================================================================== #

class _CellDelegate(QStyledItemDelegate):
    """Базовый делегат: рисует иконку карандаша (Material Icons) в ячейках,
    отредактированных вручную."""

    _CHAR = ""
    _ICON_FONT: QFont | None = None

    @classmethod
    def _icon_font(cls) -> QFont:
        if cls._ICON_FONT is None:
            cls._ICON_FONT = icons.icon_font(14)
        return cls._ICON_FONT

    def paint(self, painter, option, index):
        super().paint(painter, option, index)
        if not index.data(MANUAL_ROLE):
            return
        painter.save()
        painter.setFont(self._icon_font())
        if option.state & QStyle.StateFlag.State_Selected:
            painter.setPen(QColor(255, 255, 255, 180))
        else:
            painter.setPen(QColor("#9CA3AF"))
        painter.drawText(
            option.rect.adjusted(0, 0, -4, 0),
            Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight,
            self._CHAR,
        )
        painter.restore()

    # Инлайн-редактор ячейки: непрозрачный белый фон ОБЯЗАТЕЛЕН — без явного
    # QSS дефолтный QLineEdit делегата остаётся прозрачным (при активных
    # стилях приложения Qt не рисует его нативную панель), и сквозь редактор
    # просвечивает текст ячейки, отрисованный делегатом, — на экране «каша»
    # из двух наложенных текстов. Бирюзовая рамка — тот же язык, что и у
    # полей ввода приложения в фокусе.
    _EDITOR_SS = (
        "QLineEdit{background:#FFFFFF;border:1px solid #07414F;border-radius:4px;"
        "padding:0 6px;font-size:13px;color:#1F2937;"
        "selection-background-color:#C9D8E2;selection-color:#07414F;}"
    )

    def createEditor(self, parent, option, index):
        editor = super().createEditor(parent, option, index)
        if editor is not None:
            editor.setStyleSheet(self._EDITOR_SS)
            # QLineEdit по умолчанию САМ подгоняет себя под sizeHint() уже
            # ПОСЛЕ updateEditorGeometry (Qt пересчитывает геометрию редактора
            # при показе/фокусе) — sizeHint строится по ПОЛНОМУ, неусечённому
            # тексту DisplayRole. Для длинных значений (например,
            # «Контрагент» многооперационной строки — реальный текст не
            # умещается в столбец и обрезается только в paint()) редактор
            # раздувался на сотни пикселей поверх соседних столбцов.
            # setGeometry() в updateEditorGeometry этого не остановит — сам
            # по себе он вызывается раньше и корректно, но редактор всё равно
            # перерастягивает себя следом; setMaximumWidth — жёсткий предел,
            # который Qt соблюдает при ЛЮБОМ последующем resize.
            editor.setMaximumWidth(option.rect.width())
        return editor

    def updateEditorGeometry(self, editor, option, index):
        # Тот же приём, что и в _CategoryDelegate/_PlotDelegate — жёстко
        # привязываем редактор к границам ячейки.
        editor.setGeometry(option.rect)


class _DetailCheckDelegate(QStyledItemDelegate):
    """Делегат столбца чекбокса выбора строк для удаления."""

    selectionChanged = pyqtSignal()

    # Два разных глифа (не один + FILL-переключение) — тот же приём, что и
    # у чекбоксов участков в ui/plots_widget.py (_PlotRowDelegate).
    _IC_ON    = chr(0xE834)   # check_box
    _IC_OFF   = chr(0xE835)   # check_box_outline_blank
    _IC_FONT  = "Material Symbols Rounded"
    _IC_COLOR_ON  = QColor("#07414F")
    _IC_COLOR_OFF = QColor("#C3CAD3")

    def __init__(self, view):
        super().__init__(view)
        self._view = view
        self._hover_idx = QModelIndex()
        self._pointing = False
        self._fill_tag = QFont.Tag.fromString("FILL")
        self._selected: set[int] = set()
        view.viewport().installEventFilter(self)

    def get_selected(self) -> set[int]:
        return set(self._selected)

    def select_all(self, df_indices) -> None:
        self._selected.update(df_indices)
        self.selectionChanged.emit()
        self._view.viewport().update()

    def clear_selection(self):
        if self._selected:
            self._selected.clear()
            self.selectionChanged.emit()
            self._view.viewport().update()

    def _is_btn(self, index: QModelIndex) -> bool:
        node = index.internalPointer() if index.isValid() else None
        if not node or node.kind != "op":
            return False
        cols = self._columns
        return 0 <= index.column() < len(cols) and cols[index.column()] == _CHECK_COL

    @property
    def _columns(self):
        return self._view.model()._columns if self._view.model() else []

    def paint(self, painter, option, index):
        super().paint(painter, option, index)
        node = index.internalPointer()
        if node is None or node.kind != "op":
            return
        cols = self._columns
        if index.column() >= len(cols) or cols[index.column()] != _CHECK_COL:
            return
        df_idx = node.df_idx
        checked = df_idx in self._selected
        icon = self._IC_ON if checked else self._IC_OFF
        painter.save()
        f = QFont(self._IC_FONT)
        f.setPixelSize(18)
        f.setVariableAxis(self._fill_tag, 1.0 if checked else 0.0)
        painter.setFont(f)
        painter.setPen(self._IC_COLOR_ON if checked else self._IC_COLOR_OFF)
        painter.drawText(option.rect, Qt.AlignmentFlag.AlignCenter, icon)
        painter.restore()

    def editorEvent(self, event, model, option, index):
        if (event.type() == QEvent.Type.MouseButtonRelease
                and event.button() == Qt.MouseButton.LeftButton
                and self._is_btn(index)):
            node = index.internalPointer()
            if node and node.df_idx is not None:
                df_idx = node.df_idx
                if df_idx in self._selected:
                    self._selected.discard(df_idx)
                else:
                    self._selected.add(df_idx)
                self._view.viewport().update(self._view.visualRect(index))
                self.selectionChanged.emit()
            return True
        return super().editorEvent(event, model, option, index)

    def eventFilter(self, obj, event):
        try:
            viewport = self._view.viewport()
        except RuntimeError:
            # self._view уже удалён — хвост событий при закрытии приложения.
            return False
        if obj is viewport:
            if event.type() == QEvent.Type.MouseMove:
                idx = self._view.indexAt(event.position().toPoint())
                on_btn = self._is_btn(idx)

                if on_btn and not self._pointing:
                    self._pointing = True
                    QApplication.setOverrideCursor(Qt.CursorShape.PointingHandCursor)
                elif not on_btn and self._pointing:
                    self._pointing = False
                    QApplication.restoreOverrideCursor()

                new_hover = idx if on_btn else QModelIndex()
                if new_hover != self._hover_idx:
                    old = self._hover_idx
                    self._hover_idx = new_hover
                    if old.isValid():
                        self._view.viewport().update(self._view.visualRect(old))
                    if new_hover.isValid():
                        self._view.viewport().update(self._view.visualRect(new_hover))

            elif event.type() == QEvent.Type.Leave:
                if self._pointing:
                    self._pointing = False
                    QApplication.restoreOverrideCursor()
                if self._hover_idx.isValid():
                    old = self._hover_idx
                    self._hover_idx = QModelIndex()
                    self._view.viewport().update(self._view.visualRect(old))

        return super().eventFilter(obj, event)


class _CategoryDelegate(_CellDelegate):
    """Делегат колонки «Категория»: рисует цветной овальный badge;
    при редактировании открывает popup выбора категории (_CategoryPillButton /
    _SingleCatPopup) — те же пилюли, что и в диалогах Добавить/Редактировать."""

    # Цвета фона/hover/выделения совпадают с _TREE_STYLE
    _BG       = QColor("#FFFFFF")
    _BG_ALT   = QColor("#F0F4F8")
    _BG_HOVER = QColor("#DDE4EE")
    _BG_SEL   = QColor("#C9D8E2")
    _BORDER   = QColor("#E3E8EF")
    _TXT_SEL  = QColor("#07414F")
    _ARROW_W  = 16   # шеврон раскрытия — ВНУТРИ пилюли у её правого края

    editCategoriesRequested = pyqtSignal()

    def __init__(self, items: list[str], parent=None):
        super().__init__(parent)
        self._items = items
        # Qt для QTreeView отдаёт State_MouseOver на ВСЮ строку (все колонки
        # сразу), а не на конкретную ячейку под курсором — поэтому ховер
        # пилюли отслеживаем сами, тем же приёмом, что и зона кнопки
        # «Свернуть/Показать» в _BranchColumnDelegate ниже.
        self._view = parent
        self._hover_index = QModelIndex()
        if parent is not None:
            parent.viewport().installEventFilter(self)

    def _update_hover(self, pos):
        new_hover = self._view.indexAt(pos) if pos is not None else QModelIndex()
        if new_hover != self._hover_index:
            old = self._hover_index
            self._hover_index = new_hover
            if old.isValid():
                self._view.update(old)
            if new_hover.isValid():
                self._view.update(new_hover)

    # ---- отрисовка -------------------------------------------------------- #

    def paint(self, painter, option, index):
        text = str(index.data(Qt.ItemDataRole.DisplayRole) or "")

        if text == _MULTI_OP_LABEL:
            self._paint_hatched(painter, option)
            return

        if text == _MULTI_CAT_LABEL:
            self._paint_multi_cat(painter, option, text)
            return

        color = CATEGORY_COLORS.get(text) if text else None
        if color is None:
            super().paint(painter, option, index)
            return

        painter.save()

        rect = option.rect
        selected = bool(option.state & QStyle.StateFlag.State_Selected)
        # option.state & State_MouseOver — это ховер ВСЕЙ строки в QTreeView
        # (все колонки разом), а не конкретной ячейки. Настоящий ховер
        # именно этой ячейки — self._hover_index, который сами отслеживаем
        # в eventFilter() выше.
        hovered = (index == self._hover_index)

        # 1. Фон ячейки. Ховер сюда НЕ подмешиваем — на ховере темнеет сама
        # пилюля (см. ниже), а общее потемнение всей ячейки убрано, чтобы
        # два эффекта не спорили друг с другом и не мигали.
        is_alt = bool(option.features & QStyleOptionViewItem.ViewItemFeature.Alternate)
        if selected:
            painter.fillRect(rect, self._BG_SEL)
        else:
            painter.fillRect(rect, self._BG_ALT if is_alt else self._BG)

        # 2. Нижняя граница строки
        painter.setPen(QPen(self._BORDER, 1))
        painter.drawLine(rect.bottomLeft(), rect.bottomRight())
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        # 3. Цвета овала из HSL базового цвета — на ховере темнее (сигнал
        # «кликабельно»), без изменения фона всей ячейки.
        h, s, l, _ = color.getHslF()
        if h < 0:
            h, s = 0.0, 0.0
        bg_l, bd_l = (0.91, 0.52) if not hovered else (0.82, 0.44)
        pill_bg = QColor.fromHslF(h, min(s * 0.65, 1.0), bg_l)
        pill_bd = QColor.fromHslF(h, min(s * 1.00, 1.0), bd_l)
        pill_tx = QColor.fromHslF(h, min(s * 1.20, 1.0), 0.18)

        # 4. Геометрия овала — ВСЕГДА одна и та же ширина, независимо от
        # MANUAL_ROLE. Раньше карандаш резервировал 20px СНАРУЖИ пилюли,
        # из-за чего у вручную изменённых строк она была заметно уже, чем
        # у остальных с тем же текстом — карандаш переехал ВНУТРЬ (см. п.6).
        v = max(3, (rect.height() - 20) // 2)
        pill   = rect.adjusted(6, v, -6, -v)
        pill_f = QRectF(pill).adjusted(0.5, 0.5, -0.5, -0.5)
        radius_f = pill_f.height() / 2.0

        painter.setPen(QPen(pill_bd, 1))
        painter.setBrush(pill_bg)
        painter.drawRoundedRect(pill_f, radius_f, radius_f)

        # 5. Текст — отступ справа под шеврон (+карандаш, если он есть).
        manual = bool(index.data(MANUAL_ROLE))
        pencil_w = 14 if manual else 0
        painter.setPen(self._TXT_SEL if selected else pill_tx)
        painter.setFont(option.font)
        text_rect = pill.adjusted(8, 0, -self._ARROW_W - pencil_w - 4, 0)
        fm = QFontMetrics(option.font)
        elided = fm.elidedText(text, Qt.TextElideMode.ElideRight, text_rect.width())
        painter.drawText(
            text_rect,
            int(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft),
            elided,
        )

        # 6. Карандаш (ручное редактирование) — ВНУТРИ пилюли, левее
        # шеврона (не резервирует место снаружи — пилюля не сужается).
        if manual:
            pencil_rect = QRect(pill.right() - self._ARROW_W - pencil_w, pill.top(),
                                pencil_w, pill.height())
            painter.setFont(self._icon_font())
            painter.setPen(
                QColor(255, 255, 255, 180) if selected else pill_tx.lighter(160)
            )
            painter.drawText(pencil_rect, Qt.AlignmentFlag.AlignCenter, self._CHAR)

        # 7. Шеврон раскрытия списка — ВНУТРИ пилюли у её правого края
        # (как в референсе), тем же цветом текста пилюли.
        arrow_rect = QRect(pill.right() - self._ARROW_W, pill.top(),
                           self._ARROW_W, pill.height())
        painter.setFont(self._icon_font())
        painter.setPen(self._TXT_SEL if selected else pill_tx)
        painter.drawText(arrow_rect, Qt.AlignmentFlag.AlignCenter, chr(0xE5CF))  # expand_more

        painter.restore()

    def _paint_multi_cat(self, painter, option, text):
        """Серый овал с текстом (без штриховки) для строки с несколькими категориями."""
        painter.save()

        rect = option.rect
        selected = bool(option.state & QStyle.StateFlag.State_Selected)
        hovered  = bool(option.state & QStyle.StateFlag.State_MouseOver)

        is_alt = bool(option.features & QStyleOptionViewItem.ViewItemFeature.Alternate)
        if selected:
            painter.fillRect(rect, self._BG_SEL)
        else:
            painter.fillRect(rect, self._BG_ALT if is_alt else self._BG)

        painter.setPen(QPen(self._BORDER, 1))
        painter.drawLine(rect.bottomLeft(), rect.bottomRight())
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        v = max(3, (rect.height() - 20) // 2)
        pill   = rect.adjusted(6, v, -6, -v)
        pill_f = QRectF(pill).adjusted(0.5, 0.5, -0.5, -0.5)
        radius_f = pill_f.height() / 2.0

        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QColor(234, 234, 238))
        painter.drawRoundedRect(pill_f, radius_f, radius_f)

        painter.setPen(QPen(QColor(165, 165, 173), 1))
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawRoundedRect(pill_f, radius_f, radius_f)

        painter.setPen(QColor(110, 110, 118))
        painter.setFont(option.font)
        text_rect = pill.adjusted(8, 0, -4, 0)
        fm = QFontMetrics(option.font)
        elided = fm.elidedText(text, Qt.TextElideMode.ElideRight, text_rect.width())
        painter.drawText(
            text_rect,
            int(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft),
            elided,
        )

        painter.restore()

    def _paint_hatched(self, painter, option):
        """Овал с диагональной штриховкой для строки-мультиоперации."""
        painter.save()

        rect = option.rect
        selected = bool(option.state & QStyle.StateFlag.State_Selected)
        hovered  = bool(option.state & QStyle.StateFlag.State_MouseOver)

        is_alt = bool(option.features & QStyleOptionViewItem.ViewItemFeature.Alternate)
        if selected:
            painter.fillRect(rect, self._BG_SEL)
        else:
            painter.fillRect(rect, self._BG_ALT if is_alt else self._BG)

        painter.setPen(QPen(self._BORDER, 1))
        painter.drawLine(rect.bottomLeft(), rect.bottomRight())
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        v = max(3, (rect.height() - 20) // 2)
        pill   = rect.adjusted(6, v, -6, -v)
        pill_f = QRectF(pill).adjusted(0.5, 0.5, -0.5, -0.5)
        radius_f = pill_f.height() / 2.0

        # Светлый фон овала
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QColor(234, 234, 238))
        painter.drawRoundedRect(pill_f, radius_f, radius_f)

        # Диагональная штриховка
        painter.setBrush(QBrush(QColor(185, 185, 193), Qt.BrushStyle.BDiagPattern))
        painter.drawRoundedRect(pill_f, radius_f, radius_f)

        # Граница
        painter.setPen(QPen(QColor(165, 165, 173), 1))
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawRoundedRect(pill_f, radius_f, radius_f)

        painter.restore()

    # ---- редактирование -------------------------------------------------- #

    def createEditor(self, parent, option, index):
        editor = _CategoryPillButton(parent=parent, auto_open=True, show_edit_button=True)
        editor.addItems(self._items)
        editor.userPicked.connect(lambda e=editor: self._commit(e))
        editor.popupDismissed.connect(lambda e=editor: self._cancel(e))
        editor.editCategoriesRequested.connect(self.editCategoriesRequested)
        editor.editCategoriesRequested.connect(lambda e=editor: self._cancel(e))
        return editor

    def eventFilter(self, obj, event):
        # У этого делегата ДВЕ независимые причины стоять в installEventFilter:
        # 1) obj — открытый РЕДАКТОР ячейки (_CategoryPillButton). QStyledItemDelegate
        #    по умолчанию коммитит+закрывает редактор на FocusOut — а popup выбора
        #    (_SingleCatPopup) не является потомком editor (Qt.WindowType.Popup,
        #    parent=None), поэтому его открытие выглядит для Qt как «фокус ушёл
        #    из редактора совсем», и редактор закрывается ДО того, как пользователь
        #    успевает кликнуть по категории. Закрытием управляем сами через
        #    userPicked/popupDismissed/editCategoriesRequested — стандартный
        #    FocusOut здесь подавляем.
        # 2) obj — viewport дерева (см. __init__: parent.viewport().installEventFilter).
        #    Отслеживаем реально наведённую ячейку по MouseMove/Leave — Qt для
        #    QTreeView отдаёт State_MouseOver на всю строку разом, а не на
        #    конкретную ячейку под курсором.
        if isinstance(obj, _CategoryPillButton) and event.type() == QEvent.Type.FocusOut:
            return True
        if self._view is not None:
            try:
                viewport = self._view.viewport()
            except RuntimeError:
                viewport = None
            if viewport is not None and obj is viewport:
                if event.type() == QEvent.Type.MouseMove:
                    self._update_hover(event.position().toPoint())
                elif event.type() == QEvent.Type.Leave:
                    self._update_hover(None)
                return False
        return super().eventFilter(obj, event)

    def _commit(self, editor):
        self.commitData.emit(editor)
        self.closeEditor.emit(editor)

    def _cancel(self, editor):
        self.closeEditor.emit(editor, QStyledItemDelegate.EndEditHint.RevertModelCache)

    def setEditorData(self, editor, index):
        current = index.data(Qt.ItemDataRole.EditRole)
        if isinstance(editor, _CategoryPillButton):
            pos = editor.findText(str(current)) if current else -1
            editor.setCurrentIndex(pos if pos >= 0 else 0)

    def setModelData(self, editor, model, index):
        if isinstance(editor, _CategoryPillButton):
            model.setData(index, editor.currentText(), Qt.ItemDataRole.EditRole)

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)


class _CategoryPillButton(QWidget):
    """Кнопка-пилюля для выбора категории.

    Внешне: нейтральная кнопка «Все категории ▾» → при выборе категории
    превращается в цветной овал с её цветом из CATEGORY_COLORS.

    Реализует подмножество API QComboBox, чтобы существующий код работал
    без изменений: currentText(), blockSignals(), clear(), addItem(),
    addItems(), findText(), setCurrentIndex().
    """

    currentTextChanged = pyqtSignal()
    userPicked = pyqtSignal()          # категория выбрана ЖИВЫМ кликом в popup
    popupDismissed = pyqtSignal()      # popup закрыт без выбора (клик мимо)
    editCategoriesRequested = pyqtSignal()

    _NEUTRAL = "Все категории"

    def __init__(self, neutral_label: str = "Все категории", parent=None, *,
                 auto_open: bool = False, show_edit_button: bool = False):
        super().__init__(parent)
        self._NEUTRAL = neutral_label
        self._items: list[str] = []
        self._current_idx: int = 0
        self._show_edit_button = show_edit_button
        self._picked_in_popup = False
        # auto_open=True — это редактор ячейки таблицы (_CategoryDelegate),
        # а не поле формы. QSS border-radius (даже заведомо огромный) не
        # авто-клэмпится Qt до капсулы, как в CSS браузеров — реальную
        # форму получить можно только ручной отрисовкой, той же, что и в
        # _CategoryDelegate.paint(). Поэтому в компакт-режиме кнопка
        # становится полностью прозрачной (только кликабельна), а овал,
        # текст и шеврон рисует paintEvent САМОГО _CategoryPillButton.
        self._compact = auto_open
        self._pill_bg = QColor("#FFFFFF")
        self._pill_bd = QColor("#D5DCE4")
        self._pill_tx = QColor("#1F2937")
        self._display_text = ""

        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        self._btn = QPushButton()
        self._btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._btn.clicked.connect(self._open_menu)
        if self._compact:
            # Пустая QPushButton даже без текста/паддингов сохраняет
            # нативный минимальный размер стиля (Windows/Fusion) — через
            # layout он раздувает ВЕСЬ _CategoryPillButton выше высоты
            # строки, которую задаёт editor.setGeometry(option.rect).
            # Явный minimumSize(0,0) отменяет это (в отличие от
            # minimumSizeHint(), explicit minimumSize побеждает в layout).
            self._btn.setMinimumSize(0, 0)
        lay.addWidget(self._btn)

        self._update_display()
        if auto_open:
            # Редактор ячейки таблицы: popup открывается сразу, без
            # дополнительного клика по самой кнопке (см. _CategoryDelegate).
            QTimer.singleShot(0, self._open_menu)

    def paintEvent(self, event):
        if not self._compact:
            return super().paintEvent(event)
        # Формулы геометрии/шрифтов ДОЛЖНЫ побайтово совпадать с
        # _CategoryDelegate.paint() (пп. 4, 5, 7) — иначе при входе в
        # редактирование пилюля заметно меняет размер/форму.
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        rect = self.rect()
        v = max(3, (rect.height() - 20) // 2)
        pill = rect.adjusted(6, v, -6, -v)
        pill_f = QRectF(pill).adjusted(0.5, 0.5, -0.5, -0.5)
        if pill_f.width() <= 0 or pill_f.height() <= 0:
            return
        radius = pill_f.height() / 2.0
        p.setPen(QPen(self._pill_bd, 1))
        p.setBrush(self._pill_bg)
        p.drawRoundedRect(pill_f, radius, radius)

        arrow_w = _CategoryDelegate._ARROW_W
        text_rect = pill.adjusted(8, 0, -arrow_w - 4, 0)
        p.setFont(self.font())          # шрифт таблицы, как option.font у делегата
        fm = QFontMetrics(self.font())
        elided = fm.elidedText(self._display_text, Qt.TextElideMode.ElideRight, text_rect.width())
        p.setPen(self._pill_tx)
        p.drawText(text_rect, int(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft), elided)

        arrow_rect = QRect(pill.right() - arrow_w, pill.top(), arrow_w, pill.height())
        p.setFont(_CellDelegate._icon_font())
        p.drawText(arrow_rect, Qt.AlignmentFlag.AlignCenter, chr(0xE5CF))  # expand_more

    # ---- QComboBox-compatible API ---------------------------------------- #

    def currentText(self) -> str:
        if 0 <= self._current_idx < len(self._items):
            return self._items[self._current_idx]
        return ""

    def clear(self):
        self._items.clear()
        self._current_idx = 0
        self._update_display()

    def addItem(self, text: str):
        self._items.append(text)

    def addItems(self, texts):
        self._items.extend(list(texts))

    def findText(self, text: str) -> int:
        try:
            return self._items.index(text)
        except ValueError:
            return -1

    def setCurrentIndex(self, idx: int):
        if self._items:
            idx = max(0, min(idx, len(self._items) - 1))
        else:
            idx = 0
        self._current_idx = idx
        self._update_display()
        self.currentTextChanged.emit()

    # ---- Display --------------------------------------------------------- #

    def _chip_style(self, color: QColor) -> str:
        h, s, l, _ = color.getHslF()
        if h < 0:
            h, s = 0.0, 0.0
        bg = QColor.fromHslF(h, min(s * 0.65, 1.0), 0.91)
        bd = QColor.fromHslF(h, min(s * 1.00, 1.0), 0.52)
        tx = QColor.fromHslF(h, min(s * 1.20, 1.0), 0.18)
        bg_h = bg.darker(108)
        return (
            f"QPushButton {{"
            f"  background: {bg.name()};"
            f"  border: 1px solid {bd.name()};"
            f"  border-radius: 11px;"
            f"  color: {tx.name()};"
            f"  padding: 4px 14px; font-size: 12px; font-weight: 500;"
            f"  min-height: 22px;"
            f"}}"
            f"QPushButton:hover {{ background: {bg_h.name()}; }}"
        )

    def _neutral_style(self) -> str:
        return (
            "QPushButton {"
            "  background: #FFFFFF; border: 1px solid #D5DCE4; border-radius: 6px;"
            "  color: #1F2937; padding: 6px 10px; font-size: 13px; min-height: 22px;"
            "}"
            "QPushButton:hover { background: #F9FAFB; }"
        )

    def _update_display(self):
        text = self.currentText()
        is_neutral = not text or text == self._NEUTRAL
        color = None if is_neutral else CATEGORY_COLORS.get(text)

        if self._compact:
            # Кнопка полностью прозрачна — овал/текст/шеврон рисует
            # paintEvent (см. __init__): только так форма гарантированно
            # совпадает с _CategoryDelegate.paint() при любой высоте строки.
            self._display_text = self._NEUTRAL if is_neutral else text
            if color is not None:
                h, s, l, _ = color.getHslF()
                if h < 0:
                    h, s = 0.0, 0.0
                self._pill_bg = QColor.fromHslF(h, min(s * 0.65, 1.0), 0.91)
                self._pill_bd = QColor.fromHslF(h, min(s * 1.00, 1.0), 0.52)
                self._pill_tx = QColor.fromHslF(h, min(s * 1.20, 1.0), 0.18)
            elif is_neutral:
                self._pill_bg = QColor("#FFFFFF")
                self._pill_bd = QColor("#D5DCE4")
                self._pill_tx = QColor("#1F2937")
            else:
                self._pill_bg = QColor("#EFF1F5")
                self._pill_bd = QColor("#C4C9D4")
                self._pill_tx = QColor("#374151")
            self._btn.setText("")
            self._btn.setStyleSheet(
                "QPushButton{background:transparent;border:none;}")
            self.update()
            return

        if is_neutral:
            self._btn.setText(f"{self._NEUTRAL}  ▾")
            self._btn.setStyleSheet(self._neutral_style())
            return

        if color:
            self._btn.setStyleSheet(self._chip_style(color))
        else:
            self._btn.setStyleSheet(
                "QPushButton {"
                "  background: #EFF1F5; border: 1px solid #C4C9D4; border-radius: 11px;"
                "  color: #374151; padding: 4px 14px; font-size: 12px; font-weight: 500;"
                "  min-height: 22px;"
                "}"
                "QPushButton:hover { background: #E2E5EC; }"
            )
        self._btn.setText(text)

    def _open_menu(self):
        if not self._items:
            return
        self._picked_in_popup = False
        popup = _SingleCatPopup(self._items, self._current_idx,
                                show_edit_button=self._show_edit_button)
        popup.itemSelected.connect(self._on_popup_selected)
        popup.editCategoriesRequested.connect(self._on_popup_edit_categories)
        popup.hidden.connect(self._on_popup_hidden)
        popup.show_at(self._btn.mapToGlobal(self._btn.rect().bottomLeft() + QPoint(0, 2)))

    def _on_popup_selected(self, idx: int):
        self._picked_in_popup = True
        self.setCurrentIndex(idx)
        self.userPicked.emit()

    def _on_popup_edit_categories(self):
        # _SingleCatPopup сразу закроется следом (self.close() в
        # _on_edit_categories) — _picked_in_popup=True гасит дублирующий
        # popupDismissed из _on_popup_hidden.
        self._picked_in_popup = True
        self.editCategoriesRequested.emit()

    def _on_popup_hidden(self):
        if not self._picked_in_popup:
            self.popupDismissed.emit()


class _PlotDelegate(_CellDelegate):
    """Делегат колонки «Участок»: выпадающий список номеров участков из БД.
    Для строк-мультиопераций рисует заштрихованный овал вместо значения."""

    _BG       = QColor("#FFFFFF")
    _BG_ALT   = QColor("#F0F4F8")
    _BG_HOVER = QColor("#DDE4EE")
    _BG_SEL   = QColor("#C9D8E2")
    _BORDER   = QColor("#E3E8EF")

    def __init__(self, items: list[str], parent=None):
        super().__init__(parent)
        self._items = items

    def paint(self, painter, option, index):
        text = index.data(Qt.ItemDataRole.DisplayRole)

        if text == _MULTI_PLOT_LABEL:
            self._paint_multi_plot(painter, option, text)
            return

        if text == _MULTI_OP_LABEL:
            painter.save()

            rect = option.rect
            is_alt = bool(option.features & QStyleOptionViewItem.ViewItemFeature.Alternate)
            if option.state & QStyle.StateFlag.State_Selected:
                painter.fillRect(rect, self._BG_SEL)
            elif option.state & QStyle.StateFlag.State_MouseOver:
                painter.fillRect(rect, self._BG_HOVER)
            else:
                painter.fillRect(rect, self._BG_ALT if is_alt else self._BG)

            painter.setPen(QPen(self._BORDER, 1))
            painter.drawLine(rect.bottomLeft(), rect.bottomRight())
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)

            v = max(3, (rect.height() - 20) // 2)
            pill   = rect.adjusted(6, v, -6, -v)
            pill_f = QRectF(pill).adjusted(0.5, 0.5, -0.5, -0.5)
            radius_f = pill_f.height() / 2.0

            painter.setPen(Qt.PenStyle.NoPen)
            painter.setBrush(QColor(234, 234, 238))
            painter.drawRoundedRect(pill_f, radius_f, radius_f)

            painter.setBrush(QBrush(QColor(185, 185, 193), Qt.BrushStyle.BDiagPattern))
            painter.drawRoundedRect(pill_f, radius_f, radius_f)

            painter.setPen(QPen(QColor(165, 165, 173), 1))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawRoundedRect(pill_f, radius_f, radius_f)

            painter.restore()
            return

        super().paint(painter, option, index)

    def _paint_multi_plot(self, painter, option, text):
        painter.save()

        rect = option.rect
        selected = bool(option.state & QStyle.StateFlag.State_Selected)
        hovered  = bool(option.state & QStyle.StateFlag.State_MouseOver)

        is_alt = bool(option.features & QStyleOptionViewItem.ViewItemFeature.Alternate)
        if selected:
            painter.fillRect(rect, self._BG_SEL)
        else:
            painter.fillRect(rect, self._BG_ALT if is_alt else self._BG)

        painter.setPen(QPen(self._BORDER, 1))
        painter.drawLine(rect.bottomLeft(), rect.bottomRight())
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        v = max(3, (rect.height() - 20) // 2)
        pill   = rect.adjusted(6, v, -6, -v)
        pill_f = QRectF(pill).adjusted(0.5, 0.5, -0.5, -0.5)
        radius_f = pill_f.height() / 2.0

        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QColor(234, 234, 238))
        painter.drawRoundedRect(pill_f, radius_f, radius_f)

        painter.setPen(QPen(QColor(165, 165, 173), 1))
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawRoundedRect(pill_f, radius_f, radius_f)

        painter.setPen(QColor(110, 110, 118))
        painter.setFont(option.font)
        text_rect = pill.adjusted(8, 0, -4, 0)
        fm = QFontMetrics(option.font)
        elided = fm.elidedText(text, Qt.TextElideMode.ElideRight, text_rect.width())
        painter.drawText(
            text_rect,
            int(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft),
            elided,
        )

        painter.restore()

    def createEditor(self, parent, option, index):
        combo = QComboBox(parent)
        combo.addItem("")           # пустой пункт — участок не указан
        combo.addItems(self._items)
        combo.activated.connect(lambda: self.commitData.emit(combo))
        return combo

    def setEditorData(self, editor, index):
        current = index.data(Qt.ItemDataRole.EditRole)
        if isinstance(editor, QComboBox):
            text = str(current).strip() if current else ""
            if text:
                pos = editor.findText(text)
                if pos < 0:
                    # Значения нет в базе — вставляем как первый пункт
                    editor.insertItem(1, text)
                    pos = 1
                editor.setCurrentIndex(pos)
            else:
                editor.setCurrentIndex(0)

    def setModelData(self, editor, model, index):
        if isinstance(editor, QComboBox):
            model.setData(index, editor.currentText(), Qt.ItemDataRole.EditRole)

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)


class _BranchColumnDelegate(_CellDelegate):
    """Делегат колонки «Контрагент»:
    - op-строки с детьми: текст + пилюля «Показать/Свернуть (N)»
    - split-строки: линии дерева + отступ текста
    - op-строки без детей: стандартный _CellDelegate"""

    toggleRequested = pyqtSignal(QModelIndex)

    _BTN_BG     = QColor("#E8F0F5")
    _BTN_BG_H   = QColor("#C9D8E2")
    _BTN_FG     = QColor("#07414F")
    _BTN_BORDER = QColor("#B5C8D5")
    _BTN_H      = 22
    _LINE_COLOR = QColor("#B5C8D5")
    _BG         = QColor("#FFFFFF")
    _BG_ALT     = QColor("#F0F4F8")
    _BG_HOVER   = QColor("#DDE4EE")
    _BG_SEL     = QColor("#C9D8E2")
    _BORDER     = QColor("#E3E8EF")
    _TXT        = QColor("#1F2937")
    _TXT_SEL    = QColor("#07414F")
    _TRUNK_X    = 12
    _SPLIT_PAD  = 24

    def __init__(self, view):
        super().__init__(view)
        self._view = view
        self._hover_btn_idx = QModelIndex()
        view.viewport().installEventFilter(self)

    @staticmethod
    def _btn_font() -> QFont:
        f = QFont()
        f.setPixelSize(11)
        f.setBold(True)
        return f

    def _btn_rect(self, cell_rect: QRect, n: int) -> QRect:
        f_btn = self._btn_font()
        btn_w = QFontMetrics(f_btn).horizontalAdvance(f"Свернуть ({n})") + 20
        y = cell_rect.top() + (cell_rect.height() - self._BTN_H) // 2
        x = cell_rect.right() - btn_w - 6
        return QRect(x, y, btn_w, self._BTN_H)

    def _update_btn_hover(self, pos):
        new_hover = QModelIndex()
        if pos is not None:
            idx = self._view.indexAt(pos)
            if idx.isValid():
                node = idx.internalPointer()
                if node is not None and node.kind == "op" and node.children:
                    rect = self._view.visualRect(idx)
                    if self._btn_rect(rect, len(node.children)).contains(pos):
                        new_hover = idx
        if new_hover != self._hover_btn_idx:
            old = self._hover_btn_idx
            self._hover_btn_idx = new_hover
            if old.isValid():
                self._view.update(old)
            if new_hover.isValid():
                self._view.update(new_hover)
        if new_hover.isValid():
            self._view.viewport().setCursor(Qt.CursorShape.PointingHandCursor)
        else:
            self._view.viewport().unsetCursor()

    def eventFilter(self, obj, event):
        try:
            viewport = self._view.viewport()
        except RuntimeError:
            # self._view уже удалён — хвост событий при закрытии приложения.
            return False
        if obj is viewport:
            if event.type() == QEvent.Type.MouseMove:
                self._update_btn_hover(event.position().toPoint())
            elif event.type() == QEvent.Type.Leave:
                self._update_btn_hover(None)
        return False

    def paint(self, painter, option, index):
        node = index.internalPointer()
        if node is None:
            super().paint(painter, option, index)
            return

        rect = option.rect
        selected = bool(option.state & QStyle.StateFlag.State_Selected)
        hovered  = bool(option.state & QStyle.StateFlag.State_MouseOver)

        if node.kind == "split":
            painter.save()
            painter.setRenderHint(QPainter.RenderHint.Antialiasing, False)

            is_alt = bool(option.features & QStyleOptionViewItem.ViewItemFeature.Alternate)
            if selected:
                bg = self._BG_SEL
            else:
                bg = self._BG_ALT if is_alt else self._BG
            painter.fillRect(rect, bg)
            painter.setPen(QPen(self._BORDER, 1))
            painter.drawLine(rect.bottomLeft(), rect.bottomRight())

            model_obj = index.model()
            is_last = index.row() == model_obj.rowCount(index.parent()) - 1
            midy = rect.top() + rect.height() // 2
            cx = rect.left() + self._TRUNK_X
            painter.setPen(QPen(self._LINE_COLOR, 1))
            painter.drawLine(cx, rect.top(), cx, midy if is_last else rect.bottom())
            painter.drawLine(cx, midy, cx + 10, midy)

            text = index.data(Qt.ItemDataRole.DisplayRole) or ""
            if text:
                painter.setFont(option.font)
                painter.setPen(self._TXT_SEL if selected else self._TXT)
                txt_rect = rect.adjusted(self._SPLIT_PAD, 0, -6, 0)
                elided = painter.fontMetrics().elidedText(
                    str(text), Qt.TextElideMode.ElideRight, txt_rect.width())
                painter.drawText(
                    txt_rect,
                    int(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft),
                    elided,
                )
            painter.restore()

        elif node.kind == "op" and node.children:
            painter.save()

            # Фон и нижняя граница — строго без антиалиасинга, иначе 1px-линия
            # размазывается и не совпадает с QSS-границей соседних колонок.
            is_alt = bool(option.features & QStyleOptionViewItem.ViewItemFeature.Alternate)
            if selected:
                bg = self._BG_SEL
            else:
                bg = self._BG_ALT if is_alt else self._BG
            painter.fillRect(rect, bg)
            painter.setPen(QPen(self._BORDER, 1))
            painter.drawLine(rect.bottomLeft(), rect.bottomRight())

            n = len(node.children)
            col0 = self._view.model().index(index.row(), 0,
                                            self._view.model().parent(index))
            is_expanded = self._view.isExpanded(col0)
            label = f"Свернуть ({n})" if is_expanded else f"Показать ({n})"
            text = index.data(Qt.ItemDataRole.DisplayRole) or ""
            btn = self._btn_rect(rect, n)

            painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
            btn_bg = self._BTN_BG_H if self._hover_btn_idx == index else self._BTN_BG
            painter.setBrush(btn_bg)
            painter.setPen(QPen(self._BTN_BORDER, 1))
            painter.drawRoundedRect(
                QRectF(btn).adjusted(0.5, 0.5, -0.5, -0.5),
                self._BTN_H / 2.0, self._BTN_H / 2.0,
            )

            f_btn = self._btn_font()
            painter.setPen(self._BTN_FG)
            painter.setFont(f_btn)
            painter.drawText(btn, Qt.AlignmentFlag.AlignCenter, label)

            text_rect = QRect(
                rect.left() + 8, rect.top(),
                btn.left() - rect.left() - 10, rect.height(),
            )
            painter.setPen(self._TXT_SEL if selected else self._TXT)
            painter.setFont(option.font)
            elided = painter.fontMetrics().elidedText(
                str(text), Qt.TextElideMode.ElideRight, text_rect.width())
            painter.drawText(
                text_rect,
                Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
                elided,
            )

            if index.data(MANUAL_ROLE):
                painter.setFont(self._icon_font())
                painter.setPen(
                    QColor(255, 255, 255, 180) if selected else QColor("#9CA3AF")
                )
                painter.drawText(
                    rect.adjusted(0, 0, -4, 0),
                    Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight,
                    self._CHAR,
                )

            painter.restore()

        else:
            # Обычные op-строки рисуем той же процедурой, что и строки с кнопкой:
            # один путь рендеринга на всю колонку — иначе фон/отступы чуть «гуляют».
            painter.save()
            painter.setRenderHint(QPainter.RenderHint.Antialiasing, False)

            is_alt = bool(option.features & QStyleOptionViewItem.ViewItemFeature.Alternate)
            if selected:
                bg = self._BG_SEL
            else:
                bg = self._BG_ALT if is_alt else self._BG
            painter.fillRect(rect, bg)
            painter.setPen(QPen(self._BORDER, 1))
            painter.drawLine(rect.bottomLeft(), rect.bottomRight())

            text = index.data(Qt.ItemDataRole.DisplayRole) or ""
            if text:
                painter.setFont(option.font)
                painter.setPen(self._TXT_SEL if selected else self._TXT)
                txt_rect = rect.adjusted(8, 0, -6, 0)
                elided = painter.fontMetrics().elidedText(
                    str(text), Qt.TextElideMode.ElideRight, txt_rect.width())
                painter.drawText(
                    txt_rect,
                    int(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft),
                    elided,
                )

            if index.data(MANUAL_ROLE):
                painter.setFont(self._icon_font())
                painter.setPen(
                    QColor(255, 255, 255, 180) if selected else QColor("#9CA3AF")
                )
                painter.drawText(
                    rect.adjusted(0, 0, -4, 0),
                    Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight,
                    self._CHAR,
                )
            painter.restore()

    def editorEvent(self, event, model, option, index):
        node = index.internalPointer()
        if (node is not None and node.kind == "op" and node.children
                and event.type() == QEvent.Type.MouseButtonRelease
                and event.button() == Qt.MouseButton.LeftButton):
            btn = self._btn_rect(option.rect, len(node.children))
            if btn.contains(event.position().toPoint()):
                self.toggleRequested.emit(index)
                return True
        return super().editorEvent(event, model, option, index)


# =========================================================================== #
#  Вспомогательные UI-виджеты                                                 #
# =========================================================================== #

# =========================================================================== #
#  Всплывающая палитра цветов                                                 #
# =========================================================================== #

class _ColorPickerPopup(QFrame):
    """Всплывающая панель выбора цвета из предустановленной палитры."""

    colorSelected = pyqtSignal(QColor)

    # Палитра из 16 цветов (2 строки × 8): светлые + тёмные тона
    PALETTE: list[QColor] = [
        QColor(0x1B, 0x2A, 0x4A), QColor(0x3D, 0x2F, 0x7A), QColor(0x15, 0x50, 0xA0), QColor(0x3A, 0x9B, 0xD5),
        QColor(0x90, 0xCC, 0xF0), QColor(0x2A, 0xB8, 0xB0), QColor(0x0A, 0x5C, 0x3A), QColor(0x20, 0xA8, 0x40),
        QColor(0x90, 0xD8, 0x88), QColor(0x5C, 0x8C, 0x20), QColor(0xC8, 0xD8, 0x30), QColor(0x6B, 0x5F, 0xA8),
        QColor(0xA8, 0x55, 0xC8), QColor(0xD8, 0x80, 0xC0), QColor(0xC4, 0x58, 0x88), QColor(0xD0, 0x40, 0x60),
        QColor(0xC4, 0x40, 0x30), QColor(0x7A, 0x1A, 0x3A), QColor(0xC8, 0x68, 0x20), QColor(0xE8, 0x98, 0x20),
        QColor(0xF0, 0xD0, 0x30), QColor(0xC4, 0x7A, 0x5A), QColor(0x8C, 0x50, 0x30), QColor(0x6A, 0x60, 0x58),
    ]

    def __init__(self):
        super().__init__(None, Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint)
        self.setStyleSheet("""
            QFrame {
                background: #FFFFFF;
                border: 1px solid #C9D8E2;
                border-radius: 8px;
            }
        """)
        grid = QGridLayout(self)
        grid.setContentsMargins(8, 8, 8, 8)
        grid.setSpacing(5)
        for i, color in enumerate(self.PALETTE):
            h, s, l, _ = color.getHslF()
            if h < 0:
                h, s = 0.0, 0.0
            pill_bg = QColor.fromHslF(h, min(s * 0.65, 1.0), 0.91)
            pill_bd = QColor.fromHslF(h, min(s * 1.00, 1.0), 0.52)

            btn = QPushButton()
            btn.setFixedSize(22, 22)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: {pill_bg.name()};
                    border: 1.5px solid {pill_bd.name()};
                    border-radius: 11px;
                }}
                QPushButton:hover {{ border: 2.5px solid #07414F; }}
            """)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            btn.clicked.connect(lambda checked, c=color: self._pick(c))
            grid.addWidget(btn, i // 8, i % 8)

    def _pick(self, color: QColor):
        self.colorSelected.emit(color)
        self.close()

    def show_near(self, widget: QWidget):
        pos = widget.mapToGlobal(QPoint(0, widget.height() + 2))
        self.adjustSize()
        self.move(pos)
        self.show()
        self.raise_()


class _PeriodFilterPopup(QFrame):
    """Всплывающая панель фильтра по периоду — иконка в заголовке столбца
    «Дата» (тот же визуальный язык, что и _CatFilterPopup: строка действий
    «Сбросить» + разделитель над содержимым, вместо отдельной кнопки
    «Период» с подписями «С:»/«По:»).

    Поля — та же компактная кнопка-дата с разворотом стрелки, что и
    «Дата начала» в карточке участка (см. inp_since в ui/plots_widget.py) —
    единый визуальный язык дата-пикеров приложения."""

    periodChanged = pyqtSignal(object, object)   # (date_from, date_to) | (None, None)
    _ACTIONS_BTN_H = 22

    def __init__(self):
        super().__init__(None, Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint)
        # Карточку рисует paintEvent на прозрачном окне (не QSS-фон) — та же
        # причина, что и у BaseDialog (ui/dialogs.py) и _CatFilterPopup: без
        # WA_TranslucentBackground реальная форма окна остаётся
        # прямоугольной, и в углах "скруглённого" QSS-фона проступает
        # исходный прямоугольный фон окна.
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        self._hide_time: float = 0.0

        def _mk_date_edit() -> NoJumpDateEdit:
            de = NoJumpDateEdit(calendarPopup=True, displayFormat="dd.MM.yyyy")
            style_date_popup(de)
            de.setFixedWidth(114)
            de.setFixedHeight(29)
            arr_dn = icons.icon_png_path("expand_more", 12, color="#6B7280")
            arr_up = icons.icon_png_path("expand_less", 12, color="#6B7280")
            de.setStyleSheet(
                "QDateEdit{background:#F8F9FA;border:1px solid #D5DCE4;"
                "border-radius:4px;padding:4px 8px;font-size:12px;color:#1F2937;}"
                "QDateEdit:focus{border:1px solid #07414F;}"
                "QDateEdit::drop-down{subcontrol-origin:padding;subcontrol-position:right;"
                "width:18px;border:none;border-left:1px solid #D5DCE4;background:transparent;"
                "border-top-right-radius:4px;border-bottom-right-radius:4px;}"
                "QDateEdit::drop-down:hover{background:#F3F4F6;}"
                f"QDateEdit::down-arrow{{image:url({arr_dn});width:12px;height:12px;}}"
                f'QDateEdit[calOpen="true"]::down-arrow{{image:url({arr_up});}}')
            CalendarArrowFlip(de)
            return de

        # Строка действий над содержимым: «Сбросить» — тот же визуальный
        # приём, что и в _CatFilterPopup (LinkButton + разделитель под ней).
        action_row = QHBoxLayout()
        action_row.setContentsMargins(10, 6, 10, 6)
        action_row.addStretch()
        btn_reset = LinkButton("Сбросить")
        btn_reset.setFixedHeight(self._ACTIONS_BTN_H)
        btn_reset.clicked.connect(self._reset)
        action_row.addWidget(btn_reset)

        divider = QFrame()
        divider.setFixedHeight(1)
        divider.setStyleSheet("background:#E5E9EF; border:none;")

        # Одна строка: [С] — [По], без подписей — так компактнее и достаточно
        # понятно из контекста (иконка фильтра уже в заголовке «Дата»).
        date_row = QHBoxLayout()
        date_row.setContentsMargins(12, 10, 12, 12)
        date_row.setSpacing(8)
        self.inp_from = _mk_date_edit()
        date_row.addWidget(self.inp_from)
        dash = QLabel("—")
        dash.setStyleSheet("font-size:12px;color:#9CA3AF;border:none;background:transparent;")
        date_row.addWidget(dash)
        self.inp_to = _mk_date_edit()
        date_row.addWidget(self.inp_to)

        today = QDate.currentDate()
        self.inp_from.setDate(today)
        self.inp_to.setDate(today)

        outer = QVBoxLayout(self)
        # Отступ >= радиуса скругления (8px в paintEvent) — та же причина,
        # что и в _CatFilterPopup.
        outer.setContentsMargins(3, 3, 3, 3)
        outer.setSpacing(0)
        outer.addLayout(action_row)
        outer.addWidget(divider)
        outer.addLayout(date_row)

        # Сигналы подключаем ПОСЛЕ setDate() выше — иначе первичная
        # установка даты по умолчанию сама выстрелила бы periodChanged.
        self.inp_from.dateChanged.connect(self._on_date_changed)
        self.inp_to.dateChanged.connect(self._on_date_changed)

    def _on_date_changed(self, _new_date=None):
        # "С" не может быть позже "По" — подтягиваем второе поле следом.
        if self.inp_from.date() > self.inp_to.date():
            sender = self.sender()
            other = self.inp_to if sender is self.inp_from else self.inp_from
            other.blockSignals(True)
            other.setDate(self.inp_from.date() if sender is self.inp_from else self.inp_to.date())
            other.blockSignals(False)
        self.periodChanged.emit(self.inp_from.date().toPyDate(), self.inp_to.date().toPyDate())

    def _reset(self):
        # Попап НЕ закрывается — та же логика, что и «Сбросить» в
        # _CatFilterPopup (снимает фильтр, но оставляет панель открытой).
        self.periodChanged.emit(None, None)

    def set_period(self, date_from, date_to):
        """Обновляет поля попапа значениями текущего фильтра (без открытия);
        если фильтр не активен — показывает сегодняшнюю дату."""
        d_from = QDate(date_from.year, date_from.month, date_from.day) if date_from else QDate.currentDate()
        d_to = QDate(date_to.year, date_to.month, date_to.day) if date_to else QDate.currentDate()
        self.inp_from.blockSignals(True)
        self.inp_to.blockSignals(True)
        self.inp_from.setDate(d_from)
        self.inp_to.setDate(d_to)
        self.inp_from.blockSignals(False)
        self.inp_to.blockSignals(False)

    def show_at(self, global_pos: "QPoint"):
        self.adjustSize()
        self.move(global_pos)
        self.show()
        self.raise_()

    def paintEvent(self, a0):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        rect = QRectF(self.rect()).adjusted(0.5, 0.5, -0.5, -0.5)
        p.setPen(QPen(QColor("#C9D8E2")))
        p.setBrush(QColor("#FFFFFF"))
        p.drawRoundedRect(rect, 8, 8)

    def hideEvent(self, event):
        self._hide_time = time.monotonic()
        super().hideEvent(event)

    def was_just_hidden(self) -> bool:
        return time.monotonic() - self._hide_time < 0.2


# =========================================================================== #
#  Панель редактора категорий                                                 #
# =========================================================================== #

class _CatPillRow(QWidget):
    """Строка-пилюля в редакторе категорий: цветной овальный badge с
    редактируемым именем и кнопкой удаления."""

    _PILL_H = 34

    def __init__(self, cat: str, is_protected: bool, parent=None):
        super().__init__(parent)
        self._cat = cat
        self._is_protected = is_protected
        self.setFixedHeight(self._PILL_H)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setMouseTracking(True)

        hl = QHBoxLayout(self)
        hl.setContentsMargins(10, 0, 6, 0)
        hl.setSpacing(6)

        color = CATEGORY_COLORS.get(cat, _DEFAULT_ROW_COLOR)
        h, s, l, _ = color.getHslF()
        if h < 0:
            h, s = 0.0, 0.0
        pill_bg = QColor.fromHslF(h, min(s * 0.65, 1.0), 0.91)
        pill_bd = QColor.fromHslF(h, min(s * 1.00, 1.0), 0.52)

        self._color_btn = QPushButton()
        self._color_btn.setFixedSize(18, 18)
        self._color_btn.setStyleSheet(f"""
            QPushButton {{
                background: {pill_bg.name()};
                border: 1.5px solid {pill_bd.name()};
                border-radius: 9px;
            }}
            QPushButton:hover {{ border: 2px solid #07414F; }}
        """)
        self._color_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._color_btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._color_tip = _TooltipFilter("Изменить цвет", self._color_btn)
        self._color_btn.installEventFilter(self._color_tip)
        hl.addWidget(self._color_btn)

        self._lbl = QLineEdit(cat)
        self._lbl.setMinimumWidth(0)
        self._lbl.setReadOnly(is_protected)
        self._lbl.setStyleSheet("""
            QLineEdit {
                background: transparent; border: none;
                font-size: 12px; font-weight: 500; padding: 1px 2px;
            }
            QLineEdit:focus {
                background: rgba(255,255,255,0.7); border: 1px solid rgba(0,0,0,0.15);
                border-radius: 4px; padding: 1px 6px;
            }
        """)
        hl.addWidget(self._lbl, stretch=1)

        if is_protected:
            action_btn = QLabel(icons.icon_char("lock"))
            action_btn.setFixedSize(24, 24)
            action_btn.setAlignment(Qt.AlignmentFlag.AlignCenter)
            action_btn.setFont(icons.icon_font(16))
            action_btn.setStyleSheet(
                "background: transparent; color: #9CA3AF; border: none;"
            )
            self._lock_tip = _TooltipFilter(
                "Обязательная категория — удаление недоступно", action_btn
            )
            action_btn.installEventFilter(self._lock_tip)
        else:
            action_btn = QPushButton("✕")
            action_btn.setFixedSize(24, 24)
            action_btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            action_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            action_btn.setStyleSheet("""
                QPushButton {
                    background: transparent; border: none;
                    color: rgba(0,0,0,0.25); font-size: 13px; font-weight: 600;
                    border-radius: 12px;
                }
                QPushButton:hover { color: #B91C1C; background: rgba(185,28,28,0.08); }
            """)
            self._del_tip = _TooltipFilter("Удалить категорию", action_btn)
            action_btn.installEventFilter(self._del_tip)
        hl.addWidget(action_btn)

        self._color_btn.clicked.connect(
            lambda: self.window()._open_color_picker(self._cat, self._color_btn)
        )
        if not is_protected:
            self._lbl.editingFinished.connect(
                lambda: self.window()._on_rename(
                    self._cat, self._lbl.text().strip(), self._lbl
                )
            )
        if not is_protected:
            action_btn.clicked.connect(
                lambda: self.window()._on_delete(self._cat)
            )

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)

        color = CATEGORY_COLORS.get(self._cat, _DEFAULT_ROW_COLOR)
        h, s, l, _ = color.getHslF()
        if h < 0:
            h, s = 0.0, 0.0
        pill_bg = QColor.fromHslF(h, min(s * 0.65, 1.0), 0.91)
        pill_bd = QColor.fromHslF(h, min(s * 1.00, 1.0), 0.52)

        rf = QRectF(self.rect()).adjusted(0.5, 0.5, -0.5, -0.5)
        radius_f = rf.height() / 2.0

        if self.underMouse():
            bg = QColor(pill_bd)
            bg.setAlphaF(0.08)
        else:
            bg = QColor(pill_bg)

        p.setPen(QPen(pill_bd, 1.0))
        p.setBrush(bg)
        p.drawRoundedRect(rf, radius_f, radius_f)

        p.end()

    def enterEvent(self, event):
        self.update()
        super().enterEvent(event)

    def leaveEvent(self, event):
        self.update()
        super().leaveEvent(event)

    def mouseDoubleClickEvent(self, event):
        if not self._is_protected:
            self._lbl.setFocus()
            self._lbl.selectAll()
        super().mouseDoubleClickEvent(event)


# =========================================================================== #
#  Каркас кастомных диалогов вкладки — без нативного чрома ОС                 #
# =========================================================================== #

# _FramelessDialog, _exec_dialog, _AlertDialog → ui.dialogs
# (BaseDialog / exec_dialog / AlertDialog); импортируются в шапке модуля.


class CategoryEditorPanel(_FramelessDialog):
    """Диалог редактирования списка категорий."""

    categoriesChanged = pyqtSignal(list)
    categoryRenamed   = pyqtSignal(str, str)   # (old_name, new_name)

    def _panel_style(self) -> str:
        return self.base_qss() + """
        QWidget#scrollContents { background: transparent; }
    """

    def __init__(self, categories: list[str], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Редактор категорий")
        self.setModal(False)
        self.setMinimumSize(380, 500)
        self._categories = list(categories)
        self._active_color_cat: str | None = None
        self._color_popup = _ColorPickerPopup()
        self._color_popup.colorSelected.connect(self._on_color_selected)
        self._setup_ui()
        self.setStyleSheet(self._panel_style())

    def set_categories(self, cats: list[str]):
        self._categories = list(cats)
        self._rebuild_list()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        layout.addLayout(self.make_header("Редактор категорий", closable=True))

        self._scroll_contents = QWidget(objectName="scrollContents")
        self._list_layout = QVBoxLayout(self._scroll_contents)
        self._list_layout.setContentsMargins(0, 4, 14, 4)
        self._list_layout.setSpacing(4)

        scroll = QScrollArea()
        scroll.setWidget(self._scroll_contents)
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        # Win11: нативный стиль рисует overlay-скроллбар поверх контента и
        # игнорирует QSS (та же ловушка, что и в разделе "Участки" —
        # ui/plots_widget.py) — принудительно переводим на Fusion.
        # Стиль хранится на self, иначе GC его соберёт и скроллбар
        # вернётся к overlay.
        self._cat_scroll_style = QStyleFactory.create("Fusion")
        if self._cat_scroll_style is not None:
            scroll.setStyle(self._cat_scroll_style)
            scroll.verticalScrollBar().setStyle(self._cat_scroll_style)
        # transparent-трек внутри QScrollArea ненадёжен (просвечивает базовый
        # фон #F0F3F9 серой "рельсой") — красим трек сплошным цветом фона
        # панели. Стиль на самом QScrollBar, а не на контейнере.
        scroll.verticalScrollBar().setStyleSheet(scrollbar_qss(track=C.BG_SURFACE))
        layout.addWidget(scroll, stretch=1)

        add_row = QHBoxLayout()
        add_row.setSpacing(6)
        self._new_input = QLineEdit(objectName="newCatInput")
        self._new_input.setPlaceholderText("Новая категория...")
        self._new_input.returnPressed.connect(self._on_add)
        add_btn = PrimaryButton("Добавить")
        add_btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        add_btn.clicked.connect(self._on_add)
        add_row.addWidget(self._new_input, stretch=1)
        add_row.addWidget(add_btn)
        layout.addLayout(add_row)

        self._rebuild_list()

    def _rebuild_list(self):
        while self._list_layout.count():
            item = self._list_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        for cat in self._categories:
            is_protected = cat in PROTECTED_CATEGORIES
            row = _CatPillRow(cat, is_protected, self)
            self._list_layout.addWidget(row)
        self._list_layout.addStretch()

    def _on_rename(self, old_name: str, new_name: str, editor: "QLineEdit"):
        if old_name in PROTECTED_CATEGORIES:
            editor.setText(old_name)
            return
        if not new_name or new_name == old_name:
            editor.setText(old_name)
            return
        if new_name in self._categories:
            editor.setText(old_name)
            confirmed = _ConfirmDialog.confirm(
                self,
                "Объединить категории?",
                f"Категория «{new_name}» уже существует.\n\n"
                f"Объединить «{old_name}» с «{new_name}»?\n"
                f"Все строки с категорией «{old_name}» получат категорию «{new_name}»,\n"
                f"а «{old_name}» будет удалена из списка.",
                confirm_text="Объединить", cancel_text="Отмена",
            )
            if confirmed:
                self._do_merge(old_name, new_name)
            return
        rename_user_category(old_name, new_name)
        self._categories[self._categories.index(old_name)] = new_name
        self._rebuild_list()
        self.categoryRenamed.emit(old_name, new_name)

    def _do_merge(self, old_name: str, new_name: str):
        """Удаляет old_name из списка; все строки df с этой категорией
        получат new_name через сигнал categoryRenamed."""
        delete_user_category(old_name)
        if old_name in self._categories:
            self._categories.remove(old_name)
        self._rebuild_list()
        self.categoryRenamed.emit(old_name, new_name)

    def _open_color_picker(self, cat: str, btn: QPushButton):
        self._active_color_cat = cat
        self._color_popup.show_near(btn)

    def _on_color_selected(self, color: QColor):
        if not self._active_color_cat:
            return
        import ui.categorization as _cat_mod
        _cat_mod.CATEGORY_COLORS[self._active_color_cat] = color
        save_user_category_color(self._active_color_cat, color)
        self._active_color_cat = None
        self._rebuild_list()
        self.categoriesChanged.emit(list(self._categories))

    def _on_add(self):
        name = self._new_input.text().strip()
        if not name or name in self._categories:
            return
        self._categories.append(name)
        import ui.categorization as _cat_mod
        if name not in _cat_mod.CATEGORY_COLORS:
            # Иначе категория без цвета рисуется обычным текстом без
            # пилюли (см. _CategoryDelegate.paint: color is None) — берём
            # цвет по кругу из готовой палитры, чтобы у каждой новой
            # категории сразу была своя пилюля.
            palette = _ColorPickerPopup.PALETTE
            color = palette[(len(self._categories) - 1) % len(palette)]
            _cat_mod.CATEGORY_COLORS[name] = color
            save_user_category_color(name, color)
        self._new_input.clear()
        self._rebuild_list()
        self._persist_and_emit()

    def _on_delete(self, cat: str):
        if cat in self._categories:
            self._categories.remove(cat)
            delete_user_category(cat)
            self._rebuild_list()
            self.categoriesChanged.emit(list(self._categories))
            # Ячейки, у которых стояла удалённая категория, не должны
            # остаться «осиротевшими» (пустой текст без пилюли) — переносим
            # их в служебную «Без категории» тем же путём, что и слияние
            # категорий при переименовании (см. _do_merge/categoryRenamed).
            self.categoryRenamed.emit(cat, NO_CATEGORY_LABEL)

    def _persist_and_emit(self):
        save_user_categories(self._categories)
        self.categoriesChanged.emit(list(self._categories))


# =========================================================================== #
#  Диалоги (без изменений)                                                    #
# =========================================================================== #

class LoadSettingsDialog(_FramelessDialog):
    """Диалог настроек перед загрузкой файла выписки."""

    # Пилюли (полностью скруглённые) — тот же визуальный язык, что и выбор
    # формата в «Сохранить как файл» (ui.plots_widget._ExportFormatDialog):
    # радиус/паддинг/шрифт совпадают, различие — «выбранное» состояние тут
    # стойкое (залито брендом), а не мгновенное действие, как там.
    #
    # Стиль ставится ПРЯМО на кнопку (setStyleSheet), а не через objectName +
    # каскад от диалога: у objectName-селекторов (#fmtActive/#fmtInactive)
    # на этом диалоге почему-то не срабатывал каскад от родителя (кнопки
    # оставались нестилизованными, как на скриншоте — обе выглядели
    # одинаково блёкло) — прямая установка стиля надёжна независимо от
    # причины.
    _FMT_ACTIVE_QSS = f"""
        QPushButton {{
            background: {C.BRAND}; color: #FFFFFF; border: none;
            border-radius: 16px; padding: 7px 16px; font-size: {FS.BODY}px;
            font-weight: 600;
        }}
        QPushButton:hover {{ background: {C.BRAND_HOVER}; }}
    """
    # СберБизнес — фирменный зелёный вместо общего бренда приложения,
    # узнаваемо как «это про Сбер», в отличие от «Мой Садовод».
    _FMT_ACTIVE_SBER_QSS = f"""
        QPushButton {{
            background: #148F2B; color: #FFFFFF; border: none;
            border-radius: 16px; padding: 7px 16px; font-size: {FS.BODY}px;
            font-weight: 600;
        }}
        QPushButton:hover {{ background: #106F22; }}
    """
    _FMT_INACTIVE_QSS = f"""
        QPushButton {{
            background: {C.BG_SURFACE}; color: {C.TEXT_MUTED};
            border: 1px solid {C.BORDER};
            border-radius: 16px; padding: 7px 16px; font-size: {FS.BODY}px;
            font-weight: 600;
        }}
        QPushButton:hover {{ background: {C.BG_HOVER}; color: {C.TEXT_BODY}; }}
    """

    def __init__(self, parent=None, has_existing_data: bool = False):
        super().__init__(parent)
        self.setWindowTitle("Загрузка детализации")
        self.setModal(True)
        self.setFixedWidth(400)
        self._fmt = "sber"
        self._has_existing = has_existing_data
        self._setup_ui()
        self.setStyleSheet(self.base_qss())

    def _setup_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(24, 24, 24, 20)
        lay.setSpacing(14)

        lay.addLayout(self.make_header("Загрузка детализации"))

        div0 = QFrame(objectName="divider")
        div0.setFixedHeight(1)
        lay.addWidget(div0)

        lay.addWidget(QLabel("ФОРМАТ ФАЙЛА", objectName="sectionLabel"))

        fmt_row = QHBoxLayout()
        fmt_row.setSpacing(8)
        self._btn_sber = QPushButton("СберБизнес")
        self._btn_snt  = QPushButton("Мой Садовод")
        self._btn_sber.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._btn_snt .setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._btn_sber.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_snt .setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_sber.clicked.connect(lambda: self._set_fmt("sber"))
        self._btn_snt .clicked.connect(lambda: self._set_fmt("snt"))
        fmt_row.addWidget(self._btn_sber)
        fmt_row.addWidget(self._btn_snt)
        fmt_row.addStretch()
        lay.addLayout(fmt_row)
        self._btn_sber.setStyleSheet(self._FMT_ACTIVE_SBER_QSS)
        self._btn_snt.setStyleSheet(self._FMT_INACTIVE_QSS)

        self._fmt_hint = QLabel()
        self._fmt_hint.setWordWrap(True)
        self._fmt_hint.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
        self._fmt_hint.setStyleSheet("color:#9CA3AF; font-size:11px; background:transparent;")
        # Высота зарезервирована под ХУДШИЙ случай (2 строки — подсказка
        # «Мой Садовод» длиннее и переносится, а «СберБизнес» умещается в
        # одну) — иначе переключение формата меняет высоту всего диалога, и
        # он «дёргается»/скачет при клике (BaseDialog пересчитывает маску
        # скругления на resize, отсюда и рывок).
        _hint_font = QFont(self.font())
        _hint_font.setPixelSize(11)
        _hint_fm = QFontMetrics(_hint_font)
        self._fmt_hint.setFixedHeight(_hint_fm.lineSpacing() * 2 + 2)
        lay.addWidget(self._fmt_hint)
        self._update_hint()

        div1 = QFrame(objectName="divider")
        div1.setFixedHeight(1)
        lay.addWidget(div1)

        lay.addWidget(QLabel("АВТОМАТИЧЕСКОЕ РАСПРЕДЕЛЕНИЕ", objectName="sectionLabel"))

        # Тумблеры toggle_on/toggle_off вместо чекбоксов — тот же виджет,
        # что и «Показывать переплату» в PlotsWidget (ui.plots_widget).
        def _make_toggle(text: str) -> QPushButton:
            btn = QPushButton(f" {text}")
            btn.setCheckable(True)
            btn.setChecked(True)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setIconSize(QSize(20, 20))
            btn.setStyleSheet(
                "QPushButton{background:transparent;border:none;border-radius:6px;"
                "padding:4px 8px;font-size:13px;color:#374151;text-align:left;}"
                "QPushButton:hover{background:#F3F4F6;}")
            return btn

        def _refresh_toggle_icon(btn: QPushButton):
            checked = btn.isChecked()
            cp = 0xE9F6 if checked else 0xE9F5  # toggle_on / toggle_off
            btn.setIcon(icons.get_icon(cp, 20, fill=1 if checked else 0,
                                       color="#07414F" if checked else "#9CA3AF"))

        self.chk_cat  = _make_toggle("Категория")
        self.chk_plot = _make_toggle("Участок")
        auto_row = QHBoxLayout()
        auto_row.setContentsMargins(0, 0, 0, 0)
        auto_row.setSpacing(8)
        for btn in (self.chk_cat, self.chk_plot):
            _refresh_toggle_icon(btn)
            btn.toggled.connect(lambda _checked, b=btn: _refresh_toggle_icon(b))
            auto_row.addWidget(btn)
        auto_row.addStretch()
        lay.addLayout(auto_row)

        div2 = QFrame(objectName="divider")
        div2.setFixedHeight(1)
        lay.addWidget(div2)

        if self._has_existing:
            lay.addWidget(QLabel("РЕЖИМ ЗАГРУЗКИ", objectName="sectionLabel"))
            self.chk_merge = _make_toggle("Добавить к существующим данным")
            _refresh_toggle_icon(self.chk_merge)
            self.chk_merge.toggled.connect(
                lambda _checked, b=self.chk_merge: _refresh_toggle_icon(b))
            lay.addWidget(self.chk_merge)

            merge_hint = QLabel(
                "Новые операции будут добавлены к уже загруженным. Операции, "
                "которые уже есть в списке, не задваиваются — существующая "
                "строка подсвечивается жёлтым, правой кнопкой мыши по ней "
                "можно восстановить исходные данные повтора или оставить "
                "как есть.")
            merge_hint.setWordWrap(True)
            merge_hint.setStyleSheet(
                "color:#9CA3AF; font-size:11px; background:transparent;")
            # Высота задаётся явно по факту переноса текста на ширину
            # содержимого диалога (400 - поля по 24px) — авто-высота
            # WordWrap-лейбла в этом frameless-диалоге считается ДО того,
            # как layout узнаёт финальную ширину, и текст обрезается снизу
            # (см. скриншот с «...оставить как» без конца фразы).
            _hint_font2 = QFont(self.font())
            _hint_font2.setPixelSize(11)
            _fm2 = QFontMetrics(_hint_font2)
            _wrap_w = 400 - 24 - 24  # setFixedWidth(400) - поля 24px слева/справа
            _rect = _fm2.boundingRect(
                0, 0, _wrap_w, 2000, Qt.TextFlag.TextWordWrap, merge_hint.text())
            merge_hint.setFixedHeight(_rect.height() + 4)
            lay.addWidget(merge_hint)

            div3 = QFrame(objectName="divider")
            div3.setFixedHeight(1)
            lay.addWidget(div3)
        else:
            self.chk_merge = None

        btn_cancel = SecondaryButton("Отмена")
        btn_ok     = PrimaryButton("Выбрать файл")
        btn_cancel.clicked.connect(self.reject)
        btn_ok.clicked.connect(self.accept)
        lay.addLayout(self.make_button_row(btn_cancel, btn_ok))

    def _set_fmt(self, fmt: str):
        self._fmt = fmt
        self._btn_sber.setStyleSheet(
            self._FMT_ACTIVE_SBER_QSS if fmt == "sber" else self._FMT_INACTIVE_QSS)
        self._btn_snt.setStyleSheet(
            self._FMT_ACTIVE_QSS if fmt == "snt" else self._FMT_INACTIVE_QSS)
        self._update_hint()

    def _update_hint(self):
        if self._fmt == "sber":
            self._fmt_hint.setText(
                "Стандартная выгрузка операций из СберБизнес (.xlsx)")
        else:
            self._fmt_hint.setText(
                "Файл в формате программы Мой Садовод — столбцы уже приведены к нужному виду")

    @property
    def fmt(self) -> str:
        return self._fmt

    @property
    def auto_cat(self) -> bool:
        return self.chk_cat.isChecked()

    @property
    def auto_plot(self) -> bool:
        return self.chk_plot.isChecked()

    @property
    def merge_mode(self) -> bool:
        return self.chk_merge.isChecked() if self.chk_merge else False


# =========================================================================== #
#  Виджет строки разбивки (для Add/Edit диалогов)                             #
# =========================================================================== #

class _PlotComboBox(QComboBox):
    """Выпадающий список участков с фильтрацией по вводу.

    При вводе текста список фильтруется в реальном времени (MatchContains).
    Допускается только выбор существующего значения — при потере фокуса
    или нажатии Enter невалидное значение сбрасывается.
    """

    _STYLE = """
        QComboBox {
            background: #F8F9FA; border: 1px solid #D1D5DB;
            border-radius: 5px; color: #374151; padding: 7px 10px;
            padding-right: 28px; font-size: 13px;
        }
        QComboBox:focus { border: 1px solid #07414F; }
        QComboBox::drop-down {
            subcontrol-origin: padding; subcontrol-position: center right;
            width: 28px; border: none; border-left: 1px solid #E5E7EB;
            border-radius: 0 4px 4px 0;
        }
        QComboBox::down-arrow {
            image: none; width: 10px; height: 10px;
        }
        QComboBox QAbstractItemView {
            background: #FFFFFF; border: 1px solid #D1D5DB;
            border-radius: 6px; padding: 4px;
            selection-background-color: #E8F0F5; selection-color: #07414F;
        }
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setEditable(True)
        self.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.setStyleSheet(self._STYLE)
        self.addItem("")
        self._valid_values: set[str] = {""}
        self._fill_items()

        completer = self.completer()
        completer.setFilterMode(Qt.MatchFlag.MatchContains)
        completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)

        self.lineEdit().editingFinished.connect(self._validate)

    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self)
        painter.save()
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        # Стрелка вниз
        arrow_x = self.width() - 22
        painter.setPen(QColor("#6B7280"))
        f = QFont("Material Symbols Rounded")
        f.setPixelSize(18)
        painter.setFont(f)
        arrow_rect = QRectF(arrow_x, 0, 22, self.height())
        painter.drawText(arrow_rect, Qt.AlignmentFlag.AlignCenter, chr(0xE5C5))
        painter.restore()

    def _fill_items(self):
        try:
            if os.path.exists(_PLOTS_FILE):
                with open(_PLOTS_FILE, "r", encoding="utf-8") as f:
                    plots = json.load(f)
                nums = sorted(
                    set(str(p.get("num", "")) for p in plots if p.get("num")),
                    key=lambda s: (0, int(s), s) if s.isdigit() else (1, 0, s),
                )
                for num in nums:
                    self.addItem(num)
                self._valid_values = {"", *nums}
        except Exception:
            pass

    def _validate(self):
        text = self.currentText().strip()
        if text and text not in self._valid_values:
            self.setCurrentIndex(0)


class _SplitRowWidget(QWidget):
    """Одна строка разбивки: Сумма / Категория / Участок / кнопка удаления."""

    deleteRequested = pyqtSignal(object)  # self
    sumChanged = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 2, 0, 2)
        lay.setSpacing(6)

        self.inp_summa = QLineEdit()
        self.inp_summa.setPlaceholderText("Сумма")
        self.inp_summa.setMaximumWidth(120)
        self.inp_summa.textChanged.connect(lambda: self.sumChanged.emit())
        lay.addWidget(self.inp_summa)

        self.combo_cat = _CategoryPillButton(neutral_label="")
        for cat in ALL_CATEGORIES:
            self.combo_cat.addItem(cat)
        if ALL_CATEGORIES:
            self.combo_cat.setCurrentIndex(0)
        lay.addWidget(self.combo_cat, stretch=1)

        self.combo_plot = _PlotComboBox()
        self.combo_plot.setMaximumWidth(100)
        lay.addWidget(self.combo_plot)

        btn_del = QPushButton("✕")
        btn_del.setFixedSize(26, 26)
        btn_del.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_del.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        btn_del.setStyleSheet(
            "QPushButton { background: transparent; border: none;"
            " color: #9CA3AF; font-size: 14px; font-weight: 600; border-radius: 4px; }"
            "QPushButton:hover { background: #FEE2E2; color: #B91C1C; }"
        )
        btn_del.clicked.connect(lambda: self.deleteRequested.emit(self))
        lay.addWidget(btn_del)

    def get_data(self) -> dict:
        raw = (self.inp_summa.text().strip()
               .replace(",", ".").replace("−", "-").replace("–", "-").replace(" ", ""))
        summa = 0.0
        try:
            summa = float(raw)
        except ValueError:
            pass
        return {
            "Сумма": summa,
            "Категория": self.combo_cat.currentText(),
            "Участок": self.combo_plot.currentText().strip(),
        }

    def set_data(self, data: dict):
        num = _to_num(data.get("Сумма"))
        if num is not None:
            self.inp_summa.setText(_num_edit_str(num))
        cat = str(data.get("Категория") or "")
        idx = self.combo_cat.findText(cat)
        if idx >= 0:
            self.combo_cat.setCurrentIndex(idx)
        self.combo_plot.setCurrentText(str(data.get("Участок") or ""))


# Кнопки «Разделить операцию» / «Добавить строку» — общий стиль
# диалогов добавления/редактирования операции.
_SPLIT_BTN_QSS = f"""
    QPushButton#btnSplit, QPushButton#btnAddSplit {{
        background: transparent; color: {C.BRAND}; border: 1px solid #B5C8D5;
        border-radius: 6px; padding: 5px 14px; font-size: {FS.SMALL}px;
        font-weight: 600; text-align: left;
    }}
    QPushButton#btnSplit:hover, QPushButton#btnAddSplit:hover {{
        background: {C.BRAND_FAINT};
    }}
"""


class AddRowDialog(_FramelessDialog):
    """Диалог ручного добавления операции в таблицу Детализации."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Добавить операцию")
        self.setMinimumWidth(500)
        self.setModal(True)
        self._setup_ui()
        self._apply_styles()

    def _setup_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(24, 22, 24, 20)
        lay.setSpacing(14)

        lay.addLayout(self.make_header("Добавить операцию"))

        form = QFormLayout()
        form.setSpacing(10)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        self.date_edit = QDateEdit(QDate.currentDate(), calendarPopup=True)
        style_date_popup(self.date_edit)
        self.date_edit.setDisplayFormat("dd.MM.yyyy")
        self.date_edit.setObjectName("datePicker")
        form.addRow("Дата:", self.date_edit)

        self.inp_summa = QLineEdit()
        self.inp_summa.setPlaceholderText("например: 1500 (поступление) или -500 (списание)")
        self.inp_summa.textChanged.connect(self._update_save_state)
        form.addRow("Сумма, ₽:", self.inp_summa)

        self.inp_cont = QLineEdit()
        self.inp_cont.setPlaceholderText("Организация или ФИО")
        form.addRow("Контрагент:", self.inp_cont)

        self.inp_nazn = QLineEdit()
        self.inp_nazn.setPlaceholderText("Назначение платежа")
        form.addRow("Назначение:", self.inp_nazn)

        self.combo_cat = _CategoryPillButton(neutral_label="")
        for cat in ALL_CATEGORIES:
            self.combo_cat.addItem(cat)
        if ALL_CATEGORIES:
            self.combo_cat.setCurrentIndex(0)
        form.addRow("Категория:", self.combo_cat)

        self.combo_plot = _PlotComboBox()
        form.addRow("Участок:", self.combo_plot)

        lay.addLayout(form)

        # ---- секция разбивки операции ----
        self._split_rows: list[_SplitRowWidget] = []

        self._btn_split = QPushButton("  Разделить операцию")
        self._btn_split.setObjectName("btnSplit")
        self._btn_split.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_split.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._btn_split.clicked.connect(self._on_split_toggle)
        lay.addWidget(self._btn_split)

        self._split_section = QWidget()
        self._split_section.setVisible(False)
        split_lay = QVBoxLayout(self._split_section)
        split_lay.setContentsMargins(0, 0, 0, 0)
        split_lay.setSpacing(4)

        hdr = QHBoxLayout()
        for lbl in ("Сумма", "Категория", "Участок", ""):
            l = QLabel(lbl)
            l.setStyleSheet("color:#6B7280; font-size:11px; font-weight:600;")
            hdr.addWidget(l, stretch=(2 if lbl == "Категория" else (0 if lbl == "" else 1)))
        split_lay.addLayout(hdr)

        self._split_rows_lay = QVBoxLayout()
        self._split_rows_lay.setSpacing(2)
        split_lay.addLayout(self._split_rows_lay)

        self._btn_add_split = QPushButton(" Добавить строку")
        self._btn_add_split.setIcon(icons.get_icon("add", 14, color=C.BRAND))
        self._btn_add_split.setIconSize(QSize(14, 14))
        self._btn_add_split.setObjectName("btnAddSplit")
        self._btn_add_split.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_add_split.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._btn_add_split.clicked.connect(self._add_split_row)
        split_lay.addWidget(self._btn_add_split)

        lay.addWidget(self._split_section)
        # ----------------------------------

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color:#E5E7EB;background:#E5E7EB;max-height:1px;")
        lay.addWidget(sep)

        self._split_warning = QLabel()
        self._split_warning.setObjectName("splitWarning")
        self._split_warning.setWordWrap(True)
        self._split_warning.setStyleSheet(
            "color:#9CA3AF; background:transparent; font-size:12px;"
        )
        lay.addWidget(self._split_warning)

        btn_cancel = SecondaryButton("Отмена")
        btn_cancel.clicked.connect(self.reject)
        self._btn_save = PrimaryButton("Добавить")
        self._btn_save.clicked.connect(self._on_accept)
        lay.addLayout(self.make_button_row(btn_cancel, self._btn_save))
        self._update_save_state()

    def _on_split_toggle(self):
        if self._split_section.isVisible():
            self._split_section.setVisible(False)
            self._btn_split.setText("  Разделить операцию")
            for row in list(self._split_rows):
                self._split_rows_lay.removeWidget(row)
                row.deleteLater()
            self._split_rows.clear()
        else:
            self._split_section.setVisible(True)
            self._btn_split.setText("  Отменить разделение")
            self._add_split_row()
            self._add_split_row()
        self.adjustSize()
        self._update_save_state()

    def _add_split_row(self):
        row = _SplitRowWidget(self._split_section)
        row.deleteRequested.connect(self._remove_split_row)
        self._split_rows_lay.addWidget(row)
        self._split_rows.append(row)
        self.adjustSize()
        self._update_save_state()

    def _remove_split_row(self, row: "_SplitRowWidget"):
        if row in self._split_rows:
            self._split_rows.remove(row)
            self._split_rows_lay.removeWidget(row)
            row.deleteLater()
            self.adjustSize()
            self._update_save_state()

    def _update_save_state(self):
        if not hasattr(self, '_btn_save') or self._btn_save is None:
            return
        split_ok = True
        if self._split_rows and self._split_section.isVisible():
            raw = (self.inp_summa.text().strip()
                   .replace(",", ".").replace("\u2212", "-").replace("\u2013", "-").replace(" ", ""))
            try:
                main_sum = float(raw) if raw else 0.0
            except ValueError:
                main_sum = 0.0
            split_sum = 0.0
            for row in self._split_rows:
                rd = row.get_data()
                split_sum += rd.get("Сумма", 0.0)
            split_ok = abs(main_sum - split_sum) < 0.01
        self._btn_save.setEnabled(split_ok)
        self._btn_save.setCursor(
            Qt.CursorShape.PointingHandCursor if split_ok else Qt.CursorShape.ArrowCursor
        )
        _text = "Сумма разбитых строк должна равняться сумме операции"
        if split_ok:
            self._split_warning.setStyleSheet(
                "color:#9CA3AF; background:transparent; font-size:12px;"
            )
        else:
            self._split_warning.setStyleSheet(
                "color:#B45309; background:transparent; font-size:12px; font-weight:600;"
            )
        self._split_warning.setText(f"\u2139  {_text}")
        self._split_warning.setVisible(self._split_section.isVisible() and not split_ok)

    def _prefill(self, data: dict, breakdown: list):
        ts = _to_ts(data.get("Дата"))
        if ts is not None:
            self.date_edit.setDate(QDate(ts.year, ts.month, ts.day))
        num = _to_num(data.get("Сумма"))
        if num is not None:
            self.inp_summa.setText(_num_edit_str(num))
        self.inp_cont.setText(str(data.get("Контрагент") or ""))
        self.inp_cont.setCursorPosition(0)
        self.inp_nazn.setText(str(data.get("Назначение") or ""))
        self.inp_nazn.setCursorPosition(0)
        cat = str(data.get("Категория") or "")
        idx = self.combo_cat.findText(cat)
        if idx >= 0:
            self.combo_cat.setCurrentIndex(idx)
        self.combo_plot.setCurrentText(str(data.get("Участок") or ""))
        if breakdown:
            self._on_split_toggle()
            for bd in breakdown:
                self._add_split_row(bd)
            self._update_save_state()

    def _on_accept(self):
        raw = (self.inp_summa.text().strip()
               .replace(",", ".").replace("−", "-").replace("–", "-").replace(" ", ""))
        if not raw:
            _AlertDialog.show_alert(self, "Ошибка", "Укажите сумму операции")
            return
        try:
            float(raw)
        except ValueError:
            _AlertDialog.show_alert(self, "Ошибка", "Некорректный формат суммы")
            return
        self.accept()

    def get_result(self) -> dict:
        raw = (self.inp_summa.text().strip()
               .replace(",", ".").replace("−", "-").replace("–", "-").replace(" ", ""))
        d = self.date_edit.date()
        result = {
            "Дата":        pd.Timestamp(d.year(), d.month(), d.day()),
            "Контрагент":  self.inp_cont.text().strip(),
            "Сумма":       float(raw),
            "Назначение":  self.inp_nazn.text().strip(),
            "Категория":   self.combo_cat.currentText(),
            "Участок":     self.combo_plot.currentText().strip(),
        }
        if self._split_rows:
            contragent = self.inp_cont.text().strip()
            result["_breakdown"] = [
                {"Контрагент": contragent, **row.get_data()}
                for row in self._split_rows
            ]
        return result

    def _apply_styles(self):
        self.setStyleSheet(self.base_qss() + _SPLIT_BTN_QSS)


# =========================================================================== #
#  Делегат кнопки редактирования                                              #
# =========================================================================== #

class _DetailEditDelegate(QStyledItemDelegate):
    """Иконка карандаша в последнем столбце; filled при hover, cursor-рука."""

    _IC_EDIT  = chr(0xE3C9)
    _IC_FONT  = "Material Symbols Rounded"
    _IC_COLOR = QColor("#07414F")

    def __init__(self, view):
        super().__init__(view)
        self._view      = view
        self._hover_idx = QModelIndex()
        self._pointing  = False
        self._fill_tag  = QFont.Tag.fromString("FILL")
        view.viewport().installEventFilter(self)

    def _is_btn(self, index: QModelIndex) -> bool:
        if not index.isValid():
            return False
        node = index.internalPointer()
        m    = index.model()
        cols = m.columns() if m else []
        col  = cols[index.column()] if 0 <= index.column() < len(cols) else ""
        return bool(node and node.kind == "op" and col == _EDIT_COL)

    def paint(self, painter, option, index):
        super().paint(painter, option, index)
        if not self._is_btn(index):
            return
        painter.save()
        f = QFont(self._IC_FONT)
        f.setPixelSize(18)
        f.setVariableAxis(self._fill_tag, 1.0 if self._hover_idx == index else 0.0)
        painter.setFont(f)
        painter.setPen(self._IC_COLOR)
        painter.drawText(option.rect, Qt.AlignmentFlag.AlignCenter, self._IC_EDIT)
        painter.restore()

    def eventFilter(self, obj, event):
        try:
            viewport = self._view.viewport()
        except RuntimeError:
            # self._view уже удалён — хвост событий при закрытии приложения.
            return False
        if obj is viewport:
            if event.type() == QEvent.Type.MouseMove:
                idx    = self._view.indexAt(event.position().toPoint())
                on_btn = self._is_btn(idx)
                if on_btn and not self._pointing:
                    self._pointing = True
                    QApplication.setOverrideCursor(Qt.CursorShape.PointingHandCursor)
                elif not on_btn and self._pointing:
                    self._pointing = False
                    QApplication.restoreOverrideCursor()
                new_hover = idx if on_btn else QModelIndex()
                if new_hover != self._hover_idx:
                    old = self._hover_idx
                    self._hover_idx = new_hover
                    if old.isValid():
                        self._view.viewport().update(self._view.visualRect(old))
                    if new_hover.isValid():
                        self._view.viewport().update(self._view.visualRect(new_hover))
            elif event.type() == QEvent.Type.Leave:
                if self._pointing:
                    self._pointing = False
                    QApplication.restoreOverrideCursor()
                if self._hover_idx.isValid():
                    old = self._hover_idx
                    self._hover_idx = QModelIndex()
                    self._view.viewport().update(self._view.visualRect(old))
        return super().eventFilter(obj, event)


# =========================================================================== #
#  Диалог редактирования операции                                             #
# =========================================================================== #

class EditOperationDialog(_FramelessDialog):
    """Диалог редактирования существующей операции."""

    def __init__(self, data: dict, breakdown: list = None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Редактировать операцию")
        self.setMinimumWidth(500)
        self.setModal(True)
        self._split_rows: list[_SplitRowWidget] = []
        self._setup_ui()
        self._prefill(data, breakdown or [])
        self._apply_styles()

    def _setup_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(24, 22, 24, 20)
        lay.setSpacing(14)

        lay.addLayout(self.make_header("Редактировать операцию"))

        form = QFormLayout()
        form.setSpacing(10)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        self.date_edit = QDateEdit(QDate.currentDate(), calendarPopup=True)
        style_date_popup(self.date_edit)
        self.date_edit.setDisplayFormat("dd.MM.yyyy")
        self.date_edit.setObjectName("datePicker")
        form.addRow("Дата:", self.date_edit)

        self.inp_summa = QLineEdit()
        self.inp_summa.setPlaceholderText("например: 1500 (поступление) или -500 (списание)")
        self.inp_summa.textChanged.connect(self._update_save_state)
        form.addRow("Сумма, ₽:", self.inp_summa)

        self.inp_cont = QLineEdit()
        self.inp_cont.setPlaceholderText("Организация или ФИО")
        form.addRow("Контрагент:", self.inp_cont)

        self.inp_nazn = QLineEdit()
        self.inp_nazn.setPlaceholderText("Назначение платежа")
        form.addRow("Назначение:", self.inp_nazn)

        self.combo_cat = _CategoryPillButton(neutral_label="")
        for cat in ALL_CATEGORIES:
            self.combo_cat.addItem(cat)
        if ALL_CATEGORIES:
            self.combo_cat.setCurrentIndex(0)
        form.addRow("Категория:", self.combo_cat)

        self.combo_plot = _PlotComboBox()
        form.addRow("Участок:", self.combo_plot)

        lay.addLayout(form)

        # ---- секция разбивки операции ----
        self._btn_split = QPushButton("  Разделить операцию")
        self._btn_split.setObjectName("btnSplit")
        self._btn_split.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_split.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._btn_split.clicked.connect(self._on_split_toggle)
        lay.addWidget(self._btn_split)

        self._split_section = QWidget()
        self._split_section.setVisible(False)
        split_lay = QVBoxLayout(self._split_section)
        split_lay.setContentsMargins(0, 0, 0, 0)
        split_lay.setSpacing(4)

        hdr = QHBoxLayout()
        for lbl in ("Сумма", "Категория", "Участок", ""):
            l = QLabel(lbl)
            l.setStyleSheet("color:#6B7280; font-size:11px; font-weight:600;")
            hdr.addWidget(l, stretch=(2 if lbl == "Категория" else (0 if lbl == "" else 1)))
        split_lay.addLayout(hdr)

        self._split_rows_lay = QVBoxLayout()
        self._split_rows_lay.setSpacing(2)
        split_lay.addLayout(self._split_rows_lay)

        self._btn_add_split = QPushButton(" Добавить строку")
        self._btn_add_split.setIcon(icons.get_icon("add", 14, color=C.BRAND))
        self._btn_add_split.setIconSize(QSize(14, 14))
        self._btn_add_split.setObjectName("btnAddSplit")
        self._btn_add_split.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_add_split.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._btn_add_split.clicked.connect(self._add_split_row)
        split_lay.addWidget(self._btn_add_split)

        lay.addWidget(self._split_section)
        # ----------------------------------

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color:#E5E7EB;background:#E5E7EB;max-height:1px;")
        lay.addWidget(sep)

        self._split_warning = QLabel()
        self._split_warning.setObjectName("splitWarning")
        self._split_warning.setWordWrap(True)
        self._split_warning.setStyleSheet(
            "color:#9CA3AF; background:transparent; font-size:12px;"
        )
        lay.addWidget(self._split_warning)

        btn_cancel = SecondaryButton("Отмена")
        btn_cancel.clicked.connect(self.reject)
        self._btn_save = PrimaryButton("Сохранить")
        self._btn_save.clicked.connect(self._on_accept)
        lay.addLayout(self.make_button_row(btn_cancel, self._btn_save))
        self._update_save_state()

    def _on_split_toggle(self):
        if self._split_section.isVisible():
            self._split_section.setVisible(False)
            self._btn_split.setText("  Разделить операцию")
            for row in list(self._split_rows):
                self._split_rows_lay.removeWidget(row)
                row.deleteLater()
            self._split_rows.clear()
        else:
            self._split_section.setVisible(True)
            self._btn_split.setText("  Отменить разделение")
            self._add_split_row()
            self._add_split_row()
        self.adjustSize()
        self._update_save_state()

    def _add_split_row(self, data: dict = None):
        row = _SplitRowWidget(self._split_section)
        if data:
            row.set_data(data)
        row.deleteRequested.connect(self._remove_split_row)
        row.sumChanged.connect(self._update_save_state)
        self._split_rows_lay.addWidget(row)
        self._split_rows.append(row)
        self.adjustSize()
        self._update_save_state()

    def _remove_split_row(self, row: "_SplitRowWidget"):
        if row in self._split_rows:
            self._split_rows.remove(row)
            self._split_rows_lay.removeWidget(row)
            row.deleteLater()
            self.adjustSize()
            self._update_save_state()

    def _update_save_state(self):
        if not hasattr(self, '_btn_save') or self._btn_save is None:
            return
        split_ok = True
        if self._split_rows and self._split_section.isVisible():
            raw = (self.inp_summa.text().strip()
                   .replace(",", ".").replace("\u2212", "-").replace("\u2013", "-").replace(" ", ""))
            try:
                main_sum = float(raw) if raw else 0.0
            except ValueError:
                main_sum = 0.0
            split_sum = 0.0
            for row in self._split_rows:
                rd = row.get_data()
                split_sum += rd.get("Сумма", 0.0)
            split_ok = abs(main_sum - split_sum) < 0.01
        self._btn_save.setEnabled(split_ok)
        self._btn_save.setCursor(
            Qt.CursorShape.PointingHandCursor if split_ok else Qt.CursorShape.ArrowCursor
        )
        _text = "Сумма разбитых строк должна равняться сумме операции"
        if split_ok:
            self._split_warning.setStyleSheet(
                "color:#9CA3AF; background:transparent; font-size:12px;"
            )
        else:
            self._split_warning.setStyleSheet(
                "color:#B45309; background:transparent; font-size:12px; font-weight:600;"
            )
        self._split_warning.setText(f"\u2139  {_text}")
        self._split_warning.setVisible(self._split_section.isVisible() and not split_ok)

    def _prefill(self, data: dict, breakdown: list):
        ts = _to_ts(data.get("Дата"))
        if ts is not None:
            self.date_edit.setDate(QDate(ts.year, ts.month, ts.day))
        num = _to_num(data.get("Сумма"))
        if num is not None:
            self.inp_summa.setText(_num_edit_str(num))
        self.inp_cont.setText(str(data.get("Контрагент") or ""))
        self.inp_cont.setCursorPosition(0)
        self.inp_nazn.setText(str(data.get("Назначение") or ""))
        self.inp_nazn.setCursorPosition(0)
        cat = str(data.get("Категория") or "")
        idx = self.combo_cat.findText(cat)
        if idx >= 0:
            self.combo_cat.setCurrentIndex(idx)
        self.combo_plot.setCurrentText(str(data.get("Участок") or ""))
        if breakdown:
            self._on_split_toggle()
            for bd in breakdown:
                self._add_split_row(bd)
            self._update_save_state()

    def _on_accept(self):
        raw = (self.inp_summa.text().strip()
               .replace(",", ".").replace("\u2212", "-").replace("\u2013", "-").replace(" ", ""))
        if not raw:
            _AlertDialog.show_alert(self, "Ошибка", "Укажите сумму операции")
            return
        try:
            float(raw)
        except ValueError:
            _AlertDialog.show_alert(self, "Ошибка", "Некорректный формат суммы")
            return
        self.accept()

    def get_result(self) -> dict:
        raw = (self.inp_summa.text().strip()
               .replace(",", ".").replace("−", "-").replace("–", "-").replace(" ", ""))
        d = self.date_edit.date()
        result = {
            "Дата":       pd.Timestamp(d.year(), d.month(), d.day()),
            "Контрагент": self.inp_cont.text().strip(),
            "Сумма":      float(raw),
            "Назначение": self.inp_nazn.text().strip(),
            "Категория":  self.combo_cat.currentText(),
            "Участок":    self.combo_plot.currentText().strip(),
        }
        if self._split_rows:
            contragent = self.inp_cont.text().strip()
            result["_breakdown"] = [
                {"Контрагент": contragent, **row.get_data()}
                for row in self._split_rows
            ]
        else:
            result["_breakdown"] = []
        return result

    def _apply_styles(self):
        self.setStyleSheet(self.base_qss() + _SPLIT_BTN_QSS)


# =========================================================================== #
#  Попап фильтра по категориям                                                #
# =========================================================================== #

class _PopupPillButton(QPushButton):
    """Кнопка-пилюля в попапе фильтра — рисуется вручную, без QSS-бордер-артефактов."""

    def __init__(self, text: str, color: "QColor | None", parent=None, *,
                 show_checkmark: bool = True):
        super().__init__(text, parent)
        self._pill_color = color
        # Галочка слева нужна для попапов-ФИЛЬТРОВ (мультивыбор категорий/
        # участков) — там она отличает активные фильтры от неактивных. В
        # _SingleCatPopup эта же кнопка используется для ВЫБОРА значения
        # (не фильтрации): «отмечена» там означает «это текущее значение»,
        # а не «применён фильтр», поэтому там галочка отключается.
        self._show_checkmark = show_checkmark
        self.setCheckable(True)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.setFixedHeight(28)
        fm = QFontMetrics(QFont())
        self.setMinimumWidth(min(fm.horizontalAdvance(text) + 36, 300))

    def paintEvent(self, _event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        color = self._pill_color
        if color:
            h, s, l, _ = color.getHslF()
            if h < 0:
                h, s = 0.0, 0.0
            pill_bd = QColor.fromHslF(h, min(s * 1.0, 1.0), 0.52)
            pill_bg = QColor.fromHslF(h, min(s * 0.65, 1.0), 0.91)
            pill_tx = QColor.fromHslF(h, min(s * 1.2, 1.0), 0.18)
        else:
            pill_bd = QColor("#9CA3AF")
            pill_bg = QColor("#F3F4F6")
            pill_tx = QColor("#374151")

        rf       = QRectF(self.rect()).adjusted(0.5, 0.5, -0.5, -0.5)
        radius_f = rf.height() / 2.0

        if self.isChecked():
            bg = QColor(pill_bg)
            if self.underMouse():
                bg = bg.darker(108)
        elif self.underMouse():
            bg = QColor(pill_bd)
            bg.setAlphaF(0.12)
        else:
            bg = QColor(0, 0, 0, 0)

        painter.setPen(QPen(pill_bd, 1.0))
        painter.setBrush(bg)
        painter.drawRoundedRect(rf, radius_f, radius_f)

        text_rect = self.rect()
        if self.isChecked() and self._show_checkmark:
            # Галочка слева у выбранных пилюль — заметнее показывает, что
            # именно отмечено (особенно в мультивыборе, где иначе выбранные
            # и невыбранные пилюли отличаются только оттенком фона).
            ic_w = 16
            ic_rect = QRect(6, 0, ic_w, self.height())
            f_ic = QFont("Material Symbols Rounded")
            f_ic.setPixelSize(15)
            painter.setFont(f_ic)
            painter.setPen(pill_bd)
            painter.drawText(ic_rect, Qt.AlignmentFlag.AlignCenter, chr(0xF88B))
            text_rect = self.rect().adjusted(ic_w + 4, 0, 0, 0)

        f = QFont()
        f.setPixelSize(12)
        f.setWeight(QFont.Weight.Medium)
        painter.setFont(f)
        painter.setPen(pill_tx)
        fm = QFontMetrics(f)
        elided = fm.elidedText(self.text(), Qt.TextElideMode.ElideRight, text_rect.width() - 12)
        painter.drawText(text_rect, Qt.AlignmentFlag.AlignCenter, elided)


class _SingleCatPopup(QFrame):
    """Popup единичного выбора категории для диалогов Add/Edit и для
    редактирования категории прямо в ячейке таблицы (см. _CategoryDelegate)."""

    itemSelected = pyqtSignal(int)
    editCategoriesRequested = pyqtSignal()
    hidden = pyqtSignal()   # попап скрылся (и выбором, и кликом мимо)
    _MAX_H = 360
    _ACTIONS_BTN_H = 28

    def __init__(self, items: list, current_idx: int, parent=None, *,
                 show_edit_button: bool = False):
        super().__init__(None, Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint)
        # Карточку рисует paintEvent на прозрачном окне — см. докстринг
        # у _CatFilterPopup (тот же приём, та же причина).
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)

        self._items = items
        container = QWidget()
        inner_lay = QVBoxLayout(container)
        inner_lay.setContentsMargins(8, 8, 14, 8)
        inner_lay.setSpacing(4)

        self._btns: list[_PopupPillButton] = []
        group = QButtonGroup(self)
        group.setExclusive(True)

        for i, item in enumerate(items):
            btn = _PopupPillButton(item, CATEGORY_COLORS.get(item), container,
                                    show_checkmark=False)
            btn.setChecked(i == current_idx)
            btn.clicked.connect(lambda _, idx=i: self._select(idx))
            inner_lay.addWidget(btn)
            group.addButton(btn)
            self._btns.append(btn)

        scroll = QScrollArea(self)
        scroll.setWidget(container)
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll.setStyleSheet("QScrollArea{background:#FFFFFF;border:none;}")
        scroll.viewport().setStyleSheet("background:#FFFFFF;")
        self._scroll_style = QStyleFactory.create("Fusion")
        if self._scroll_style is not None:
            scroll.setStyle(self._scroll_style)
            scroll.verticalScrollBar().setStyle(self._scroll_style)
        scroll.verticalScrollBar().setStyleSheet(scrollbar_qss(track="#FFFFFF"))

        outer = QVBoxLayout(self)
        outer.setContentsMargins(3, 3, 3, 3)
        outer.setSpacing(0)

        self._actions_h = 0
        if show_edit_button:
            btn_edit = SecondaryButton("Редактировать категории", icon="edit")
            btn_edit.clicked.connect(self._on_edit_categories)
            row = QHBoxLayout()
            row.setContentsMargins(8, 6, 8, 6)
            row.addWidget(btn_edit)
            divider = QFrame()
            divider.setFixedHeight(1)
            divider.setStyleSheet("background:#E5E9EF; border:none;")
            outer.addLayout(row)
            outer.addWidget(divider)
            self._actions_h = 6 + self._ACTIONS_BTN_H + 6 + 1

        outer.addWidget(scroll)

    def paintEvent(self, a0):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        rect = QRectF(self.rect()).adjusted(0.5, 0.5, -0.5, -0.5)
        p.setPen(QPen(QColor("#C9D8E2")))
        p.setBrush(QColor("#FFFFFF"))
        p.drawRoundedRect(rect, 8, 8)

    def _select(self, idx: int):
        self.itemSelected.emit(idx)
        self.close()

    def _on_edit_categories(self):
        self.editCategoriesRequested.emit()
        self.close()

    def hideEvent(self, event):
        super().hideEvent(event)
        self.hidden.emit()

    def show_at(self, global_pos: "QPoint"):
        n = len(self._btns)
        if n > 0:
            pill_h    = 28
            spacing   = 4
            vmargin   = 16
            content_h = n * pill_h + max(0, n - 1) * spacing + vmargin

            max_btn_w = max(
                (btn.minimumWidth() for btn in self._btns),
                default=180
            )
            popup_w = min(max_btn_w, 300) + 34

            screen      = QApplication.primaryScreen().availableGeometry()
            available_h = screen.bottom() - global_pos.y() - 16 - self._actions_h
            list_h      = min(content_h, self._MAX_H, max(80, available_h))
            popup_h     = list_h + self._actions_h

            self.setFixedWidth(max(popup_w, 200))
            self.setFixedHeight(max(popup_h, 40 + self._actions_h))

        self.move(global_pos)
        self.show()
        self.raise_()


class _CatFilterPopup(QFrame):
    """Выпадающий попап мультивыбора категорий из заголовка таблицы."""

    selectionChanged = pyqtSignal(object)   # set[str]
    _MAX_H = 360
    _ACTIONS_BTN_H = 22
    # Высота строки «Выбрать все / Сбросить» + разделитель: 6+22+6 (action_row) + 1 (divider).
    _ACTIONS_ROW_H = 6 + _ACTIONS_BTN_H + 6 + 1

    def __init__(self):
        super().__init__(None, Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint)
        # Карточку рисует paintEvent на прозрачном окне (не QSS-фон) — та же
        # причина, что и у BaseDialog (ui/dialogs.py): без
        # WA_TranslucentBackground реальная форма окна остаётся
        # прямоугольной, и в углах "скруглённого" QSS-фона проступает
        # исходный прямоугольный фон окна.
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        self._selected: set = set()
        self._btns: dict[str, _PopupPillButton] = {}
        self._hide_time: float = 0.0

        self._container = QWidget()
        self._lay = QVBoxLayout(self._container)
        # Справа отступ больше (14 вместо 8) — зазор между пилюлей и
        # желобом скроллбара, иначе они визуально соприкасаются.
        self._lay.setContentsMargins(8, 8, 14, 8)
        self._lay.setSpacing(4)

        self._scroll = QScrollArea(self)
        self._scroll.setWidget(self._container)
        self._scroll.setWidgetResizable(True)
        self._scroll.setFrameShape(QFrame.Shape.NoFrame)
        self._scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self._scroll.setStyleSheet("QScrollArea{background:#FFFFFF;border:none;}")
        # Вьюпорт красит фон САМ поверх стиля QScrollArea (та же ловушка с
        # просвечивающим базовым фоном QAbstractScrollArea, что и в
        # list_view "Участков", ui/plots_widget.py) — стилизуем его отдельно.
        self._scroll.viewport().setStyleSheet("background:#FFFFFF;")
        # Win11: без принудительного Fusion нативный стиль рисует
        # overlay-скроллбар поверх контента и игнорирует QSS (та же
        # ловушка, что и в "Категориях"/"Участках" — см. CategoryEditorPanel).
        self._scroll_style = QStyleFactory.create("Fusion")
        if self._scroll_style is not None:
            self._scroll.setStyle(self._scroll_style)
            self._scroll.verticalScrollBar().setStyle(self._scroll_style)
        self._scroll.verticalScrollBar().setStyleSheet(scrollbar_qss(track="#FFFFFF"))

        # Строка действий над списком: «Выбрать все» / «Сбросить».
        action_row = QHBoxLayout()
        action_row.setContentsMargins(10, 6, 10, 6)
        action_row.setSpacing(12)
        btn_select_all = LinkButton("Выбрать все")
        btn_select_all.setFixedHeight(self._ACTIONS_BTN_H)
        btn_select_all.clicked.connect(self.select_all)
        action_row.addWidget(btn_select_all)
        action_row.addStretch()
        btn_clear = LinkButton("Сбросить")
        btn_clear.setFixedHeight(self._ACTIONS_BTN_H)
        btn_clear.clicked.connect(self.clear_selection)
        action_row.addWidget(btn_clear)

        divider = QFrame()
        divider.setFixedHeight(1)
        divider.setStyleSheet("background:#E5E9EF; border:none;")

        outer = QVBoxLayout(self)
        # Отступ >= радиуса скругления (8px в paintEvent) — иначе непрозрачный
        # белый вьюпорт квадратом перекрывает антиалиасинг скруглённого угла.
        outer.setContentsMargins(3, 3, 3, 3)
        outer.setSpacing(0)
        outer.addLayout(action_row)
        outer.addWidget(divider)
        outer.addWidget(self._scroll)

    def paintEvent(self, a0):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        rect = QRectF(self.rect()).adjusted(0.5, 0.5, -0.5, -0.5)
        p.setPen(QPen(QColor("#C9D8E2")))
        p.setBrush(QColor("#FFFFFF"))
        p.drawRoundedRect(rect, 8, 8)

    # ------------------------------------------------------------------ build #

    def rebuild(self, categories: list):
        while self._lay.count():
            item = self._lay.takeAt(0)
            if (w := item.widget()):
                w.deleteLater()
        self._btns.clear()

        for cat in categories:
            btn = _PopupPillButton(cat, CATEGORY_COLORS.get(cat), self._container)
            btn.setChecked(cat in self._selected)
            btn.toggled.connect(lambda checked, c=cat: self._on_toggle(c, checked))
            self._lay.addWidget(btn)
            self._btns[cat] = btn

    # ----------------------------------------------------------------- state #

    def _on_toggle(self, cat: str, checked: bool):
        if checked:
            self._selected.add(cat)
        else:
            self._selected.discard(cat)
        self.selectionChanged.emit(set(self._selected))

    def set_selected(self, sel: set):
        self._selected = set(sel)
        for cat, btn in self._btns.items():
            btn.blockSignals(True)
            btn.setChecked(cat in sel)
            btn.blockSignals(False)

    def clear_selection(self):
        self._selected.clear()
        for btn in self._btns.values():
            btn.blockSignals(True)
            btn.setChecked(False)
            btn.blockSignals(False)
        self.selectionChanged.emit(set())

    def select_all(self):
        self._selected = set(self._btns.keys())
        for btn in self._btns.values():
            btn.blockSignals(True)
            btn.setChecked(True)
            btn.blockSignals(False)
        self.selectionChanged.emit(set(self._selected))

    def get_selected(self) -> set:
        return set(self._selected)

    # ---------------------------------------------------------------- popup #

    def show_at(self, global_pos: "QPoint"):
        n = len(self._btns)
        if n > 0:
            # Высота: аналитически (не sizeHint — он ненадёжен до первого show)
            pill_h   = 28   # _PopupPillButton.setFixedHeight(28)
            spacing  = 4    # self._lay.setSpacing(4)
            vmargin  = 16   # 8 top + 8 bottom
            content_h = n * pill_h + max(0, n - 1) * spacing + vmargin

            max_btn_w = max(
                (btn.minimumWidth() for btn in self._btns.values()),
                default=180
            )
            # Буфер вокруг пилюли по ширине: 3+3 внешнее поле окна (paintEvent) +
            # 8+14 отступы _lay (слева/справа, справа увеличен под зазор до
            # скроллбара) + 6 сам скроллбар = 34. Потолок ширины — от ширины
            # САМОЙ ПИЛЮЛИ (у неё свой потолок 300, см. _PopupPillButton),
            # а не от итоговой ширины попапа: иначе для длинных названий
            # буфер не помещается и layout съедает правый отступ до нуля.
            popup_w = min(max_btn_w, 300) + 34

            screen      = QApplication.primaryScreen().availableGeometry()
            available_h = screen.bottom() - global_pos.y() - 16
            list_h      = min(content_h, self._MAX_H, max(80, available_h))
            popup_h     = list_h + self._ACTIONS_ROW_H

            self.setFixedWidth(max(popup_w, 200))
            self.setFixedHeight(max(popup_h, 40 + self._ACTIONS_ROW_H))

        self.move(global_pos)
        self.show()
        self.raise_()

    def hideEvent(self, event):
        self._hide_time = time.monotonic()
        super().hideEvent(event)

    def was_just_hidden(self) -> bool:
        return time.monotonic() - self._hide_time < 0.2


class _PlotFilterPopup(QFrame):
    """Выпадающий попап мультивыбора участков из заголовка таблицы —
    тот же визуальный язык, что и _CatFilterPopup, но вместо «Выбрать
    все» — строка поиска: участков в проекте обычно много больше, чем
    категорий, и пролистывать список нет смысла."""

    selectionChanged = pyqtSignal(object)   # set[str]
    _MAX_H = 360
    _ACTIONS_BTN_H = 22
    # Высота строки поиска/«Сбросить» + разделитель: 6+22+6 (action_row) + 1 (divider).
    _ACTIONS_ROW_H = 6 + _ACTIONS_BTN_H + 6 + 1

    def __init__(self):
        super().__init__(None, Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint)
        # Карточку рисует paintEvent на прозрачном окне (не QSS-фон) — та же
        # причина, что и у _CatFilterPopup.
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        self._selected: set = set()
        self._btns: dict[str, _PopupPillButton] = {}
        self._plot_order: list = []   # порядок из rebuild() — база для поиска
        self._hide_time: float = 0.0

        self._container = QWidget()
        self._lay = QVBoxLayout(self._container)
        self._lay.setContentsMargins(8, 8, 14, 8)
        self._lay.setSpacing(4)

        self._scroll = QScrollArea(self)
        self._scroll.setWidget(self._container)
        self._scroll.setWidgetResizable(True)
        self._scroll.setFrameShape(QFrame.Shape.NoFrame)
        self._scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self._scroll.setStyleSheet("QScrollArea{background:#FFFFFF;border:none;}")
        self._scroll.viewport().setStyleSheet("background:#FFFFFF;")
        self._scroll_style = QStyleFactory.create("Fusion")
        if self._scroll_style is not None:
            self._scroll.setStyle(self._scroll_style)
            self._scroll.verticalScrollBar().setStyle(self._scroll_style)
        self._scroll.verticalScrollBar().setStyleSheet(scrollbar_qss(track="#FFFFFF"))

        # Строка действий над списком: поиск участка + «Сбросить».
        action_row = QHBoxLayout()
        action_row.setContentsMargins(10, 6, 10, 6)
        action_row.setSpacing(8)
        self._search_input = QLineEdit()
        self._search_input.setPlaceholderText("Поиск участка...")
        self._search_input.setClearButtonEnabled(True)
        self._search_input.setFixedHeight(self._ACTIONS_BTN_H)
        self._search_input.setStyleSheet(
            "QLineEdit{background:#F8F9FA;border:1px solid #D5DCE4;border-radius:5px;"
            "padding:2px 8px;font-size:12px;color:#1F2937;}"
            "QLineEdit:focus{border:1px solid #07414F;}")
        self._search_input.textChanged.connect(self._on_search_text)
        action_row.addWidget(self._search_input, stretch=1)
        btn_clear = LinkButton("Сбросить")
        btn_clear.setFixedHeight(self._ACTIONS_BTN_H)
        btn_clear.clicked.connect(self.clear_selection)
        action_row.addWidget(btn_clear)

        divider = QFrame()
        divider.setFixedHeight(1)
        divider.setStyleSheet("background:#E5E9EF; border:none;")

        outer = QVBoxLayout(self)
        # Отступ >= радиуса скругления (8px в paintEvent) — та же причина,
        # что и в _CatFilterPopup.
        outer.setContentsMargins(3, 3, 3, 3)
        outer.setSpacing(0)
        outer.addLayout(action_row)
        outer.addWidget(divider)
        outer.addWidget(self._scroll)

    def paintEvent(self, a0):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        rect = QRectF(self.rect()).adjusted(0.5, 0.5, -0.5, -0.5)
        p.setPen(QPen(QColor("#C9D8E2")))
        p.setBrush(QColor("#FFFFFF"))
        p.drawRoundedRect(rect, 8, 8)

    # ------------------------------------------------------------------ build #

    def rebuild(self, plots: list):
        while self._lay.count():
            item = self._lay.takeAt(0)
            if (w := item.widget()):
                w.deleteLater()
        self._btns.clear()
        self._plot_order = list(plots)

        for num in plots:
            btn = _PopupPillButton(num, None, self._container)
            btn.setChecked(num in self._selected)
            btn.toggled.connect(lambda checked, n=num: self._on_toggle(n, checked))
            self._btns[num] = btn
        # Раскладка (порядок + завершающий stretch) собирается в
        # _apply_search_filter — единая точка добавления в self._lay.
        self._apply_search_filter()

    # ----------------------------------------------------------------- state #

    def _on_toggle(self, num: str, checked: bool):
        if checked:
            self._selected.add(num)
        else:
            self._selected.discard(num)
        self.selectionChanged.emit(set(self._selected))

    def set_selected(self, sel: set):
        self._selected = set(sel)
        for num, btn in self._btns.items():
            btn.blockSignals(True)
            btn.setChecked(num in sel)
            btn.blockSignals(False)

    def clear_selection(self):
        self._selected.clear()
        for btn in self._btns.values():
            btn.blockSignals(True)
            btn.setChecked(False)
            btn.blockSignals(False)
        self.selectionChanged.emit(set())

    def get_selected(self) -> set:
        return set(self._selected)

    # ----------------------------------------------------------------- поиск #

    def _on_search_text(self, _text: str):
        self._apply_search_filter()

    def _apply_search_filter(self):
        """Совпавшие с поиском пилюли поднимаются в начало списка (не просто
        скрывают несовпавшие на месте — иначе результаты оказываются
        вперемешку, разделённые пустыми местами скрытых кнопок).

        Раскладка каждый раз собирается заново, а не просто переставляется —
        внутри QScrollArea (widgetResizable=True) окно попапа держит
        ФИКСИРОВАННУЮ высоту, выставленную под полный список ещё при
        открытии (show_at); когда фильтр прячет большинство пилюль, реальный
        контент становится намного ниже этой высоты, и QVBoxLayout БЕЗ
        завершающего stretch-элемента размазывает освободившееся место
        поровну МЕЖДУ видимыми пилюлями (это не связано с порядком/hide —
        воспроизводится и с двумя виджетами без скрытия вовсе). Обязательный
        addStretch() в хвосте отдаёт весь излишек ему, а не пилюлям."""
        text = self._search_input.text().strip().lower()
        if text:
            matched = [n for n in self._plot_order if text in n.lower()]
            rest    = [n for n in self._plot_order if text not in n.lower()]
            ordered = matched + rest
        else:
            ordered = self._plot_order

        while self._lay.count():
            self._lay.takeAt(0)
        for num in ordered:
            self._lay.addWidget(self._btns[num])
        self._lay.addStretch(1)

        for num in self._plot_order:
            self._btns[num].setVisible(not text or text in num.lower())

    # ---------------------------------------------------------------- popup #

    def show_at(self, global_pos: "QPoint"):
        n = len(self._btns)
        if n > 0:
            pill_h   = 28   # _PopupPillButton.setFixedHeight(28)
            spacing  = 4    # self._lay.setSpacing(4)
            vmargin  = 16   # 8 top + 8 bottom
            content_h = n * pill_h + max(0, n - 1) * spacing + vmargin

            max_btn_w = max(
                (btn.minimumWidth() for btn in self._btns.values()),
                default=180
            )
            popup_w = min(max_btn_w, 300) + 34

            screen      = QApplication.primaryScreen().availableGeometry()
            available_h = screen.bottom() - global_pos.y() - 16
            list_h      = min(content_h, self._MAX_H, max(80, available_h))
            popup_h     = list_h + self._ACTIONS_ROW_H

            self.setFixedWidth(max(popup_w, 240))
            self.setFixedHeight(max(popup_h, 40 + self._ACTIONS_ROW_H))

        self._search_input.clear()
        self.move(global_pos)
        self.show()
        self.raise_()
        self._search_input.setFocus()

    def hideEvent(self, event):
        self._hide_time = time.monotonic()
        super().hideEvent(event)

    def was_just_hidden(self) -> bool:
        return time.monotonic() - self._hide_time < 0.2


# =========================================================================== #
#  Шапка таблицы «Детализация»                                                #
# =========================================================================== #

class _DetailHeaderView(QHeaderView):
    """Кастомная шапка таблицы в стиле вкладки «Список участков»."""

    catFilterChanged    = pyqtSignal(object)   # set[str]
    periodFilterChanged = pyqtSignal(object, object)   # (date_from, date_to) | (None, None)
    plotFilterChanged    = pyqtSignal(object)   # set[str]
    checkAllToggled      = pyqtSignal()

    _BG      = QColor("#C9D8E2")
    _FG      = QColor("#07414F")
    _BORDER  = QColor("#B5C8D5")
    _ARR_ON  = QColor("#07414F")
    _ARR_OFF = QColor("#9AABB6")
    _IC_W    = 22

    # Мастер-чекбокс «выбрать все» — те же глифы/цвета, что и у чекбоксов
    # строк (_DetailCheckDelegate) и у мастер-чекбокса «Участков».
    _CB_ON     = chr(0xE834)   # check_box
    _CB_OFF    = chr(0xE835)   # check_box_outline_blank
    _CB_MIXED  = chr(0xE15B)   # remove (промежуточное состояние)
    _CB_COLOR_ON  = QColor("#07414F")
    _CB_COLOR_OFF = QColor("#C3CAD3")

    def __init__(self, parent=None):
        super().__init__(Qt.Orientation.Horizontal, parent)
        # False — иначе Qt сортирует по клику в ЛЮБОМ месте секции; клик
        # обрабатываем сами (см. mousePressEvent), только в зоне стрелки.
        self.setSectionsClickable(False)
        self.setSortIndicatorShown(False)
        # Qt по умолчанию отдаёт sortIndicatorSection()==0 ДО первого явного
        # setSortIndicator — из-за этого первый клик именно по нулевому
        # столбцу выглядел бы как «уже отсортировано по возрастанию» и сразу
        # прыгал на убывание. -1 однозначно значит «сортировки нет».
        self.setSortIndicator(-1, Qt.SortOrder.AscendingOrder)
        self.setFixedHeight(34)
        self.setMouseTracking(True)
        self._cat_col:     int                  = -1
        self._cat_active:  bool                 = False
        self._cat_hovered: bool                 = False
        self._cat_popup:   "_CatFilterPopup | None" = None
        self._period_col:     int                     = -1
        self._period_active:  bool                    = False
        self._period_hovered: bool                    = False
        self._period_popup:   "_PeriodFilterPopup | None" = None
        self._period_from = None
        self._period_to = None
        self._plot_col:     int                  = -1
        self._plot_active:  bool                 = False
        self._plot_hovered: bool                 = False
        self._plot_popup:   "_PlotFilterPopup | None" = None
        self._check_col:      int = -1
        self._check_total:    int = 0
        self._check_selected: int = 0
        self._fill_tag = QFont.Tag.fromString("FILL")

    _SORT_W = 18   # ширина зоны иконки сортировки (слева, перед текстом)

    def _content_left(self, sec_left: int) -> int:
        """X, с которого начинается текст — после иконки сортировки
        (см. sort_rect в paintSection)."""
        return sec_left + 4 + self._SORT_W + 4

    def _sort_icon_zone(self, sec_rect: QRect) -> QRect:
        """QRect стрелки сортировки — та же геометрия, что и sort_rect
        в paintSection (слева, перед текстом)."""
        return QRect(sec_rect.left() + 4, sec_rect.top(), self._SORT_W, sec_rect.height())

    # ---------------------------------------------------------- фильтр категорий #

    def set_cat_col(self, col: int, categories: list, selected: set = None):
        self._cat_col = col
        if self._cat_popup is None:
            self._cat_popup = _CatFilterPopup()
            self._cat_popup.selectionChanged.connect(self._on_cat_selection_changed)
        self._cat_popup.rebuild(categories)
        sel = selected or set()
        if sel:
            self._cat_popup.set_selected(sel)
        self._cat_active = bool(sel)
        self.viewport().update()

    def _on_cat_selection_changed(self, selected: set):
        self._cat_active = bool(selected)
        self.viewport().update()
        self.catFilterChanged.emit(selected)

    def _cat_icon_zone(self, sec_rect: QRect) -> QRect:
        IC_W = self._IC_W
        return QRect(sec_rect.right() - IC_W - 4, sec_rect.top(), IC_W, sec_rect.height())

    # ----------------------------------------------------------- фильтр периода #

    def set_period_col(self, col: int, date_from, date_to):
        self._period_col = col
        if self._period_popup is None:
            self._period_popup = _PeriodFilterPopup()
            self._period_popup.periodChanged.connect(self._on_period_selection_changed)
        self._period_from = date_from
        self._period_to = date_to
        self._period_active = bool(date_from or date_to)
        self.viewport().update()

    def _on_period_selection_changed(self, date_from, date_to):
        self._period_from = date_from
        self._period_to = date_to
        self._period_active = bool(date_from or date_to)
        self.viewport().update()
        self.periodFilterChanged.emit(date_from, date_to)

    def _period_icon_zone(self, sec_rect: QRect) -> QRect:
        IC_W = self._IC_W
        return QRect(sec_rect.right() - IC_W - 4, sec_rect.top(), IC_W, sec_rect.height())

    # ----------------------------------------------------------- фильтр участков #

    def set_plot_col(self, col: int, plots: list, selected: set = None):
        self._plot_col = col
        if self._plot_popup is None:
            self._plot_popup = _PlotFilterPopup()
            self._plot_popup.selectionChanged.connect(self._on_plot_selection_changed)
        self._plot_popup.rebuild(plots)
        sel = selected or set()
        if sel:
            self._plot_popup.set_selected(sel)
        self._plot_active = bool(sel)
        self.viewport().update()

    def _on_plot_selection_changed(self, selected: set):
        self._plot_active = bool(selected)
        self.viewport().update()
        self.plotFilterChanged.emit(selected)

    def _plot_icon_zone(self, sec_rect: QRect) -> QRect:
        IC_W = self._IC_W
        return QRect(sec_rect.right() - IC_W - 4, sec_rect.top(), IC_W, sec_rect.height())

    # ----------------------------------------------------------- чекбокс «все» #

    def set_check_col(self, col: int):
        self._check_col = col

    def set_check_state(self, total: int, selected: int):
        if (self._check_total, self._check_selected) != (total, selected):
            self._check_total = total
            self._check_selected = selected
            self.viewport().update()

    def _paint_master_checkbox(self, painter: QPainter, rect: QRect):
        if self._check_total <= 0 or self._check_selected <= 0:
            icon, color = self._CB_OFF, self._CB_COLOR_OFF
        elif self._check_selected >= self._check_total:
            icon, color = self._CB_ON, self._CB_COLOR_ON
        else:
            icon, color = self._CB_MIXED, self._CB_COLOR_ON
        f = QFont("Material Symbols Rounded")
        f.setPixelSize(18)
        f.setVariableAxis(self._fill_tag, 1.0 if icon != self._CB_OFF else 0.0)
        painter.setFont(f)
        painter.setPen(color)
        painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, icon)

    # ------------------------------------------------------------------ paint #

    def paintSection(self, painter: QPainter, rect: QRect, logical_index: int):
        if not rect.isValid():
            return
        painter.save()
        painter.fillRect(rect, self._BG)

        painter.setPen(QPen(self._BORDER, 1))
        painter.drawLine(rect.right(), rect.top() + 4, rect.right(), rect.bottom() - 4)

        if logical_index == self._check_col:
            self._paint_master_checkbox(painter, rect)
            painter.restore()
            return

        model = self.model()
        label = (
            str(model.headerData(logical_index, Qt.Orientation.Horizontal,
                                 Qt.ItemDataRole.DisplayRole) or "")
            if model else ""
        )
        if label:
            # Одиночный символ Material Symbols — рисуем как иконку, без стрелок
            if len(label) == 1 and 0xE000 <= ord(label) <= 0xF8FF:
                f_ic = QFont("Material Symbols Rounded")
                f_ic.setPixelSize(18)
                painter.setPen(self._FG)
                painter.setFont(f_ic)
                painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, label)
            else:
                is_cat    = (logical_index == self._cat_col)
                is_period = (logical_index == self._period_col)
                is_plot   = (logical_index == self._plot_col)
                IC_W     = self._IC_W
                # Иконка сортировки — слева, перед текстом (не у правого края).
                sort_rect = QRect(rect.left() + 4, rect.top(), self._SORT_W, rect.height())
                content_left = self._content_left(rect.left())
                cat_rect = self._cat_icon_zone(rect) if is_cat else None
                period_rect = self._period_icon_zone(rect) if is_period else None
                plot_rect = self._plot_icon_zone(rect) if is_plot else None

                if is_cat:
                    title_max_w = max(0, cat_rect.left() - content_left - 6)
                    text_rect = QRect(content_left, rect.top(),
                                      title_max_w, rect.height())
                elif is_period:
                    title_max_w = max(0, period_rect.left() - content_left - 6)
                    text_rect = QRect(content_left, rect.top(),
                                      title_max_w, rect.height())
                elif is_plot:
                    title_max_w = max(0, plot_rect.left() - content_left - 6)
                    text_rect = QRect(content_left, rect.top(),
                                      title_max_w, rect.height())
                else:
                    text_rect = QRect(content_left, rect.top(),
                                      rect.right() - content_left - 4, rect.height())

                painter.setPen(self._FG)
                f = QFont(); f.setPixelSize(12); f.setBold(True)
                painter.setFont(f)
                painter.drawText(text_rect,
                                 Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
                                 label)

                if is_cat:
                    painter.setRenderHint(QPainter.RenderHint.Antialiasing)
                    if self._cat_hovered:
                        painter.setPen(Qt.PenStyle.NoPen)
                        painter.setBrush(QColor(C.BRAND_GHOST))
                        painter.drawRoundedRect(
                            cat_rect.adjusted(1, 3, -1, -3), 6, 6)
                    f_ico = QFont("Material Symbols Rounded"); f_ico.setPixelSize(18)
                    f_ico.setVariableAxis(self._fill_tag, 1.0 if self._cat_active else 0.0)
                    painter.setFont(f_ico)
                    painter.setPen(self._FG)
                    painter.drawText(cat_rect, Qt.AlignmentFlag.AlignCenter, chr(0xEF4F))
                elif is_period:
                    painter.setRenderHint(QPainter.RenderHint.Antialiasing)
                    if self._period_hovered:
                        painter.setPen(Qt.PenStyle.NoPen)
                        painter.setBrush(QColor(C.BRAND_GHOST))
                        painter.drawRoundedRect(
                            period_rect.adjusted(1, 3, -1, -3), 6, 6)
                    f_ico = QFont("Material Symbols Rounded"); f_ico.setPixelSize(18)
                    f_ico.setVariableAxis(self._fill_tag, 1.0 if self._period_active else 0.0)
                    painter.setFont(f_ico)
                    painter.setPen(self._FG)
                    painter.drawText(period_rect, Qt.AlignmentFlag.AlignCenter, chr(0xEBCC))
                elif is_plot:
                    painter.setRenderHint(QPainter.RenderHint.Antialiasing)
                    if self._plot_hovered:
                        painter.setPen(Qt.PenStyle.NoPen)
                        painter.setBrush(QColor(C.BRAND_GHOST))
                        painter.drawRoundedRect(
                            plot_rect.adjusted(1, 3, -1, -3), 6, 6)
                    f_ico = QFont("Material Symbols Rounded"); f_ico.setPixelSize(18)
                    f_ico.setVariableAxis(self._fill_tag, 1.0 if self._plot_active else 0.0)
                    painter.setFont(f_ico)
                    painter.setPen(self._FG)
                    painter.drawText(plot_rect, Qt.AlignmentFlag.AlignCenter, chr(0xEF4F))

                painter.setRenderHint(QPainter.RenderHint.Antialiasing)
                cx = sort_rect.left() + sort_rect.width() // 2
                cy = sort_rect.top() + sort_rect.height() // 2
                is_sorted = (self.sortIndicatorSection() == logical_index)
                asc  = is_sorted and self.sortIndicatorOrder() == Qt.SortOrder.AscendingOrder
                desc = is_sorted and self.sortIndicatorOrder() == Qt.SortOrder.DescendingOrder

                painter.setPen(Qt.PenStyle.NoPen)
                painter.setBrush(self._ARR_ON if asc else self._ARR_OFF)
                painter.drawPolygon(QPolygon([
                    QPoint(cx - 4, cy - 1), QPoint(cx + 4, cy - 1), QPoint(cx, cy - 6),
                ]))
                painter.setBrush(self._ARR_ON if desc else self._ARR_OFF)
                painter.drawPolygon(QPolygon([
                    QPoint(cx - 4, cy + 1), QPoint(cx + 4, cy + 1), QPoint(cx, cy + 6),
                ]))

        painter.restore()

    # --------------------------------------------------------------- mouse #

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            pos     = event.position().toPoint()
            logical = self.logicalIndexAt(pos.x())

            # Мастер-чекбокс «выбрать все» — вся секция кликабельна, не
            # выделенная под-зона (столбец узкий и целиком занят чекбоксом).
            if self._check_col >= 0 and logical == self._check_col:
                self.checkAllToggled.emit()
                return

            # Фильтр категорий
            if self._cat_col >= 0 and logical == self._cat_col:
                sec_x    = self.sectionViewportPosition(self._cat_col)
                sec_rect = QRect(sec_x, 0, self.sectionSize(self._cat_col), self.height())
                if self._cat_icon_zone(sec_rect).contains(pos):
                    if self._cat_popup and self._cat_popup.isVisible():
                        self._cat_popup.hide()
                        return
                    if self._cat_popup and self._cat_popup.was_just_hidden():
                        return
                    if self._cat_popup:
                        vp_pt     = QPoint(sec_x, self.viewport().height())
                        global_pt = self.viewport().mapToGlobal(vp_pt)
                        self._cat_popup.show_at(global_pt)
                    return

            # Фильтр периода
            if self._period_col >= 0 and logical == self._period_col:
                sec_x    = self.sectionViewportPosition(self._period_col)
                sec_rect = QRect(sec_x, 0, self.sectionSize(self._period_col), self.height())
                if self._period_icon_zone(sec_rect).contains(pos):
                    if self._period_popup and self._period_popup.isVisible():
                        self._period_popup.hide()
                        return
                    if self._period_popup and self._period_popup.was_just_hidden():
                        return
                    if self._period_popup:
                        self._period_popup.set_period(self._period_from, self._period_to)
                        vp_pt     = QPoint(sec_x, self.viewport().height())
                        global_pt = self.viewport().mapToGlobal(vp_pt)
                        self._period_popup.show_at(global_pt)
                    return

            # Фильтр участков
            if self._plot_col >= 0 and logical == self._plot_col:
                sec_x    = self.sectionViewportPosition(self._plot_col)
                sec_rect = QRect(sec_x, 0, self.sectionSize(self._plot_col), self.height())
                if self._plot_icon_zone(sec_rect).contains(pos):
                    if self._plot_popup and self._plot_popup.isVisible():
                        self._plot_popup.hide()
                        return
                    if self._plot_popup and self._plot_popup.was_just_hidden():
                        return
                    if self._plot_popup:
                        vp_pt     = QPoint(sec_x, self.viewport().height())
                        global_pt = self.viewport().mapToGlobal(vp_pt)
                        self._plot_popup.show_at(global_pt)
                    return

            # Сортировка — только клик по стрелке, не по всей ячейке.
            # 3 состояния по кругу: по возрастанию → по убыванию → без
            # сортировки (возврат к исходному порядку).
            if logical >= 0:
                cols = self.model().columns() if self.model() else []
                col_name = cols[logical] if 0 <= logical < len(cols) else None
                if col_name not in (_EDIT_COL, _CHECK_COL):
                    sec_x    = self.sectionViewportPosition(logical)
                    sec_rect = QRect(sec_x, 0, self.sectionSize(logical), self.height())
                    if self._sort_icon_zone(sec_rect).contains(pos):
                        cur_col   = self.sortIndicatorSection()
                        cur_order = self.sortIndicatorOrder()
                        if cur_col != logical:
                            self.setSortIndicator(logical, Qt.SortOrder.AscendingOrder)
                        elif cur_order == Qt.SortOrder.AscendingOrder:
                            self.setSortIndicator(logical, Qt.SortOrder.DescendingOrder)
                        else:
                            self.setSortIndicator(-1, Qt.SortOrder.AscendingOrder)
                        return
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        # Базовый QHeaderView.mouseReleaseEvent переключает сортировку по
        # клику в ЛЮБОМ месте секции (старая 2-состояний логика), если
        # sectionsClickable=True. Основной источник — setHeader() в
        # _setup_ui (DetailWidget), который синхронизирует sectionsClickable
        # с tree.isSortingEnabled(); держим здесь как доп. страховку на
        # случай, если это состояние где-то ещё включат заново — всё решение
        # о сортировке уже принято на mousePressEvent (клик именно по стрелке).
        event.accept()

    def mouseMoveEvent(self, event):
        pos  = event.position().toPoint()
        hand = False
        logical = self.logicalIndexAt(pos.x())

        if self._check_col >= 0 and logical == self._check_col:
            hand = True

        # Иконка фильтра категорий — hover-фон + курсор-рука
        if self._cat_col >= 0:
            sec_x    = self.sectionViewportPosition(self._cat_col)
            sec_rect = QRect(sec_x, 0, self.sectionSize(self._cat_col), self.height())
            cat_hov  = (logical == self._cat_col
                        and self._cat_icon_zone(sec_rect).contains(pos))
            if cat_hov != self._cat_hovered:
                self._cat_hovered = cat_hov
                self.viewport().update()
            if cat_hov:
                hand = True

        # Иконка фильтра периода — hover-фон + курсор-рука
        if self._period_col >= 0:
            sec_x    = self.sectionViewportPosition(self._period_col)
            sec_rect = QRect(sec_x, 0, self.sectionSize(self._period_col), self.height())
            period_hov = (logical == self._period_col
                          and self._period_icon_zone(sec_rect).contains(pos))
            if period_hov != self._period_hovered:
                self._period_hovered = period_hov
                self.viewport().update()
            if period_hov:
                hand = True

        # Иконка фильтра участков — hover-фон + курсор-рука
        if self._plot_col >= 0:
            sec_x    = self.sectionViewportPosition(self._plot_col)
            sec_rect = QRect(sec_x, 0, self.sectionSize(self._plot_col), self.height())
            plot_hov = (logical == self._plot_col
                        and self._plot_icon_zone(sec_rect).contains(pos))
            if plot_hov != self._plot_hovered:
                self._plot_hovered = plot_hov
                self.viewport().update()
            if plot_hov:
                hand = True

        self.viewport().setCursor(
            Qt.CursorShape.PointingHandCursor if hand else Qt.CursorShape.ArrowCursor
        )
        super().mouseMoveEvent(event)

    def leaveEvent(self, event):
        if self._cat_hovered:
            self._cat_hovered = False
            self.viewport().update()
        if self._period_hovered:
            self._period_hovered = False
            self.viewport().update()
        if self._plot_hovered:
            self._plot_hovered = False
            self.viewport().update()
        _AppTooltip.hide()
        super().leaveEvent(event)

    def event(self, e):
        if e.type() == QEvent.Type.ToolTip and (
            self._cat_col >= 0 or self._period_col >= 0 or self._plot_col >= 0
        ):
            pos     = e.pos()
            logical = self.logicalIndexAt(pos.x())
            if logical == self._cat_col:
                sec_x    = self.sectionViewportPosition(self._cat_col)
                sec_rect = QRect(sec_x, 0, self.sectionSize(self._cat_col), self.height())
                if self._cat_icon_zone(sec_rect).contains(pos):
                    _AppTooltip.show_at("Фильтровать категории", e.globalPos())
                    return True
            if logical == self._period_col:
                sec_x    = self.sectionViewportPosition(self._period_col)
                sec_rect = QRect(sec_x, 0, self.sectionSize(self._period_col), self.height())
                if self._period_icon_zone(sec_rect).contains(pos):
                    _AppTooltip.show_at("Фильтровать по периоду", e.globalPos())
                    return True
            if logical == self._plot_col:
                sec_x    = self.sectionViewportPosition(self._plot_col)
                sec_rect = QRect(sec_x, 0, self.sectionSize(self._plot_col), self.height())
                if self._plot_icon_zone(sec_rect).contains(pos):
                    _AppTooltip.show_at("Фильтровать участки", e.globalPos())
                    return True
            _AppTooltip.hide()
        return super().event(e)


# =========================================================================== #
#  Виджет вкладки «Детализация»                                               #
# =========================================================================== #

class DetailWidget(QWidget):
    dataLoaded = pyqtSignal(object)

    # «Выбрано: N» — тот же стиль/логика (серый при 0, бирюзовый жирный при
    # выборе), что и в ui/plots_widget.py::PlotsWidget, для единого языка.
    _SELECTED_LBL_OFF_SS = "font-size:12px; color:#9CA3AF; background:transparent; font-weight:600;"
    _SELECTED_LBL_ON_SS  = "font-size:12px; color:#07414F; background:transparent; font-weight:600;"

    def __init__(self):
        super().__init__()
        # Прозрачный фон страницы — чтобы проступал белый contentFrame (окно
        # вкладки), как у PlotsWidget. autoFill красил страницу сплошным
        # палитровым цветом ПОВЕРХ скруглённого белого фрейма — отсюда были
        # видны квадратные углы вместо скруглённых.
        self.setAutoFillBackground(False)
        self.df_full = None
        self._manual_rows: set[int] = set()
        self._manual_cells: set[tuple[int, str]] = set()
        # df_idx -> {"Дата","Сумма","Контрагент","Назначение"} — снимок
        # СЫРЫХ данных повторного импорта, обнаруженного при загрузке в
        # режиме "Добавить к существующим данным" (см. load_file). Строка
        # НЕ дублируется — вместо этого существующая помечается, а
        # пользователь решает через контекстное меню: восстановить исходные
        # значения из повторного импорта или оставить как отредактировал.
        # Не сохраняется в проект — живёт только в рамках сессии.
        self._dup_pending: dict[int, dict] = {}
        self._cat_col: int | None = None
        self._cont_col: int | None = None
        self._plot_col: int | None = None
        self._hdr_cat_filter: set = set()
        self._hdr_plot_filter: set = set()
        self._filter_mode = "all"   # all | income | expense
        self._period_from = None    # date | None — фильтр по периоду
        self._period_to = None
        self._search_text = ""      # поиск по Контрагенту/Назначению
        self._setup_ui()

    # ----------------------------------------------------------------- UI -- #
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(14)

        # Шапка вкладки — тот же стиль, что и «Участки»: заголовок слева,
        # иконочные кнопки без подписи (только тултип) справа.
        top_bar = QHBoxLayout()
        top_bar.setSpacing(8)
        lbl_title = QLabel("Операции")
        lbl_title.setStyleSheet(
            "font-size:14px; font-weight:700; color:#1F2937; background:transparent;")
        top_bar.addWidget(lbl_title)
        top_bar.addStretch()

        def _hdr_icon_btn(tooltip: str, handler) -> QPushButton:
            b = QPushButton()
            b.setFixedSize(32, 32)
            b.setIconSize(QSize(22, 22))
            b.setCursor(Qt.CursorShape.PointingHandCursor)
            b.installEventFilter(_TooltipFilter(tooltip, b))
            b.setStyleSheet(
                "QPushButton{background:transparent;border:none;border-radius:6px;}"
                "QPushButton:hover{background:#EBF4F6;}")
            b.clicked.connect(handler)
            return b

        # «Добавить операцию» — леве́е всех остальных иконок (единственная
        # с нестандартным расположением по просьбе — остальные три идут в
        # прежнем относительном порядке).
        btn_add_row = _hdr_icon_btn("Добавить операцию", self._add_row)
        btn_add_row.setIcon(icons.get_icon(0xE145, 22, color="#9CA3AF"))  # add
        top_bar.addWidget(btn_add_row)

        self.btn_load = _hdr_icon_btn("Импорт из Excel", self.load_file)
        self.btn_load.setIcon(icons.get_icon(0xEAF3, 22, color="#9CA3AF"))
        top_bar.addWidget(self.btn_load)

        btn_cat = _hdr_icon_btn("Категории", self._show_cat_editor)
        btn_cat.setIcon(icons.get_icon("category", 22, color="#9CA3AF"))
        top_bar.addWidget(btn_cat)

        btn_excel = _hdr_icon_btn("Экспорт в Excel", self._export_excel)
        btn_excel.setIcon(icons.get_icon(0xF3B2, 22, color="#9CA3AF"))  # file_export
        top_bar.addWidget(btn_excel)

        # «Удалить выбранные» — перенесена из иконки в заголовке столбца
        # чекбоксов (как во вкладке «Участки»): обычная иконка тулбара,
        # неактивна, пока не отмечена ни одна строка (см.
        # _on_check_selection_changed).
        self._btn_bulk_delete = _hdr_icon_btn("Удалить выбранные", self._delete_selected)
        self._btn_bulk_delete.setIcon(icons.get_icon(0xE92B, 22, color="#DC2626"))  # delete
        self._btn_bulk_delete.setEnabled(False)
        self._btn_bulk_delete.setStyleSheet(
            "QPushButton{background:transparent;border:none;border-radius:6px;}"
            "QPushButton:hover{background:#FEF2F2;}"
            "QPushButton:disabled{background:transparent;}")
        top_bar.addWidget(self._btn_bulk_delete)

        layout.addLayout(top_bar)

        # ── Вкладки-фильтры: Все / Пополнения / Списания ─────────────────
        # Тот же компонент, что и в "Участках" (_FilterTabButton) — единый
        # визуальный язык вкладок-фильтров по всему приложению.
        tabs_row = QHBoxLayout()
        tabs_row.setContentsMargins(0, 4, 0, 0)
        tabs_row.setSpacing(20)
        self._op_tab_group = QButtonGroup(self)
        self._op_tab_group.setExclusive(True)
        self._op_tab_buttons: dict[str, _FilterTabButton] = {}
        for mode, label in (("all", "Все"), ("income", "Пополнения"), ("expense", "Списания")):
            btn = _FilterTabButton(label)
            btn.clicked.connect(lambda checked, m=mode: self._on_op_filter_tab(m))
            self._op_tab_group.addButton(btn)
            self._op_tab_buttons[mode] = btn
            tabs_row.addWidget(btn)
        tabs_row.addStretch()
        self._op_tab_buttons["all"].setChecked(True)

        layout.addLayout(tabs_row)

        # ── Поиск по операциям — тот же визуальный приём, что и в «Участках» ──
        # (QLineEdit с нижней линией вместо рамки, без иконки). Один поисковый
        # запрос проверяется сразу по двум столбцам: Контрагент и Назначение.
        search_row = QHBoxLayout()
        search_row.setContentsMargins(0, 4, 0, 4)

        # «Выбрано: N» — слева от поиска, видим всегда (серый при 0), как
        # в «Участках».
        self._selected_lbl = QLabel("Выбрано: 0")
        self._selected_lbl.setStyleSheet(self._SELECTED_LBL_OFF_SS)
        search_row.addWidget(self._selected_lbl)

        self._search_op = QLineEdit()
        self._search_op.setPlaceholderText("Поиск по контрагенту или назначению")
        self._search_op.setClearButtonEnabled(True)
        self._search_op.setMinimumWidth(220)
        self._search_op.setContextMenuPolicy(Qt.ContextMenuPolicy.NoContextMenu)
        self._search_op.setStyleSheet(
            "QLineEdit{background:transparent;border:none;border-bottom:2px solid #D1D5DB;"
            "border-radius:0;padding:6px 2px;font-size:13px;color:#1F2937;}"
            "QLineEdit:focus{border-bottom:2px solid #07414F;}")
        self._search_op.textChanged.connect(self._on_search_text)
        search_row.addWidget(self._search_op, stretch=1)
        layout.addLayout(search_row)

        # --- дерево (Model-View) --------------------------------------- #
        self.model = OperationsTreeModel(self._manual_cells, self._dup_pending, self)
        self.model.cellEdited.connect(self._on_cell_edited)

        self.tree = QTreeView(objectName="mainTable")
        self.tree.setModel(self.model)
        self.tree.setUniformRowHeights(True)
        # Стрелка/＋/корзина живут в служебном столбце-делегате, поэтому штатную
        # «ёлочку» дерева отключаем, а отступ обнуляем — геометрия кнопок едина.
        self.tree.setRootIsDecorated(False)
        self.tree.setIndentation(0)
        self.tree.setAlternatingRowColors(True)
        self.tree.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        # Двойной клик правит ячейку прямо в таблице (Дата/Сумма/Контрагент/
        # Назначение/Категория/Участок) — в дополнение к карандашу, который
        # открывает полный диалог с разбивкой операции.
        self.tree.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked)
        self.tree.setSortingEnabled(True)
        self.tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tree.customContextMenuRequested.connect(self._show_context_menu)
        self.tree.setMouseTracking(True)
        self.tree.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.hdr_view = _DetailHeaderView()
        self.tree.setHeader(self.hdr_view)
        # setHeader() синхронизирует sectionsClickable нового заголовка с
        # tree.isSortingEnabled() (уже True) — переписывает False, который
        # _DetailHeaderView.__init__ выставил ДО подключения. Гасим ещё раз,
        # уже после setHeader(), иначе клик по всей ячейке снова сортирует.
        self.hdr_view.setSectionsClickable(False)
        self.hdr_view.setStretchLastSection(False)
        self.hdr_view.catFilterChanged.connect(self._on_hdr_cat_filter_changed)
        self.hdr_view.periodFilterChanged.connect(self._on_period_changed)
        self.hdr_view.plotFilterChanged.connect(self._on_hdr_plot_filter_changed)
        self.hdr_view.checkAllToggled.connect(self._on_master_check_clicked)
        self.tree.setStyleSheet(self._TREE_STYLE)
        self.tree.setViewportMargins(0, self.hdr_view.height(), 0, 0)

        self._cell_delegate = _CellDelegate(self.tree)
        self._category_delegate = _CategoryDelegate(ALL_CATEGORIES, self.tree)
        self._category_delegate.editCategoriesRequested.connect(self._show_cat_editor)
        self._plot_delegate = _PlotDelegate(load_plot_numbers(), self.tree)
        self._branch_delegate = _BranchColumnDelegate(self.tree)
        self._branch_delegate.toggleRequested.connect(self._on_toggle)
        self._edit_delegate = _DetailEditDelegate(self.tree)
        self._check_delegate = _DetailCheckDelegate(self.tree)
        self._check_delegate.selectionChanged.connect(self._on_check_selection_changed)
        self.tree.setItemDelegate(self._cell_delegate)
        self.tree.clicked.connect(self._on_tree_clicked)
        self.tree.expanded.connect(
            lambda idx: self.tree.viewport().update(self.tree.visualRect(idx))
        )
        self.tree.collapsed.connect(
            lambda idx: self.tree.viewport().update(self.tree.visualRect(idx))
        )

        # Панель редактора категорий — скрыта по умолчанию, выезжает справа.
        self._cat_editor_panel = CategoryEditorPanel(ALL_CATEGORIES, self)
        self._cat_editor_panel.categoriesChanged.connect(self._on_categories_changed)
        self._cat_editor_panel.categoryRenamed.connect(self._on_category_renamed)
        self._cat_editor_overlay = None
        self._cat_editor_panel.finished.connect(self._hide_cat_editor_overlay)

        table_outer = _ClipFrame(QColor("#D5DCE4"), 6)
        outer_lay = QVBoxLayout(table_outer)
        outer_lay.setContentsMargins(0, 0, 0, 0)
        outer_lay.setSpacing(0)
        outer_lay.addWidget(self.tree, stretch=1)
        table_outer.finish_setup()
        layout.addWidget(table_outer, stretch=1)

        summary_layout = QHBoxLayout()
        self.lbl_records = QLabel("Записей: —", objectName="summaryRecords")
        summary_layout.addWidget(self.lbl_records)
        summary_layout.addStretch()
        layout.addLayout(summary_layout)

    # QTreeView нуждается в собственных правилах: глобальный QSS целит в
    # QTableWidget#mainTable и на дерево не распространяется.
    # Единая копия стиля — ui.theme.tree_qss() (через ui.common.TREE_STYLE).
    _TREE_STYLE = _TREE_STYLE_COMMON

    # ------------------------------------------ редактор категорий ---------- #
    def _show_cat_editor(self):
        self._cat_editor_panel.set_categories(list(ALL_CATEGORIES))
        panel = self._cat_editor_panel
        panel.adjustSize()
        self._cat_editor_overlay = _BasePromptDialog._show_centered(panel, self)
        panel.show()
        panel.raise_()
        panel.activateWindow()

    def _hide_cat_editor_overlay(self):
        if self._cat_editor_overlay is not None:
            self._cat_editor_overlay.hide()
            self._cat_editor_overlay.deleteLater()
            self._cat_editor_overlay = None

    def _on_categories_changed(self, new_cats: list[str]):
        import ui.categorization as _cat_mod
        _cat_mod.ALL_CATEGORIES[:] = new_cats
        self._category_delegate._items = list(new_cats)

        valid_sel = self._hdr_cat_filter & set(new_cats)
        self._hdr_cat_filter = valid_sel
        if self._cat_col is not None and self._cat_col >= 0:
            self.hdr_view.set_cat_col(self._cat_col, self._present_categories(), valid_sel)

        self.apply_filters()

    def _on_category_renamed(self, old_name: str, new_name: str):
        import ui.categorization as _cat_mod
        # Переименовываем во всех строках df_full
        if self.df_full is not None and "Категория" in self.df_full.columns:
            self.df_full.loc[self.df_full["Категория"] == old_name, "Категория"] = new_name
            # ...и внутри разбивок (_breakdown) — иначе подстроки останутся
            # со старым именем категории: не попадут в дашборд/долг под новым
            # именем (расчёты матчат по точному совпадению строки), и не
            # найдутся в выпадающем списке при повторном открытии разбивки.
            if "_breakdown" in self.df_full.columns:
                def _rename_in_breakdown(v):
                    items = _parse_breakdown(v)
                    if not items:
                        return v
                    changed = False
                    for item in items:
                        if isinstance(item, dict) and item.get("Категория") == old_name:
                            item["Категория"] = new_name
                            changed = True
                    return _dump_breakdown(items) if changed else v
                self.df_full["_breakdown"] = self.df_full["_breakdown"].apply(_rename_in_breakdown)
        # Делегат уже обновлён через rename_user_category → ALL_CATEGORIES
        self._category_delegate._items = list(_cat_mod.ALL_CATEGORIES)
        if old_name in self._hdr_cat_filter:
            self._hdr_cat_filter.discard(old_name)
            self._hdr_cat_filter.add(new_name)
        if self._cat_col is not None and self._cat_col >= 0:
            self.hdr_view.set_cat_col(
                self._cat_col, self._present_categories(), self._hdr_cat_filter
            )
        self.apply_filters()

    # --------------------------------------------------- наполнение модели -- #
    # Желаемый порядок видимых столбцов по умолчанию.
    _COLUMN_ORDER = ["Дата", "Контрагент", "Назначение", "Сумма", "Категория", "Участок"]

    def _rebuild_model(self, df: "pd.DataFrame"):
        visible = [c for c in df.columns if not str(c).startswith("_")]
        ordered = [c for c in self._COLUMN_ORDER if c in visible]
        ordered += [c for c in visible if c not in self._COLUMN_ORDER]
        columns = ordered
        # Без df.iterrows(): создание Series на каждую строку заметно
        # замедляло перестроение таблицы на больших выписках.
        bd_values = (list(df["_breakdown"]) if "_breakdown" in df.columns
                     else [None] * len(df))
        records = [
            (df_idx, data, _parse_breakdown(raw_bd))
            for df_idx, data, raw_bd in zip(
                df.index, df[columns].to_dict("records"), bd_values)
        ]

        self.model.load(columns, records)
        self._apply_header_layout(self.model.columns())

        # Делегат комбобокса — только на колонку «Категория».
        cat = self.model.category_column()
        if self._cat_col is not None and self._cat_col != cat:
            self.tree.setItemDelegateForColumn(self._cat_col, self._cell_delegate)
        if cat >= 0:
            self.tree.setItemDelegateForColumn(cat, self._category_delegate)
        self._cat_col = cat

        # Делегат с линиями дерева и корзиной — на колонку «Контрагент».
        cols = self.model.columns()
        cont = cols.index("Контрагент") if "Контрагент" in cols else -1
        if self._cont_col is not None and self._cont_col != cont:
            self.tree.setItemDelegateForColumn(self._cont_col, self._cell_delegate)
        if cont >= 0:
            self.tree.setItemDelegateForColumn(cont, self._branch_delegate)
        self._cont_col = cont

        # Делегат выпадающего списка — на колонку «Участок».
        plot = cols.index("Участок") if "Участок" in cols else -1
        if self._plot_col is not None and self._plot_col != plot:
            self.tree.setItemDelegateForColumn(self._plot_col, self._cell_delegate)
        if plot >= 0:
            self.tree.setItemDelegateForColumn(plot, self._plot_delegate)
        self._plot_col = plot

        self._refresh_summary()

    def _present_categories(self) -> list[str]:
        """Категории для попапа фильтра — только реально встречающиеся в
        текущих данных (+ NO_CATEGORY_LABEL, если такие ячейки есть), а не
        весь ALL_CATEGORIES: там оседают категории, которые когда-то
        добавили в редакторе, но ни разу не присвоили ни одной операции."""
        if self.df_full is None or "Категория" not in self.df_full.columns:
            return []
        present: set = set(self.df_full["Категория"].dropna().unique())
        present.discard("")
        if "_breakdown" in self.df_full.columns:
            bd_col = self.df_full["_breakdown"]
            for raw in bd_col[bd_col.notna()]:
                for item in _parse_breakdown(raw):
                    cat = item.get("Категория") if isinstance(item, dict) else None
                    if cat:
                        present.add(cat)
        ordered = [c for c in ALL_CATEGORIES if c in present]
        extra = sorted(present - set(ordered) - {NO_CATEGORY_LABEL})
        result = ordered + extra
        if NO_CATEGORY_LABEL in present:
            result.append(NO_CATEGORY_LABEL)
        return result

    def _present_plots(self) -> list[str]:
        """Номера участков для попапа фильтра — только реально встречающиеся
        в текущих данных (не весь реестр snt_plots.json), отсортированные
        численно там, где номер — число."""
        if self.df_full is None or "Участок" not in self.df_full.columns:
            return []
        present: set = set(self.df_full["Участок"].dropna().astype(str).unique())
        present.discard("")
        if "_breakdown" in self.df_full.columns:
            bd_col = self.df_full["_breakdown"]
            for raw in bd_col[bd_col.notna()]:
                for item in _parse_breakdown(raw):
                    num = item.get("Участок") if isinstance(item, dict) else None
                    if num:
                        present.add(str(num))
        return sorted(present, key=lambda s: (0, int(s), s) if s.isdigit() else (1, 0, s))

    def _apply_header_layout(self, columns: list[str]):
        col_widths = {
            _CHECK_COL: 36, _EDIT_COL: 46,
            "Дата": 95, "Сумма": 140, "Категория": 210, "Участок": 120,
        }
        stretch_cols = {"Контрагент", "Назначение"}
        header = self.tree.header()
        header.setStretchLastSection(False)
        for col_idx, col in enumerate(columns):
            if col in stretch_cols:
                header.setSectionResizeMode(col_idx, QHeaderView.ResizeMode.Stretch)
            elif col in col_widths:
                header.setSectionResizeMode(col_idx, QHeaderView.ResizeMode.Fixed)
                self.tree.setColumnWidth(col_idx, col_widths[col])
            else:
                header.setSectionResizeMode(col_idx, QHeaderView.ResizeMode.ResizeToContents)
        # Делегат чекбокса
        if _CHECK_COL in columns:
            check_idx = columns.index(_CHECK_COL)
            self.tree.setItemDelegateForColumn(check_idx, self._check_delegate)
            self.hdr_view.set_check_col(check_idx)
        # Делегат редактирования
        if _EDIT_COL in columns:
            self.tree.setItemDelegateForColumn(columns.index(_EDIT_COL), self._edit_delegate)
        if "Категория" in columns:
            self.hdr_view.set_cat_col(
                columns.index("Категория"), self._present_categories(), self._hdr_cat_filter
            )
        if "Дата" in columns:
            self.hdr_view.set_period_col(
                columns.index("Дата"), self._period_from, self._period_to
            )
        if "Участок" in columns:
            self.hdr_view.set_plot_col(
                columns.index("Участок"), self._present_plots(), self._hdr_plot_filter
            )

    def _refresh_summary(self):
        nodes = self.model.top_nodes()
        self.lbl_records.setText(f"Записей: {len(nodes)}")
        self._refresh_check_header()

    def _refresh_check_header(self):
        total = len(self.model.top_nodes())
        selected = len(self._check_delegate.get_selected())
        self.hdr_view.set_check_state(total, selected)

    # --------------------------------------------------------- разбивка --- #
    def _set_breakdown(self, df_idx, items: list):
        if self.df_full is None:
            return
        if "_breakdown" not in self.df_full.columns:
            self.df_full["_breakdown"] = pd.Series(dtype="object", index=self.df_full.index)
        elif self.df_full["_breakdown"].dtype != "object":
            self.df_full["_breakdown"] = self.df_full["_breakdown"].astype("object")
        self.df_full.at[df_idx, "_breakdown"] = _dump_breakdown(items)

    def _delete_split(self, split_node: _Node):
        parent = split_node.parent
        if parent is None or self.df_full is None or parent.df_idx not in self.df_full.index:
            return
        items = [dict(c.data) for c in parent.children if c is not split_node]
        self._set_breakdown(parent.df_idx, items)
        target = parent.df_idx
        self.apply_filters()
        idx = self.model.index_for_df_idx(target)
        if idx.isValid() and items:
            self.tree.expand(idx)
        # Разбивка — аннотация, на суммы/долги не влияет: dataLoaded не эмитим.

    def _on_toggle(self, index: QModelIndex):
        col0 = self.model.index(index.row(), 0, self.model.parent(index))
        if self.tree.isExpanded(col0):
            self.tree.collapse(col0)
        else:
            self.tree.expand(col0)

    # --------------------------------------------------- кнопка редакт. ----- #
    def _on_tree_clicked(self, index: QModelIndex):
        cols = self.model.columns()
        if index.isValid() and 0 <= index.column() < len(cols):
            col = cols[index.column()]
            if col == _EDIT_COL:
                node = index.internalPointer()
                if node and node.kind == "op":
                    self._edit_operation(node)
            elif col == "Категория":
                # Пилюля категории открывается уже по одному клику (обычные
                # ячейки — по двойному, см. setEditTriggers), это привычнее
                # для бейджа-переключателя. Если ячейка недоступна для
                # правки (строка с разбивкой без своей категории), model.flags()
                # молча откажет — tree.edit() ничего не сделает.
                self.tree.edit(index)

    def _edit_operation(self, node):
        breakdown = []
        if self.df_full is not None and node.df_idx in self.df_full.index:
            breakdown = _parse_breakdown(self.df_full.at[node.df_idx, "_breakdown"]
                                         if "_breakdown" in self.df_full.columns else None)
        dlg = EditOperationDialog(node.data, breakdown, self)
        if _exec_dialog(dlg, self) != QDialog.DialogCode.Accepted:
            return
        try:
            result = dlg.get_result()
        except Exception as e:
            _AlertDialog.show_alert(self, "Ошибка", f"Не удалось получить данные из формы:\n{e}")
            return
        df_idx = node.df_idx
        new_breakdown = result.pop("_breakdown", None)
        if df_idx is not None and self.df_full is not None and df_idx in self.df_full.index:
            for col, val in result.items():
                if col in self.df_full.columns:
                    cur_dtype = self.df_full[col].dtype
                    if cur_dtype.kind == "f" and isinstance(val, str) and val.strip() == "":
                        val = 0.0
                    self.df_full.at[df_idx, col] = val
                    self._manual_cells.add((df_idx, col))
            self.df_full["Дата"] = pd.to_datetime(self.df_full["Дата"], errors="coerce")
            if new_breakdown is not None:
                self._set_breakdown(df_idx, new_breakdown)
        try:
            self.apply_filters()
        except Exception as e:
            _AlertDialog.show_alert(self, "Ошибка", f"Ошибка обновления таблицы:\n{e}")
            return
        if new_breakdown:
            idx = self.model.index_for_df_idx(df_idx)
            if idx.isValid():
                self.tree.expand(idx)
        try:
            self.dataLoaded.emit(self.df_full)
        except Exception as e:
            _AlertDialog.show_alert(self, "Ошибка", f"Ошибка обновления вкладок:\n{e}")

    # ----------------------------------------------------- добавить строку -- #
    def _add_row(self):
        dlg = AddRowDialog(self)
        if _exec_dialog(dlg, self) != QDialog.DialogCode.Accepted:
            return
        try:
            row_data = dlg.get_result()
        except Exception as e:
            _AlertDialog.show_alert(self, "Ошибка", f"Не удалось получить данные из формы:\n{e}")
            return
        new_breakdown = row_data.pop("_breakdown", None)

        if self.df_full is None:
            self.df_full = pd.DataFrame(
                columns=["Дата", "Контрагент", "Сумма", "Назначение", "Категория", "Участок"]
            )
            self.df_full = self.df_full.astype({"Дата": "datetime64[ns]", "Сумма": float})

        new_idx = int(self.df_full.index.max()) + 1 if len(self.df_full) > 0 else 0

        new_hash = _compute_hash(row_data)
        row_data["_hash"] = new_hash

        new_row = {
            col: row_data.get(col, "" if col != "Сумма" else float("nan"))
            for col in self.df_full.columns
        }
        self.df_full.loc[new_idx] = new_row
        self.df_full["Дата"] = pd.to_datetime(self.df_full["Дата"], errors="coerce")

        self._manual_rows.add(new_idx)
        for col in row_data:
            if col in self.df_full.columns:
                self._manual_cells.add((new_idx, col))

        if new_breakdown:
            self._set_breakdown(new_idx, new_breakdown)

        self.apply_filters()
        if new_breakdown:
            idx = self.model.index_for_df_idx(new_idx)
            if idx.isValid():
                self.tree.expand(idx)
        try:
            self.dataLoaded.emit(self.df_full)
        except Exception as e:
            _AlertDialog.show_alert(self, "Ошибка", f"Ошибка обновления вкладок:\n{e}")

    # --------------------------------------------------------- загрузка --- #
    def load_file(self):
        settings_dlg = LoadSettingsDialog(self, has_existing_data=self.df_full is not None)
        if _exec_dialog(settings_dlg, self) != QDialog.DialogCode.Accepted:
            return

        fmt        = settings_dlg.fmt
        auto_cat   = settings_dlg.auto_cat
        auto_plot  = settings_dlg.auto_plot
        merge_mode = settings_dlg.merge_mode

        path, _ = QFileDialog.getOpenFileName(
            self, "Открыть файл выписки", "", "Excel файлы (*.xlsx *.xls)")
        if not path:
            return
        # Импорт большого файла ощутимо долгий (парсинг Excel + автокатегоризация
        # + пересчёт) — показываем курсор занятости, чтобы окно не выглядело
        # зависшим (U1 из аудита UI).
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        try:
            df = pd.read_excel(path, engine="openpyxl")
            cols = [c for c in df.columns
                    if not str(c).strip().startswith("Валюта") and str(c).strip() != ""]
            df = df[cols]

            if fmt == "sber":
                drop_cols = {"Номер", "Номер счёта", "Контрагент счёт", "Контрагент cчёт"}
                df.drop(columns=[c for c in drop_cols if c in df.columns], inplace=True)

            df = _merge_to_summa(df)

            df["Дата"] = pd.to_datetime(df["Дата"], dayfirst=True, errors="coerce")
            df = df[df["Дата"].notna()].copy()

            if auto_cat:
                df = apply_categorization(df)
            if auto_plot:
                df = apply_plot_column(df)

            if "Категория" not in df.columns:
                df["Категория"] = ""
            if "Участок" not in df.columns:
                df["Участок"] = ""

            df["Участок"] = df["Участок"].apply(
                lambda v: "" if pd.isna(v) else
                str(int(v)) if isinstance(v, float) and v == int(v) else str(v)
            )

            df = _ensure_meta_columns(df)

            if merge_mode and self.df_full is not None:
                existing = _ensure_meta_columns(self.df_full)
                hash_to_idx: dict[str, int] = {}
                for idx, h in existing["_hash"].items():
                    if h:
                        hash_to_idx.setdefault(h, idx)

                new_rows = []
                dup_count = 0
                for _, row in df.iterrows():
                    match_idx = hash_to_idx.get(row["_hash"])
                    if match_idx is not None:
                        dt = row.get("Дата")
                        self._dup_pending[match_idx] = {
                            "Дата": dt.isoformat() if pd.notna(dt) else None,
                            "Сумма": _to_num(row.get("Сумма")),
                            "Контрагент": str(row.get("Контрагент") or ""),
                            "Назначение": str(row.get("Назначение") or ""),
                        }
                        dup_count += 1
                    else:
                        new_rows.append(row)

                if new_rows:
                    new_df = pd.DataFrame(new_rows).reset_index(drop=True)
                    new_start = int(existing.index.max()) + 1 if len(existing) > 0 else 0
                    new_df.index = new_df.index + new_start
                    self.df_full = pd.concat([existing, new_df])
                else:
                    self.df_full = existing

                if dup_count:
                    QApplication.restoreOverrideCursor()
                    _AlertDialog.show_alert(
                        self, "Импорт завершён",
                        f"Добавлено новых операций: {len(new_rows)}.\n"
                        f"Обнаружено повторов уже загруженных операций: {dup_count} "
                        f"— они не задвоены, а помечены в списке (жёлтым цветом; "
                        f"правой кнопкой мыши по строке — восстановить исходные "
                        f"данные повтора или оставить как есть)."
                    )
                    QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
            else:
                self._manual_rows.clear()
                self._manual_cells.clear()
                self._dup_pending.clear()
                self.df_full = df

            self.apply_filters()
            self.dataLoaded.emit(self.df_full)
        except Exception as e:
            QApplication.restoreOverrideCursor()
            _AlertDialog.show_alert(self, "Ошибка загрузки", f"Не удалось загрузить файл:\n{e}")
        finally:
            # Лишний restore при уже пустом стеке курсоров — no-op.
            QApplication.restoreOverrideCursor()

    def load_dataframe(self, df: "pd.DataFrame"):
        """Восстанавливает DataFrame из сохранённого проекта без диалога выбора файла."""
        self._manual_rows.clear()
        self._manual_cells.clear()
        self._dup_pending.clear()
        drop_cols = {"Номер", "Номер счёта", "Контрагент счёт", "Контрагент cчёт", "Теги"}
        df = df.drop(columns=[c for c in drop_cols if c in df.columns])
        df = _merge_to_summa(df)
        if "Категория" in df.columns:
            df["Категория"] = df["Категория"].apply(
                lambda v: v[:-len(" (авто)")] if isinstance(v, str) and v.endswith(" (авто)") else v
            )
        df = _ensure_meta_columns(df)
        self.df_full = df
        self.apply_filters()
        try:
            self.dataLoaded.emit(self.df_full)
        except Exception as e:
            _AlertDialog.show_alert(self, "Ошибка", f"Ошибка восстановления данных:\n{e}")

    def get_manual_cells_data(self) -> list:
        """Сериализует _manual_cells для сохранения в проект."""
        return [[int(df_idx), col] for df_idx, col in self._manual_cells]

    def restore_manual_cells(self, data: list):
        """Восстанавливает _manual_cells и _manual_rows после загрузки проекта."""
        self._manual_cells.clear()
        self._manual_rows.clear()
        for item in data:
            if len(item) == 2:
                df_idx, col = int(item[0]), str(item[1])
                self._manual_cells.add((df_idx, col))
                self._manual_rows.add(df_idx)
        self.tree.viewport().update()

    def refresh_plot_column(self):
        """Пересчитывает столбец «Участок» по актуальным данным из snt_plots.json.
        Строки, вручную отредактированные пользователем, не перезаписываются."""
        # Обновляем список в делегате — вдруг добавились новые участки
        self._plot_delegate._items = load_plot_numbers()
        if self.df_full is None:
            return
        df_new = apply_plot_column(self.df_full)
        if "Участок" not in self.df_full.columns:
            self.df_full = df_new
        else:
            auto_mask = ~self.df_full.index.isin(self._manual_rows)
            self.df_full.loc[auto_mask, "Участок"] = df_new.loc[auto_mask, "Участок"]
        self.apply_filters()

    # ----------------------------------------------------------- фильтры -- #
    def _filtered_df(self) -> "pd.DataFrame":
        df = self.df_full.copy()

        if self._hdr_cat_filter:
            _sel = self._hdr_cat_filter
            # Строка с разбивкой фильтруется по категориям её строк, без —
            # по верхнеуровневой. JSON разбивки парсим только там, где он
            # есть (df.apply по всем строкам был излишне медленным).
            mask = df["Категория"].isin(_sel)
            if "_breakdown" in df.columns:
                bd_col = df["_breakdown"]
                for idx in df.index[bd_col.notna()]:
                    bd = _parse_breakdown(bd_col.at[idx])
                    if bd:
                        mask.at[idx] = any(it.get("Категория") in _sel for it in bd)
            df = df[mask]

        if self._hdr_plot_filter:
            _psel = self._hdr_plot_filter
            mask = df["Участок"].astype(str).isin(_psel)
            if "_breakdown" in df.columns:
                bd_col = df["_breakdown"]
                for idx in df.index[bd_col.notna()]:
                    bd = _parse_breakdown(bd_col.at[idx])
                    if bd:
                        mask.at[idx] = any(str(it.get("Участок")) in _psel for it in bd)
            df = df[mask]

        # Поиск по операциям — сразу по двум столбцам: Контрагент и Назначение.
        if self._search_text:
            text = self._search_text
            mask = df["Контрагент"].astype(str).str.lower().str.contains(text, na=False)
            if "Назначение" in df.columns:
                mask |= df["Назначение"].astype(str).str.lower().str.contains(text, na=False)
            df = df[mask]

        if self._period_from is not None:
            df = df[df["Дата"].dt.date >= self._period_from]
        if self._period_to is not None:
            df = df[df["Дата"].dt.date <= self._period_to]

        self._refresh_op_tabs(df)
        if self._filter_mode == "income":
            df = df[df["Сумма"] > 0]
        elif self._filter_mode == "expense":
            df = df[df["Сумма"] < 0]

        return df

    def _refresh_op_tabs(self, base: "pd.DataFrame"):
        amounts = base["Сумма"]
        counts = {
            "all":     len(base),
            "income":  int((amounts > 0).sum()),
            "expense": int((amounts < 0).sum()),
        }
        for mode, btn in self._op_tab_buttons.items():
            btn.set_count(counts[mode])

    def apply_filters(self):
        if self.df_full is None:
            return
        self._rebuild_model(self._filtered_df())

    def _on_op_filter_tab(self, mode: str):
        self._filter_mode = mode
        self.apply_filters()

    def _on_period_changed(self, date_from, date_to):
        self._period_from = date_from
        self._period_to = date_to
        self.apply_filters()

    def _on_hdr_cat_filter_changed(self, selected: set):
        self._hdr_cat_filter = set(selected)
        self.apply_filters()

    def _on_hdr_plot_filter_changed(self, selected: set):
        self._hdr_plot_filter = set(selected)
        self.apply_filters()

    def _on_search_text(self, text: str):
        self._search_text = text.strip().lower()
        self.apply_filters()

    def _on_check_selection_changed(self):
        n = len(self._check_delegate.get_selected())
        self._btn_bulk_delete.setEnabled(n > 0)
        self._selected_lbl.setText(f"Выбрано: {n}")
        self._selected_lbl.setStyleSheet(
            self._SELECTED_LBL_ON_SS if n else self._SELECTED_LBL_OFF_SS)
        self._refresh_check_header()

    def _on_master_check_clicked(self):
        """Клик по мастер-чекбоксу в заголовке — как в «Участках»:
        снимает выбор, если выбрано ВСЁ видимое, иначе выбирает всё видимое."""
        nodes = self.model.top_nodes()
        selected = self._check_delegate.get_selected()
        if nodes and len(selected) >= len(nodes):
            self._check_delegate.clear_selection()
        else:
            self._check_delegate.select_all(n.df_idx for n in nodes)

    def _delete_selected(self):
        indices = self._check_delegate.get_selected()
        if not indices:
            return
        confirmed = _ConfirmDialog.confirm(
            self, "Удаление строк",
            f"Удалить {len(indices)} выбранных строк из детализации?",
            confirm_text="Да, удалить", cancel_text="Нет",
        )
        if not confirmed:
            return
        try:
            self.df_full = self.df_full.drop(list(indices)).reset_index(drop=True)
            self._check_delegate.clear_selection()
            self.apply_filters()
            self.dataLoaded.emit(self.df_full)
        except Exception as e:
            _AlertDialog.show_alert(self, "Ошибка", f"Ошибка удаления строк:\n{e}")

    # ------------------------------------------------------------ экспорт -- #
    def _export_excel(self):
        if self.df_full is None:
            _AlertDialog.show_alert(self, "Нет данных", "Сначала загрузите файл выписки.")
            return

        df = self._filtered_df()
        if df.empty:
            _AlertDialog.show_alert(self, "Нет данных", "После применения фильтров нет строк для экспорта.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Экспорт детализации", "детализация.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        if not path.endswith(".xlsx"):
            path += ".xlsx"

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Детализация"

            headers = [c for c in df.columns if not str(c).startswith("_")]
            summa_col_idx = headers.index("Сумма") + 1 if "Сумма" in headers else None

            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for _, row in df.iterrows():
                out_row = []
                for col in headers:
                    val = row[col]
                    if col == "Сумма":
                        num = pd.to_numeric(val, errors="coerce")
                        out_row.append(float(num) if pd.notna(num) else "")
                    elif col == "Дата":
                        out_row.append(val.strftime("%d.%m.%Y") if pd.notna(val) else "")
                    else:
                        out_row.append("" if pd.isna(val) else val)
                ws.append(out_row)

            if summa_col_idx is not None:
                green_fill = PatternFill("solid", fgColor="C8E6C9")
                red_fill   = PatternFill("solid", fgColor="FFCDD2")
                summa_letter = get_column_letter(summa_col_idx)
                for r in range(2, ws.max_row + 1):
                    cell = ws[f"{summa_letter}{r}"]
                    if isinstance(cell.value, (int, float)):
                        cell.fill  = green_fill if cell.value >= 0 else red_fill
                        cell.number_format = '#,##0.00 ₽'
                        cell.alignment = Alignment(horizontal="right")

            for col_cells in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

            wb.save(path)
            _AlertDialog.show_alert(self, "Экспорт завершён", f"Файл сохранён:\n{path}")
        except Exception as e:
            _AlertDialog.show_alert(self, "Ошибка экспорта", str(e))

    # ------------------------------------------------------ редактирование -- #
    def _on_cell_edited(self, node: _Node, col: str):
        """Слот модели: переносит правки ячейки обратно в df_full."""
        try:
            self._on_cell_edited_impl(node, col)
        except Exception as e:
            _AlertDialog.show_alert(self, "Ошибка", f"Ошибка сохранения ячейки:\n{e}")

    def _on_cell_edited_impl(self, node: _Node, col: str):
        if self.df_full is None:
            return

        if node.kind == "op":
            df_idx = node.df_idx
            if df_idx is not None and df_idx in self.df_full.index:
                if col == "Сумма":
                    self.df_full.at[df_idx, "Сумма"] = node.data["Сумма"]
                elif col == "Дата":
                    self.df_full.at[df_idx, "Дата"] = node.data["Дата"]
                else:
                    self.df_full.at[df_idx, col] = node.data[col]
                    if col == "Категория" and node.data[col]:
                        if node.data[col] not in self._category_delegate._items:
                            self._category_delegate._items.append(node.data[col])
                    self._manual_rows.add(df_idx)
                    self._manual_cells.add((df_idx, col))
        else:  # split — сохраняем разбивку родителя
            parent = node.parent
            if parent is not None and parent.df_idx in self.df_full.index:
                items = [dict(c.data) for c in parent.children]
                self._set_breakdown(parent.df_idx, items)

        # ВАЖНО: НЕ эмитим dataLoaded на каждую правку ячейки — на этом сигнале
        # висят тяжёлые пересчёты долгов (взносы/электроэнергия/home), из-за чего
        # после выхода из редактирования ПК зависал на 5-6 сек. Долги
        # пересчитываются при загрузке/добавлении/удалении операций.
        self._refresh_summary()

    # -------------------------------------------------------- контекст-меню -- #
    def _show_context_menu(self, pos: QPoint):
        index = self.tree.indexAt(pos)
        if not index.isValid():
            return
        node = index.internalPointer()
        self.tree.setCurrentIndex(index)

        menu = QMenu(self)
        menu.setStyleSheet(menu_qss())

        if node.kind == "op":
            act_dup = QAction("Дублировать операцию", self)
            act_dup.triggered.connect(lambda: self._duplicate_op(node))
            menu.addAction(act_dup)
            act_del = QAction("Удалить операцию", self)
            act_del.triggered.connect(lambda: self._delete_op(node))
            menu.addAction(act_del)
            if node.df_idx in self._dup_pending:
                menu.addSeparator()
                act_restore = QAction("⟲ Восстановить данные повторного импорта", self)
                act_restore.triggered.connect(lambda: self._restore_dup_pending(node))
                menu.addAction(act_restore)
                act_dismiss = QAction("✓ Оставить как есть (это правка)", self)
                act_dismiss.triggered.connect(lambda: self._dismiss_dup_pending(node))
                menu.addAction(act_dismiss)
        else:
            act_del = QAction("Удалить распределение", self)
            act_del.triggered.connect(lambda: self._delete_split(node))
            menu.addAction(act_del)

        menu.exec(self.tree.viewport().mapToGlobal(pos))

    def _duplicate_op(self, op_node: _Node):
        if self.df_full is None or op_node.df_idx not in self.df_full.index:
            return
        try:
            new_idx = int(self.df_full.index.max()) + 1
            self.df_full.loc[new_idx] = self.df_full.loc[op_node.df_idx].copy()
            self.apply_filters()
            idx = self.model.index_for_df_idx(new_idx)
            if idx.isValid():
                self.tree.setCurrentIndex(idx)
            self.dataLoaded.emit(self.df_full)
        except Exception as e:
            _AlertDialog.show_alert(self, "Ошибка", f"Ошибка дублирования операции:\n{e}")

    def _restore_dup_pending(self, op_node: _Node):
        """Перезаписывает Дата/Сумма/Контрагент/Назначение строки значениями
        из повторного импорта (см. load_file) — откатывает правки пользователя
        к тому, что реально показывает банк."""
        df_idx = op_node.df_idx
        pending = self._dup_pending.get(df_idx)
        if self.df_full is None or pending is None or df_idx not in self.df_full.index:
            return
        try:
            if pending.get("Дата"):
                self.df_full.at[df_idx, "Дата"] = pd.Timestamp(pending["Дата"])
            self.df_full.at[df_idx, "Сумма"] = pending.get("Сумма")
            self.df_full.at[df_idx, "Контрагент"] = pending.get("Контрагент", "")
            self.df_full.at[df_idx, "Назначение"] = pending.get("Назначение", "")
            for col in ("Дата", "Сумма", "Контрагент", "Назначение"):
                self._manual_cells.discard((df_idx, col))
            del self._dup_pending[df_idx]
            self.apply_filters()
            self.dataLoaded.emit(self.df_full)
        except Exception as e:
            _AlertDialog.show_alert(self, "Ошибка", f"Не удалось восстановить данные:\n{e}")

    def _dismiss_dup_pending(self, op_node: _Node):
        """Снимает пометку «повторный импорт» — правки пользователя остаются."""
        self._dup_pending.pop(op_node.df_idx, None)
        self.apply_filters()

    def _delete_op(self, op_node: _Node):
        confirmed = _ConfirmDialog.confirm(
            self, "Удаление строки", "Удалить выбранную операцию?",
            confirm_text="Удалить", cancel_text="Отмена",
        )
        if not confirmed:
            return
        try:
            if self.df_full is not None and op_node.df_idx in self.df_full.index:
                self.df_full = self.df_full.drop(index=op_node.df_idx)
            self.apply_filters()
            self.dataLoaded.emit(self.df_full)
        except Exception as e:
            _AlertDialog.show_alert(self, "Ошибка", f"Ошибка удаления операции:\n{e}")
